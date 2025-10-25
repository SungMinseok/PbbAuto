"""
명령어 레지스트리 - 플러그인 방식으로 명령어를 관리합니다.
새로운 명령어를 추가하려면 이 파일에 클래스만 추가하면 됩니다.
"""

# 로그 설정을 가장 먼저 import (print 출력을 로그파일에도 저장)
import logger_setup

from abc import ABC, abstractmethod
from PyQt5.QtWidgets import (QWidget, QHBoxLayout, QVBoxLayout, QLabel, QLineEdit, 
                             QSpinBox, QComboBox, QPushButton, QMessageBox, QCheckBox,
                             QRadioButton, QButtonGroup, QTextEdit, QFileDialog, QDialog,
                             QScrollArea, QFrame)
from PyQt5.QtGui import QPixmap, QFont
from PyQt5.QtCore import Qt
import time
import os
import pyperclip
import pydirectinput as pyd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
try:
    from PIL import Image as PILImage
except ImportError:
    print("PIL(Pillow) 라이브러리가 필요합니다. 'pip install Pillow' 명령어로 설치해주세요.")
    PILImage = None
from constants import test_results_dir
from utils import take_screenshot, take_screenshot_with_coords, image_to_text, calculate_adjusted_coordinates, calculate_offset_coordinates
from datetime import datetime
import glob
import subprocess


def show_unified_ocr_test_dialog(
    x, y, width, height, 
    screenshot_path, 
    ocr_lang, 
    extracted_text="",
    expected_text=None, 
    exact_match=False,
    ocr_attempts=None,
    total_time=0
):
    """
    통합 OCR 테스트 결과 팝업 다이얼로그
    
    Args:
        x, y, width, height: 좌표 정보
        screenshot_path: 스크린샷 경로
        ocr_lang: OCR 언어 정보 (예: "영어", "한국어", "자동")
        extracted_text: 실제로 추출된 텍스트
        expected_text: 기대 텍스트 (선택사항)
        exact_match: 완전일치 모드 여부
        ocr_attempts: [(텍스트, 신뢰도, 정보), ...] 형태의 시도 목록
        total_time: 총 소요 시간
    """
    import tes
    
    dialog = QDialog()
    dialog.setWindowTitle("OCR 테스트 결과")
    dialog.setMinimumWidth(850)
    dialog.setMinimumHeight(700)
    
    # 메인 레이아웃을 스크롤 가능하게 만들기
    main_layout = QVBoxLayout()
    
    # 스크롤 영역 생성
    scroll_area = QScrollArea()
    scroll_area.setWidgetResizable(True)
    scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
    scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
    
    # 스크롤 영역 안의 컨텐츠 위젯
    content_widget = QWidget()
    layout = QVBoxLayout()
    content_widget.setLayout(layout)
    
    # 1. 기본 정보
    info_text = f"📍 좌표: ({x}, {y}, {width}, {height})\n"
    info_text += f"🌐 언어: {ocr_lang}\n"
    info_text += f"⏱️ 총 소요시간: {total_time:.2f}초\n"
    info_text += f"📁 스크린샷: {os.path.basename(screenshot_path)}"
    
    info_label = QLabel(info_text)
    info_label.setFont(QFont("맑은 고딕", 10))
    info_label.setWordWrap(True)
    info_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
    layout.addWidget(info_label)
    
    # 구분선
    line1 = QFrame()
    line1.setFrameShape(QFrame.HLine)
    line1.setFrameShadow(QFrame.Sunken)
    layout.addWidget(line1)
    
    # 2. 시도 정보 (모든 OCR 시도)
    attempts = ocr_attempts if ocr_attempts is not None else tes._last_ocr_attempts
    
    if attempts:
        attempts_label = QLabel("🔍 OCR 시도 내역:")
        attempts_label.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        layout.addWidget(attempts_label)
        
        attempts_text = ""
        for i, (text, conf, info) in enumerate(attempts, 1):
            preview = text[:50] + "..." if len(text) > 50 else text
            preview = preview.replace('\n', ' ')  # 줄바꿈 제거
            attempts_text += f"{i}. [{info}] 신뢰도:{conf:.1f}% → '{preview}'\n"
        
        attempts_display = QTextEdit()
        attempts_display.setPlainText(attempts_text)
        attempts_display.setReadOnly(True)
        attempts_display.setMaximumHeight(150)
        attempts_display.setFont(QFont("Consolas", 9))
        layout.addWidget(attempts_display)
    else:
        no_attempts_label = QLabel("⚠️ OCR 시도 정보가 없습니다.")
        no_attempts_label.setFont(QFont("맑은 고딕", 9))
        no_attempts_label.setStyleSheet("color: orange;")
        layout.addWidget(no_attempts_label)
    
    # 구분선
    line2 = QFrame()
    line2.setFrameShape(QFrame.HLine)
    line2.setFrameShadow(QFrame.Sunken)
    layout.addWidget(line2)
    
    # 3. 최종 OCR 결과
    result_label = QLabel("✅ 최종 OCR 결과:")
    result_label.setFont(QFont("맑은 고딕", 10, QFont.Bold))
    layout.addWidget(result_label)
    
    # 실제 추출된 텍스트 사용 (전달된 extracted_text 우선, 없으면 attempts에서 최고 신뢰도)
    display_text = extracted_text
    if not display_text and attempts:
        best_text, best_conf, best_info = max(attempts, key=lambda x: x[1])
        display_text = best_text
        result_label.setText(f"✅ 최종 OCR 결과 (최고 신뢰도: {best_conf:.1f}%):")
    
    result_text_edit = QTextEdit()
    result_text_edit.setPlainText(display_text or "(텍스트를 찾을 수 없음)")
    result_text_edit.setReadOnly(True)
    result_text_edit.setMaximumHeight(120)
    result_text_edit.setFont(QFont("맑은 고딕", 10))
    layout.addWidget(result_text_edit)
    
    # 4. 기대 텍스트 비교
    if expected_text:
        match_type = "완전일치" if exact_match else "일부포함"
        match_found = False
        
        # 실제 추출된 텍스트로 비교
        compare_text = display_text or ""
        
        if exact_match:
            match_found = compare_text.strip() == expected_text.strip()
        else:
            match_found = expected_text in compare_text
        
        if match_found:
            match_label = QLabel(f"✅ Pass: 기대 텍스트 '{expected_text}'를 발견했습니다! ({match_type})")
            match_label.setStyleSheet("color: green; font-weight: bold; background-color: #e8f5e9; padding: 8px; border-radius: 4px;")
        else:
            match_label = QLabel(f"❌ Fail: 기대 텍스트 '{expected_text}'를 찾지 못했습니다. ({match_type})")
            match_label.setStyleSheet("color: red; font-weight: bold; background-color: #ffebee; padding: 8px; border-radius: 4px;")
        
        match_label.setFont(QFont("맑은 고딕", 10))
        match_label.setWordWrap(True)
        layout.addWidget(match_label)
    
    # 구분선
    line3 = QFrame()
    line3.setFrameShape(QFrame.HLine)
    line3.setFrameShadow(QFrame.Sunken)
    layout.addWidget(line3)
    
    # 5. 스크린샷 이미지
    if os.path.exists(screenshot_path):
        img_label = QLabel("🖼️ 캡처된 이미지:")
        img_label.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        layout.addWidget(img_label)
        
        img_container = QLabel()
        pixmap = QPixmap(screenshot_path)
        
        # 이미지가 너무 크면 가로 크기에 맞춰 축소
        if pixmap.width() > 800:
            pixmap = pixmap.scaledToWidth(800, Qt.SmoothTransformation)
        
        img_container.setPixmap(pixmap)
        img_container.setAlignment(Qt.AlignCenter)
        layout.addWidget(img_container)
    
    # 스크롤 영역에 컨텐츠 위젯 설정
    scroll_area.setWidget(content_widget)
    main_layout.addWidget(scroll_area)
    
    # 닫기 버튼 (스크롤 영역 밖에 고정)
    close_btn = QPushButton("닫기")
    close_btn.setMinimumHeight(35)
    close_btn.clicked.connect(dialog.accept)
    main_layout.addWidget(close_btn)
    
    dialog.setLayout(main_layout)
    dialog.exec_()


class CommandBase(ABC):
    """명령어 기본 클래스"""
    
    def __init__(self):
        self.screenshot_path = None
        self.extracted_text = ""
        self.expected_text = ""
        self.last_result = "N/A"
        #self.checklist_file = "checklist.xlsx"
        #self.cl_path = os.path.join(cl_dir, self.checklist_file)
        self.main_app = None  # 메인 앱 참조
    
    def set_main_app(self, main_app):
        """메인 앱 참조 설정"""
        self.main_app = main_app
    
    def _update_current_window_info(self, window_info):
        """현재 실제 선택된 윈도우로 window_info 업데이트"""
        if not window_info:
            return
            
        if self.main_app and hasattr(self.main_app, 'window_dropdown'):
            current_window = self.main_app.window_dropdown.currentText()
            if current_window and current_window != window_info.get('target_app'):
                # 실제 윈도우와 저장된 정보가 다르면 업데이트
                window_info['target_app'] = current_window
                print(f"🔄 대상 윈도우 정보 실시간 업데이트: {current_window}")
    
    def create_window_info_layout(self):
        """현재 선택된 윈도우 정보를 보여주는 레이아웃 생성"""
        info_layout = QHBoxLayout()
        info_layout.addWidget(QLabel('현재 선택된 앱:'))
        
        self.window_info_dropdown = QComboBox()
        self.window_info_dropdown.setMinimumWidth(200)
        self.window_info_dropdown.setEnabled(True)
        
        # 드롭다운 변경 시 메인 앱의 dropdown도 변경
        self.window_info_dropdown.currentTextChanged.connect(self.on_window_changed)
        
        info_layout.addWidget(self.window_info_dropdown)
        
        # 새로고침 버튼
        refresh_btn = QPushButton('새로고침')
        refresh_btn.clicked.connect(self.refresh_window_info)
        refresh_btn.setMaximumWidth(80)
        info_layout.addWidget(refresh_btn)
        
        return info_layout
    
    def refresh_window_info(self):
        """윈도우 정보 새로고침"""
        if self.main_app and hasattr(self.main_app, 'window_dropdown'):
            # 메인 앱의 윈도우 목록 가져오기
            current_text = self.main_app.window_dropdown.currentText()
            
            # 드롭다운 업데이트
            self.window_info_dropdown.clear()
            for i in range(self.main_app.window_dropdown.count()):
                item_text = self.main_app.window_dropdown.itemText(i)
                self.window_info_dropdown.addItem(item_text)
            
            # 현재 선택된 항목으로 설정
            index = self.window_info_dropdown.findText(current_text)
            if index >= 0:
                self.window_info_dropdown.setCurrentIndex(index)
    
    def initialize_window_info(self):
        """윈도우 정보 초기화 (명령어 편집 창이 열릴 때 호출)"""
        if hasattr(self, 'window_info_dropdown'):
            # UI가 생성된 후에만 초기화
            self.refresh_window_info()
    
    def on_window_changed(self, text):
        """윈도우 선택 변경 시 메인 앱의 dropdown도 변경"""
        if self.main_app and hasattr(self.main_app, 'window_dropdown'):
            index = self.main_app.window_dropdown.findText(text)
            if index >= 0:
                self.main_app.window_dropdown.setCurrentIndex(index)
    
    @property
    @abstractmethod
    def name(self) -> str:
        """명령어 이름"""
        pass
    
    @property
    @abstractmethod  
    def description(self) -> str:
        """명령어 설명"""
        pass
    
    @abstractmethod
    def create_ui(self) -> QWidget:
        """UI 위젯 생성"""
        pass
    
    def create_ui_with_window_info(self):
        """윈도우 정보가 포함된 UI 생성 (모든 명령어에서 사용)"""
        # 실제 UI 생성
        main_widget = self.create_ui()
        
        # 윈도우 정보를 맨 위에 추가
        container = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        
        # 윈도우 정보 레이아웃 추가
        window_info_layout = self.create_window_info_layout()
        layout.addLayout(window_info_layout)
        
        # 원래 UI 추가
        layout.addWidget(main_widget)
        
        container.setLayout(layout)
        return container
    
    @abstractmethod
    def parse_params(self, params: list) -> dict:
        """파라미터 파싱"""
        pass
    
    @abstractmethod
    def set_ui_values(self, params: dict):
        """UI에 값 설정"""
        pass
    
    @abstractmethod
    def get_command_string(self) -> str:
        """명령어 문자열 생성"""
        pass
    
    @abstractmethod
    def execute(self, params: dict, window_coords=None, processor_state=None):
        """명령어 실행"""
        pass


class PressCommand(CommandBase):
    """키 입력 명령어"""
    
    @property
    def name(self) -> str:
        return "Press"
    
    @property
    def description(self) -> str:
        return "Press keyboard key(s)"
    
    def _interruptible_sleep(self, duration, params):
        """중지 플래그를 체크하면서 대기하는 함수
        
        Args:
            duration: 대기 시간 (초)
            params: 파라미터 딕셔너리 (processor_stop_flag 체크용)
        
        Returns:
            bool: True if interrupted (중지됨), False if completed (완료됨)
        """
        if duration <= 0:
            if duration > 0:
                time.sleep(duration)
            return False
            
        # 0.1초 간격으로 중지 플래그 체크
        check_interval = 0.1
        total_slept = 0
        
        while total_slept < duration:
            # params에서 CommandProcessor의 실시간 stop_flag 체크
            processor = params.get('processor') if params else None
            if processor and hasattr(processor, 'stop_flag') and processor.stop_flag:
                print(f"⚠️ Hold 중지됨 (경과시간: {total_slept:.1f}초/{duration}초)")
                return True  # 중지됨
                
            sleep_time = min(check_interval, duration - total_slept)
            time.sleep(sleep_time)
            total_slept += sleep_time
            
        return False  # 완료됨
    
    def create_ui(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Keys input row
        keys_row = QHBoxLayout()
        self.press_input = QLineEdit()
        self.press_input.setPlaceholderText('key or keys (e.g. ctrl f, esc, etc.)')
        keys_row.addWidget(QLabel('Keys:'))
        keys_row.addWidget(self.press_input)
        layout.addLayout(keys_row)
        
        # Hold duration row
        hold_row = QHBoxLayout()
        self.hold_input = QLineEdit()
        self.hold_input.setPlaceholderText('0')
        self.hold_input.setText('0')
        hold_row.addWidget(QLabel('Hold (초):'))
        hold_row.addWidget(self.hold_input)
        hold_row.addStretch()
        layout.addLayout(hold_row)
        
        widget.setLayout(layout)
        return widget
    
    def parse_params(self, params: list) -> dict:
        result = {'keys': '', 'hold': 0}
        if len(params) >= 1:
            # 마지막 파라미터가 숫자면 hold 시간으로 사용
            try:
                hold_time = float(params[-1])
                result['hold'] = hold_time
                result['keys'] = ' '.join(params[:-1])
            except ValueError:
                # 숫자가 아니면 모든 파라미터를 키로 사용
                result['keys'] = ' '.join(params)
        return result
    
    def set_ui_values(self, params: dict):
        self.press_input.setText(params.get('keys', ''))
        self.hold_input.setText(str(params.get('hold', 0)))
    
    def get_command_string(self) -> str:
        keys = self.press_input.text().strip()
        hold = self.hold_input.text().strip()
        
        if keys and hold and hold != '0':
            return f"press {keys} {hold}"
        elif keys:
            return f"press {keys}"
        else:
            return 'press'
    
    def execute(self, params: dict, window_coords=None, processor_state=None):
        keys_str = params.get('keys', '')
        hold_time = float(params.get('hold', 0))
        keys = [part.strip() for part in keys_str.split()]
        
        if hold_time > 0:
            # Hold 모드: 키를 누르고 있다가 뗌
            if len(keys) == 1:
                pyd.keyDown(keys[0])
                print(f'Key down: {keys[0]} (holding for {hold_time}초)')
                
                # 중지 플래그를 체크하면서 대기
                if self._interruptible_sleep(hold_time, params):
                    pyd.keyUp(keys[0])
                    print(f'Key up: {keys[0]} (중지됨)')
                    return
                    
                pyd.keyUp(keys[0])
                print(f'Key up: {keys[0]}')
            elif len(keys) >= 2:
                pyd.keyDown(keys[0])
                pyd.keyDown(keys[1])
                print(f'Keys down: {keys[0]}+{keys[1]} (holding for {hold_time}초)')
                
                # 중지 플래그를 체크하면서 대기
                if self._interruptible_sleep(hold_time, params):
                    pyd.keyUp(keys[1])
                    pyd.keyUp(keys[0])
                    print(f'Keys up: {keys[0]}+{keys[1]} (중지됨)')
                    return
                    
                pyd.keyUp(keys[1])
                pyd.keyUp(keys[0])
                print(f'Keys up: {keys[0]}+{keys[1]}')
        else:
            # 기본 모드: 즉시 눌렀다가 뗌
            if len(keys) == 1:
                pyd.press(keys[0])
                print(f'Pressed key: {keys[0]}')
            elif len(keys) >= 2:
                pyd.keyDown(keys[0])
                pyd.press(keys[1])
                pyd.keyUp(keys[0])
                print(f'Pressed: {keys[0]}+{keys[1]}')


class WriteCommand(CommandBase):
    """텍스트 입력 명령어"""
    
    @property
    def name(self) -> str:
        return "Write"
    
    @property
    def description(self) -> str:
        return "Type text"
    
    def create_ui(self) -> QWidget:
        widget = QWidget()
        main_layout = QVBoxLayout()
        
        # 기본 텍스트 입력 행
        text_row = QHBoxLayout()
        self.write_input = QLineEdit()
        self.write_input.setPlaceholderText('Text to type')
        text_row.addWidget(QLabel('Text:'))
        text_row.addWidget(self.write_input)
        main_layout.addLayout(text_row)
        
        # 텍스트 파일 읽기 옵션
        self.use_file_checkbox = QCheckBox('텍스트 파일 읽기')
        self.use_file_checkbox.toggled.connect(self._toggle_file_options)
        main_layout.addWidget(self.use_file_checkbox)
        
        # 파일 관련 위젯들을 담을 컨테이너
        self.file_container = QWidget()
        file_layout = QVBoxLayout()
        file_layout.setContentsMargins(20, 0, 0, 0)
        
        # 파일 경로 선택
        file_path_row = QHBoxLayout()
        file_path_row.addWidget(QLabel('파일 경로:'))
        self.file_path_input = QLineEdit()
        self.file_path_input.setPlaceholderText('파일을 드래그앤드롭 하거나 버튼으로 선택')
        self.file_path_input.setAcceptDrops(True)
        self.file_path_input.dragEnterEvent = self._drag_enter_event
        self.file_path_input.dropEvent = self._drop_event
        file_path_row.addWidget(self.file_path_input)
        self.browse_button = QPushButton('찾아보기')
        self.browse_button.clicked.connect(self._browse_file)
        file_path_row.addWidget(self.browse_button)
        file_layout.addLayout(file_path_row)
        
        # 앞 스트링
        prefix_row = QHBoxLayout()
        prefix_row.addWidget(QLabel('앞 스트링:'))
        self.prefix_input = QLineEdit()
        self.prefix_input.setPlaceholderText('파일 내용 앞에 추가할 텍스트')
        self.prefix_input.textChanged.connect(self._update_preview)
        prefix_row.addWidget(self.prefix_input)
        file_layout.addLayout(prefix_row)
        
        # 뒷 스트링
        suffix_row = QHBoxLayout()
        suffix_row.addWidget(QLabel('뒷 스트링:'))
        self.suffix_input = QLineEdit()
        self.suffix_input.setPlaceholderText('파일 내용 뒤에 추가할 텍스트')
        self.suffix_input.textChanged.connect(self._update_preview)
        suffix_row.addWidget(self.suffix_input)
        file_layout.addLayout(suffix_row)
        
        # 입력 모드 선택
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel('입력 모드:'))
        self.file_mode_group = QButtonGroup()
        self.file_mode_all = QRadioButton('한번에 입력 (모든 줄을 한번에)')
        self.file_mode_iter = QRadioButton('반복마다 입력 (반복 n번째에 n번째 줄)')
        self.file_mode_all.setChecked(True)
        self.file_mode_group.addButton(self.file_mode_all)
        self.file_mode_group.addButton(self.file_mode_iter)
        mode_row.addWidget(self.file_mode_all)
        mode_row.addWidget(self.file_mode_iter)
        mode_row.addStretch()
        file_layout.addLayout(mode_row)
        
        # 미리보기
        preview_row = QVBoxLayout()
        preview_row.addWidget(QLabel('전체 스트링 미리보기:'))
        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setMaximumHeight(60)
        self.file_path_input.textChanged.connect(self._update_preview)
        preview_row.addWidget(self.preview_text)
        file_layout.addLayout(preview_row)
        
        self.file_container.setLayout(file_layout)
        self.file_container.setVisible(False)
        main_layout.addWidget(self.file_container)
        
        # 난수 생성 옵션
        self.use_random_checkbox = QCheckBox('난수 생성')
        self.use_random_checkbox.toggled.connect(self._toggle_random_options)
        main_layout.addWidget(self.use_random_checkbox)
        
        # 난수 옵션 컨테이너
        self.random_container = QWidget()
        random_layout = QVBoxLayout()
        random_layout.setContentsMargins(20, 0, 0, 0)
        
        self.random_type_group = QButtonGroup()
        self.random_pure = QRadioButton('진짜 난수 (완전 랜덤)')
        self.random_date = QRadioButton('금일 날짜 포함 난수 (예: 1023ms1)')
        self.random_pure.setChecked(True)
        self.random_type_group.addButton(self.random_pure)
        self.random_type_group.addButton(self.random_date)
        random_layout.addWidget(self.random_pure)
        random_layout.addWidget(self.random_date)
        
        # 난수 길이 설정
        length_row = QHBoxLayout()
        length_row.addWidget(QLabel('난수 길이:'))
        self.random_length = QSpinBox()
        self.random_length.setRange(1, 20)
        self.random_length.setValue(6)
        length_row.addWidget(self.random_length)
        length_row.addWidget(QLabel('자리'))
        length_row.addStretch()
        random_layout.addLayout(length_row)
        
        self.random_container.setLayout(random_layout)
        self.random_container.setVisible(False)
        main_layout.addWidget(self.random_container)
        
        widget.setLayout(main_layout)
        return widget
    
    def _drag_enter_event(self, event):
        """드래그 이벤트 핸들러"""
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()
    
    def _drop_event(self, event):
        """드롭 이벤트 핸들러"""
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0]
            file_path = url.toLocalFile()
            self.file_path_input.setText(file_path)
            event.accept()
        else:
            event.ignore()
    
    def _browse_file(self):
        """파일 선택 다이얼로그"""
        file_path, _ = QFileDialog.getOpenFileName(
            None,
            "텍스트 파일 선택",
            "",
            "Text Files (*.txt);;All Files (*.*)"
        )
        if file_path:
            self.file_path_input.setText(file_path)
    
    def _toggle_file_options(self, checked):
        """파일 옵션 표시/숨김"""
        self.file_container.setVisible(checked)
        if checked:
            self.write_input.setEnabled(False)
            self.use_random_checkbox.setChecked(False)
            self.use_random_checkbox.setEnabled(False)
        else:
            self.write_input.setEnabled(True)
            self.use_random_checkbox.setEnabled(True)
    
    def _toggle_random_options(self, checked):
        """난수 옵션 표시/숨김"""
        self.random_container.setVisible(checked)
        if checked:
            self.write_input.setEnabled(False)
            self.use_file_checkbox.setChecked(False)
            self.use_file_checkbox.setEnabled(False)
        else:
            self.write_input.setEnabled(True)
            self.use_file_checkbox.setEnabled(True)
    
    def _update_preview(self):
        """미리보기 업데이트"""
        if not self.use_file_checkbox.isChecked():
            return
            
        file_path = self.file_path_input.text().strip()
        prefix = self.prefix_input.text()
        suffix = self.suffix_input.text()
        
        if not file_path or not os.path.exists(file_path):
            self.preview_text.setPlainText('(파일을 선택하세요)')
            return
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f.readlines() if line.strip()]
            
            if not lines:
                self.preview_text.setPlainText('(파일이 비어있습니다)')
                return
            
            # 첫 번째 줄로 미리보기 생성
            preview = f"{prefix}{lines[0]}{suffix}"
            total_lines = len(lines)
            self.preview_text.setPlainText(f"{preview}\n\n(총 {total_lines}개 줄이 입력됩니다)")
        except Exception as e:
            self.preview_text.setPlainText(f'(파일 읽기 오류: {str(e)})')
    
    def parse_params(self, params: list) -> dict:
        """파라미터 파싱
        형식: write [text] 또는 
              write --file [path] --prefix [prefix] --suffix [suffix] --mode [all|iter] 또는
              write --random [type] --length [length]
        """
        # 전체 명령어 문자열 재구성
        full_command = 'write ' + ' '.join(params)
        
        # 토큰 분할 (따옴표 고려)
        def tokenize_command(command):
            tokens = []
            current_token = ""
            in_quotes = False
            
            i = 0
            while i < len(command):
                char = command[i]
                
                if char == '"':
                    in_quotes = not in_quotes
                elif char == ' ' and not in_quotes:
                    if current_token:
                        tokens.append(current_token)
                        current_token = ""
                else:
                    current_token += char
                i += 1
            
            if current_token:
                tokens.append(current_token)
            
            return tokens
        
        tokens = tokenize_command(full_command)
        
        # 'write' 명령어 제거
        if tokens and tokens[0] == 'write':
            tokens = tokens[1:]
        
        result = {
            'text': '',
            'use_file': False,
            'file_path': '',
            'prefix': '',
            'suffix': '',
            'file_mode': 'all',  # 'all' 또는 'iter'
            'use_random': False,
            'random_type': 'pure',
            'random_length': 6
        }
        
        i = 0
        while i < len(tokens):
            if tokens[i] == '--file':
                result['use_file'] = True
                if i + 1 < len(tokens):
                    result['file_path'] = tokens[i + 1]
                    i += 2
                else:
                    i += 1
            elif tokens[i] == '--prefix':
                if i + 1 < len(tokens):
                    result['prefix'] = tokens[i + 1]
                    i += 2
                else:
                    i += 1
            elif tokens[i] == '--suffix':
                if i + 1 < len(tokens):
                    result['suffix'] = tokens[i + 1]
                    i += 2
                else:
                    i += 1
            elif tokens[i] == '--mode':
                if i + 1 < len(tokens) and tokens[i + 1] in ['all', 'iter']:
                    result['file_mode'] = tokens[i + 1]
                    i += 2
                else:
                    i += 1
            elif tokens[i] == '--random':
                result['use_random'] = True
                if i + 1 < len(tokens) and tokens[i + 1] in ['pure', 'date']:
                    result['random_type'] = tokens[i + 1]
                    i += 2
                else:
                    i += 1
            elif tokens[i] == '--length':
                if i + 1 < len(tokens):
                    try:
                        result['random_length'] = int(tokens[i + 1])
                    except ValueError:
                        pass
                    i += 2
                else:
                    i += 1
            else:
                if not result['text']:
                    result['text'] = tokens[i]
                else:
                    result['text'] += ' ' + tokens[i]
                i += 1
        
        return result
    
    def set_ui_values(self, params: dict):
        """UI에 값 설정"""
        if params.get('use_file', False):
            self.use_file_checkbox.setChecked(True)
            self.file_path_input.setText(params.get('file_path', ''))
            self.prefix_input.setText(params.get('prefix', ''))
            self.suffix_input.setText(params.get('suffix', ''))
            # 파일 모드 설정
            file_mode = params.get('file_mode', 'all')
            if file_mode == 'iter':
                self.file_mode_iter.setChecked(True)
            else:
                self.file_mode_all.setChecked(True)
        elif params.get('use_random', False):
            self.use_random_checkbox.setChecked(True)
            random_type = params.get('random_type', 'pure')
            if random_type == 'date':
                self.random_date.setChecked(True)
            else:
                self.random_pure.setChecked(True)
            self.random_length.setValue(params.get('random_length', 6))
        else:
            self.write_input.setText(params.get('text', ''))
    
    def get_command_string(self) -> str:
        """명령어 문자열 생성"""
        if self.use_file_checkbox.isChecked():
            file_path = self.file_path_input.text().strip()
            prefix = self.prefix_input.text()
            suffix = self.suffix_input.text()
            file_mode = 'iter' if self.file_mode_iter.isChecked() else 'all'
            
            cmd = f"write --file \"{file_path}\""
            if prefix:
                cmd += f" --prefix \"{prefix}\""
            if suffix:
                cmd += f" --suffix \"{suffix}\""
            cmd += f" --mode {file_mode}"
            return cmd
        elif self.use_random_checkbox.isChecked():
            random_type = 'date' if self.random_date.isChecked() else 'pure'
            length = self.random_length.value()
            return f"write --random {random_type} --length {length}"
        else:
            text = self.write_input.text().strip()
            return f"write {text}" if text else 'write'
    
    def _generate_random_string(self, random_type: str, length: int) -> str:
        """난수 문자열 생성"""
        import random
        import string
        from datetime import datetime
        
        if random_type == 'date':
            # 금일 날짜 포함 난수 (예: 1023ms1)
            today = datetime.now()
            date_part = today.strftime('%m%d')  # MMDD 형식
            
            # 남은 길이만큼 랜덤 문자 생성
            remaining_length = max(1, length - len(date_part))
            random_chars = ''.join(random.choices(string.ascii_lowercase + string.digits, k=remaining_length))
            
            # 날짜를 중간에 삽입
            mid_point = len(random_chars) // 2
            result = random_chars[:mid_point] + date_part + random_chars[mid_point:]
            
            return result
        else:
            # 진짜 난수
            return ''.join(random.choices(string.ascii_lowercase + string.digits, k=length))
    
    def execute(self, params: dict, window_coords=None, processor_state=None):
        """명령어 실행"""
        if params.get('use_file', False):
            # 파일 모드
            file_path = params.get('file_path', '')
            prefix = params.get('prefix', '')
            suffix = params.get('suffix', '')
            file_mode = params.get('file_mode', 'all')
            
            if not file_path or not os.path.exists(file_path):
                print(f'⚠️ 파일을 찾을 수 없습니다: {file_path}')
                return
            
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = [line.strip() for line in f.readlines() if line.strip()]
                
                if file_mode == 'iter':
                    # 반복마다 입력 모드
                    # processor_state에서 현재 반복 횟수 가져오기
                    iteration_count = 1
                    if processor_state:
                        iteration_count = processor_state.get('iteration_count', 1)
                    
                    # 현재 반복 횟수에 해당하는 줄 선택 (1-based)
                    line_idx = iteration_count - 1
                    
                    if line_idx < len(lines):
                        line = lines[line_idx]
                        text = f"{prefix}{line}{suffix}"
                        pyperclip.copy(text)
                        pyd.keyDown('ctrl')
                        pyd.press('v')
                        pyd.keyUp('ctrl')
                        print(f'✍️ 파일 입력 [반복 {iteration_count}] ({line_idx + 1}/{len(lines)}번째 줄): {text}')
                    else:
                        print(f'⚠️ 파일에 {iteration_count}번째 줄이 없습니다. (파일 총 {len(lines)}줄)')
                        # 마지막 줄을 반복 입력
                        if lines:
                            line = lines[-1]
                            text = f"{prefix}{line}{suffix}"
                            pyperclip.copy(text)
                            pyd.keyDown('ctrl')
                            pyd.press('v')
                            pyd.keyUp('ctrl')
                            print(f'  → 마지막 줄 재사용: {text}')
                else:
                    # 한번에 입력 모드
                    print(f'📄 파일에서 {len(lines)}개 줄 읽기: {file_path}')
                    
                    for idx, line in enumerate(lines, 1):
                        text = f"{prefix}{line}{suffix}"
                        pyperclip.copy(text)
                        pyd.keyDown('ctrl')
                        pyd.press('v')
                        pyd.keyUp('ctrl')
                        print(f'  [{idx}/{len(lines)}] 입력: {text}')
                        
                        # 마지막 줄이 아니면 짧은 대기
                        if idx < len(lines):
                            time.sleep(0.1)
                    
                    print(f'✅ 파일 내용 입력 완료')
            except Exception as e:
                print(f'❌ 파일 읽기 오류: {str(e)}')
        
        elif params.get('use_random', False):
            # 난수 모드
            random_type = params.get('random_type', 'pure')
            length = params.get('random_length', 6)
            
            text = self._generate_random_string(random_type, length)
            pyperclip.copy(text)
            pyd.keyDown('ctrl')
            pyd.press('v')
            pyd.keyUp('ctrl')
            print(f'🎲 난수 입력 ({random_type}): {text}')
        
        else:
            # 일반 텍스트 모드
            text = params.get('text', '')
            if text:
                pyperclip.copy(text)
                pyd.keyDown('ctrl')
                pyd.press('v')
                pyd.keyUp('ctrl')
                print(f'✍️ 텍스트 입력: {text}')


class WaitCommand(CommandBase):
    """대기 명령어"""
    
    @property
    def name(self) -> str:
        return "Wait"
    
    @property
    def description(self) -> str:
        return "Wait for specified seconds"
    
    def _interruptible_sleep(self, duration, params):
        """중지 플래그를 체크하면서 대기하는 함수
        
        Args:
            duration: 대기 시간 (초)
            params: 파라미터 딕셔너리 (processor_stop_flag 체크용)
        
        Returns:
            bool: True if interrupted (중지됨), False if completed (완료됨)
        """
        if duration <= 0:
            if duration > 0:
                time.sleep(duration)
            return False
            
        # 0.1초 간격으로 중지 플래그 체크
        check_interval = 0.1
        total_slept = 0
        
        # popup 참조 가져오기
        popup = None
        if params:
            processor = params.get('processor')
            if processor and hasattr(processor, 'state') and processor.state:
                popup = processor.state.get('popup')
        
        while total_slept < duration:
            # params에서 CommandProcessor의 실시간 stop_flag 체크
            processor = params.get('processor') if params else None
            if processor and hasattr(processor, 'stop_flag') and processor.stop_flag:
                print(f"⚠️ Wait 중지됨 (경과시간: {total_slept:.1f}초/{duration}초)")
                return True  # 중지됨
            
            # popup 타이머 업데이트 (0.5초마다)
            if popup and hasattr(popup, 'update_timer') and int(total_slept * 10) % 5 == 0:
                try:
                    popup.update_timer(total_slept, duration)
                except Exception:
                    pass  # popup이 닫혔을 수 있음
                
            sleep_time = min(check_interval, duration - total_slept)
            time.sleep(sleep_time)
            total_slept += sleep_time
            
        return False  # 완료됨
    
    def create_ui(self) -> QWidget:
        widget = QWidget()
        layout = QHBoxLayout()
        self.wait_input = QSpinBox()
        self.wait_input.setRange(0, 3600)
        self.wait_input.setSuffix(' s')
        layout.addWidget(QLabel('Seconds:'))
        layout.addWidget(self.wait_input)
        widget.setLayout(layout)
        return widget
    
    def parse_params(self, params: list) -> dict:
        duration = 0
        if params and len(params) > 0:
            try:
                duration = float(params[0])
            except ValueError:
                duration = 0
        return {'duration': duration}
    
    def set_ui_values(self, params: dict):
        self.wait_input.setValue(int(params.get('duration', 0)))
    
    def get_command_string(self) -> str:
        return f"wait {self.wait_input.value()}"
    
    def execute(self, params: dict, window_coords=None, processor_state=None):
        duration = params.get('duration', 0)
        if duration > 0:
            if self._interruptible_sleep(duration, params):
                print(f'Wait 중지됨')
                return
            print(f'Waited {duration} seconds')


class ScreenshotCommand(CommandBase):
    """스크린샷 명령어"""
    
    @property
    def name(self) -> str:
        return "Screenshot"
    
    @property
    def description(self) -> str:
        return "Take screenshot"
    
    def create_ui(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Coordinate input row
        coord_row = QHBoxLayout()
        self.ss_x = QLineEdit()
        self.ss_y = QLineEdit()
        self.ss_w = QLineEdit()
        self.ss_h = QLineEdit()
        self.ss_x.setPlaceholderText('상대 x')
        self.ss_y.setPlaceholderText('상대 y')
        self.ss_w.setPlaceholderText('w')
        self.ss_h.setPlaceholderText('h')
        coord_row.addWidget(QLabel('상대 x'))
        coord_row.addWidget(self.ss_x)
        coord_row.addWidget(QLabel('상대 y'))
        coord_row.addWidget(self.ss_y)
        coord_row.addWidget(QLabel('w'))
        coord_row.addWidget(self.ss_w)
        coord_row.addWidget(QLabel('h'))
        coord_row.addWidget(self.ss_h)
        layout.addLayout(coord_row)
        
        # Helper buttons row
        helper_row = QHBoxLayout()
        self.ss_get_coord_btn = QPushButton('Get Coordinates (Shift+F4)')
        self.ss_get_coord_btn.clicked.connect(lambda: self.on_ss_get_coordinates(self.ss_x, self.ss_y, self.ss_w, self.ss_h))
        self.ss_test_ocr_btn = QPushButton('Test OCR')
        self.ss_test_ocr_btn.clicked.connect(self.on_ss_test_ocr)
        helper_row.addWidget(self.ss_get_coord_btn)
        helper_row.addWidget(self.ss_test_ocr_btn)
        layout.addLayout(helper_row)
        
        widget.setLayout(layout)
        return widget
    
    def on_ss_get_coordinates(self, x_field, y_field, w_field, h_field):
        """영역 선택 도구"""
        from PyQt5.QtWidgets import QApplication, QRubberBand
        from PyQt5.QtCore import QRect, QPoint, QSize
        
        class DragSelector(QWidget):
            def __init__(self):
                super().__init__()
                self.setWindowFlag(Qt.FramelessWindowHint)
                self.setAttribute(Qt.WA_TransparentForMouseEvents, False)
                self.setWindowOpacity(0.15)
                self.screen_geo = QApplication.primaryScreen().geometry()
                self.setGeometry(self.screen_geo)
                self.rubber_band = QRubberBand(QRubberBand.Rectangle, self)
                self.origin = QPoint()
                self.setCursor(Qt.CrossCursor)
                self.rect_result = None

            def mousePressEvent(self, event):
                self.origin = event.pos()
                self.rubber_band.setGeometry(QRect(self.origin, QSize()))
                self.rubber_band.show()

            def mouseMoveEvent(self, event):
                self.rubber_band.setGeometry(QRect(self.origin, event.pos()).normalized())

            def mouseReleaseEvent(self, event):
                r = self.rubber_band.geometry().getRect()
                self.rect_result = r
                self.rubber_band.hide()
                self.close()

            def get_rect(self):
                return self.rect_result

        app = QApplication.instance()
        drag_selector = DragSelector()
        drag_selector.setWindowOpacity(0.20)
        drag_selector.setWindowModality(Qt.ApplicationModal)
        drag_selector.show()

        while drag_selector.isVisible():
            app.processEvents()

        rect = drag_selector.get_rect()
        if rect:
            x, y, w, h = rect
            
            # 선택된 윈도우 기준으로 상대 좌표 계산
            try:
                import pygetwindow as gw
                from PyQt5.QtWidgets import QApplication
                
                app = QApplication.instance()
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'window_dropdown'):
                        selected_window = widget.window_dropdown.currentText()
                        if selected_window:
                            windows = gw.getWindowsWithTitle(selected_window)
                            if windows:
                                window = windows[0]
                                win_x, win_y = window.left, window.top
                                
                                # 상대 좌표 계산
                                rel_x = x - win_x
                                rel_y = y - win_y
                                
                                x_field.setText(str(rel_x))
                                y_field.setText(str(rel_y))
                                w_field.setText(str(w))
                                h_field.setText(str(h))
                                QMessageBox.information(None, "좌표 선택", f"상대 좌표 선택됨: ({rel_x}, {rel_y}, {w}, {h})\n절대 좌표: ({x}, {y}, {w}, {h})")
                                return
                
                # 윈도우를 찾지 못했을 때 절대 좌표 사용
                x_field.setText(str(x))
                y_field.setText(str(y))
                w_field.setText(str(w))
                h_field.setText(str(h))
                QMessageBox.information(None, "좌표 선택", f"윈도우 미선택. 절대 좌표 사용: ({x}, {y}, {w}, {h})")
                
            except Exception as e:
                # 오류 발생시 절대 좌표 사용
                x_field.setText(str(x))
                y_field.setText(str(y))
                w_field.setText(str(w))
                h_field.setText(str(h))
                QMessageBox.information(None, "좌표 선택", f"상대 좌표 계산 실패. 절대 좌표 사용: ({x}, {y}, {w}, {h})")
    
    def on_ss_test_ocr(self):
        """OCR 테스트 - 이미지 미리보기 포함"""
        try:
            x = int(self.ss_x.text().strip()) if self.ss_x.text().strip() else None
            y = int(self.ss_y.text().strip()) if self.ss_y.text().strip() else None
            w = int(self.ss_w.text().strip()) if self.ss_w.text().strip() else None
            h = int(self.ss_h.text().strip()) if self.ss_h.text().strip() else None
            
            if x is not None and y is not None and w is not None and h is not None:
                screenshot_path = take_screenshot_with_coords(x, y, w, h)
                info = f"Region: ({x}, {y}, {w}, {h})"
            else:
                screenshot_path = take_screenshot()
                info = "Full screen"
            
            extracted = image_to_text(img_path=screenshot_path, lang='auto')
            if not extracted:
                extracted = "(No text found)"
            
            # 이미지 미리보기가 포함된 커스텀 다이얼로그 생성
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QScrollArea
            from PyQt5.QtGui import QPixmap
            from PyQt5.QtCore import Qt
            
            dialog = QDialog()
            dialog.setWindowTitle("OCR Test Result")
            dialog.setMinimumWidth(600)
            
            layout = QVBoxLayout()
            
            # 정보 레이블
            info_label = QLabel(f"📍 {info}\n📁 {screenshot_path}")
            info_label.setWordWrap(True)
            layout.addWidget(info_label)
            
            # 이미지 미리보기
            try:
                pixmap = QPixmap(screenshot_path)
                if not pixmap.isNull():
                    # 적절한 크기로 조정 (최대 800x600, 종횡비 유지)
                    max_width = 800
                    max_height = 600
                    
                    if pixmap.width() > max_width or pixmap.height() > max_height:
                        pixmap = pixmap.scaled(max_width, max_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    
                    image_label = QLabel()
                    image_label.setPixmap(pixmap)
                    image_label.setAlignment(Qt.AlignCenter)
                    
                    # 스크롤 영역에 이미지 추가
                    scroll_area = QScrollArea()
                    scroll_area.setWidget(image_label)
                    scroll_area.setWidgetResizable(True)
                    scroll_area.setMaximumHeight(600)
                    layout.addWidget(scroll_area)
            except Exception as e:
                error_label = QLabel(f"⚠️ 이미지 로드 실패: {e}")
                layout.addWidget(error_label)
            
            # OCR 결과
            result_label = QLabel("🔍 OCR 결과:")
            result_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
            layout.addWidget(result_label)
            
            extracted_label = QLabel(extracted)
            extracted_label.setWordWrap(True)
            extracted_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
            extracted_label.setStyleSheet("background-color: #f0f0f0; padding: 10px; border-radius: 5px;")
            layout.addWidget(extracted_label)
            
            # 닫기 버튼
            button_layout = QHBoxLayout()
            button_layout.addStretch()
            close_btn = QPushButton("닫기")
            close_btn.clicked.connect(dialog.accept)
            button_layout.addWidget(close_btn)
            layout.addLayout(button_layout)
            
            dialog.setLayout(layout)
            dialog.exec_()
        
        except ValueError:
            QMessageBox.warning(None, "Test OCR", "Invalid coordinate values. Please enter valid numbers.")
        except Exception as e:
            QMessageBox.critical(None, "Test OCR Error", f"Error during OCR test:\n{e}")
    
    def parse_params(self, params: list) -> dict:
        if len(params) >= 4:
            return {'x': params[0], 'y': params[1], 'w': params[2], 'h': params[3]}
        return {}
    
    def set_ui_values(self, params: dict):
        self.ss_x.setText(params.get('x', ''))
        self.ss_y.setText(params.get('y', ''))
        self.ss_w.setText(params.get('w', ''))
        self.ss_h.setText(params.get('h', ''))
    
    def get_command_string(self) -> str:
        x = self.ss_x.text().strip()
        y = self.ss_y.text().strip()
        w = self.ss_w.text().strip()
        h = self.ss_h.text().strip()
        return f"screenshot {x} {y} {w} {h}" if x and y and w and h else 'screenshot'
    
    def execute(self, params: dict, window_coords=None, processor_state=None):
        x = params.get('x')
        y = params.get('y')
        w = params.get('w')
        h = params.get('h')
        
        if x and y and w and h:
            try:
                x, y, w, h = int(x), int(y), int(w), int(h)
                
                # 상대 좌표를 절대 좌표로 변환
                if window_coords:
                    adjusted_x, adjusted_y = calculate_adjusted_coordinates(x, y, window_coords)
                    screenshot_path = take_screenshot_with_coords(adjusted_x, adjusted_y, w, h)
                    print(f"Screenshot taken at relative ({x},{y},{w},{h}) -> absolute ({adjusted_x},{adjusted_y},{w},{h}): {screenshot_path}")
                else:
                    # window_coords가 없으면 절대 좌표로 처리
                    screenshot_path = take_screenshot_with_coords(x, y, w, h)
                    print(f"Screenshot taken at absolute ({x},{y},{w},{h}): {screenshot_path}")
            except (ValueError, TypeError) as e:
                print(f"Invalid coordinates for screenshot: {e}")
                screenshot_path = take_screenshot()
        else:
            screenshot_path = take_screenshot()
            print(f"Screenshot taken (full screen): {screenshot_path}")
        
        if processor_state:
            processor_state['screenshot_path'] = screenshot_path


class ClickCommand(CommandBase):
    """클릭 명령어"""
    
    @property
    def name(self) -> str:
        return "Click"
    
    @property
    def description(self) -> str:
        return "Click at coordinates"
    
    def _interruptible_sleep(self, duration, params):
        """중지 플래그를 체크하면서 대기하는 함수
        
        Args:
            duration: 대기 시간 (초)
            params: 파라미터 딕셔너리 (processor_stop_flag 체크용)
        
        Returns:
            bool: True if interrupted (중지됨), False if completed (완료됨)
        """
        if duration <= 0:
            if duration > 0:
                time.sleep(duration)
            return False
            
        # 0.1초 간격으로 중지 플래그 체크
        check_interval = 0.1
        total_slept = 0
        
        while total_slept < duration:
            # params에서 CommandProcessor의 실시간 stop_flag 체크
            processor = params.get('processor') if params else None
            if processor and hasattr(processor, 'stop_flag') and processor.stop_flag:
                print(f"⚠️ Click Hold 중지됨 (경과시간: {total_slept:.1f}초/{duration}초)")
                return True  # 중지됨
                
            sleep_time = min(check_interval, duration - total_slept)
            time.sleep(sleep_time)
            total_slept += sleep_time
            
        return False  # 완료됨
    
    def create_ui(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Coordinate input row
        click_coord_row = QHBoxLayout()
        self.click_x = QLineEdit()
        self.click_y = QLineEdit()
        self.click_x.setPlaceholderText('상대 x')
        self.click_y.setPlaceholderText('상대 y')
        click_coord_row.addWidget(QLabel('상대 x'))
        click_coord_row.addWidget(self.click_x)
        click_coord_row.addWidget(QLabel('상대 y'))
        click_coord_row.addWidget(self.click_y)
        layout.addLayout(click_coord_row)
        
        # Hold duration row
        hold_row = QHBoxLayout()
        self.click_hold_input = QLineEdit()
        self.click_hold_input.setPlaceholderText('0')
        self.click_hold_input.setText('0')
        hold_row.addWidget(QLabel('Hold (초):'))
        hold_row.addWidget(self.click_hold_input)
        hold_row.addStretch()
        layout.addLayout(hold_row)
        
        # Coordinate mode selection row
        coord_mode_row = QHBoxLayout()
        self.coord_mode_combo = QComboBox()
        self.coord_mode_combo.addItems(['스케일링 (기준해상도 기반)', '오프셋 (단순 위치이동)'])
        self.coord_mode_combo.setCurrentIndex(1)  # 기본값: 스케일링
        coord_mode_row.addWidget(QLabel('좌표 모드:'))
        coord_mode_row.addWidget(self.coord_mode_combo)
        coord_mode_row.addStretch()
        layout.addLayout(coord_mode_row)
        
        # Helper button row
        click_helper_row = QHBoxLayout()
        self.click_get_coord_btn = QPushButton('Get Coordinates (F2)')
        self.click_get_coord_btn.setShortcut('F2')
        self.click_get_coord_btn.clicked.connect(lambda: self.on_click_get_coordinates(self.click_x, self.click_y))
        click_helper_row.addWidget(self.click_get_coord_btn)
        click_helper_row.addStretch()
        layout.addLayout(click_helper_row)
        
        widget.setLayout(layout)
        return widget
    
    def on_click_get_coordinates(self, x_field, y_field):
        """현재 마우스 위치를 상대 좌표로 가져오기"""
        try:
            import pyautogui
            import pygetwindow as gw
            
            # 절대 마우스 위치 가져오기
            abs_x, abs_y = pyautogui.position()
            
            # 메인 앱에서 선택된 윈도우 정보 가져오기 (해킹적이지만 필요)
            try:
                # QApplication을 통해 메인 윈도우 찾기 (임시 방법)
                from PyQt5.QtWidgets import QApplication
                app = QApplication.instance()
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'window_dropdown'):
                        selected_window = widget.window_dropdown.currentText()
                        if selected_window:
                            windows = gw.getWindowsWithTitle(selected_window)
                            if windows:
                                window = windows[0]
                                win_x, win_y = window.left, window.top
                                
                                # 상대 좌표 계산
                                rel_x = abs_x - win_x
                                rel_y = abs_y - win_y
                                
                                x_field.setText(str(rel_x))
                                y_field.setText(str(rel_y))
                                print(f"마우스 상대 좌표: ({rel_x}, {rel_y}) | 절대 좌표: ({abs_x}, {abs_y})")
                                return
                
                # 윈도우를 찾지 못했을 때 절대 좌표 사용
                x_field.setText(str(abs_x))
                y_field.setText(str(abs_y))
                print(f"윈도우가 선택되지 않음. 절대 좌표 사용: ({abs_x}, {abs_y})")
                
            except Exception as e:
                # 오류 발생시 절대 좌표 사용
                x_field.setText(str(abs_x))
                y_field.setText(str(abs_y))
                print(f"상대 좌표 계산 실패, 절대 좌표 사용: ({abs_x}, {abs_y}) - {e}")
                
        except Exception as e:
            print(f"마우스 좌표를 가져올 수 없음: {e}")
    
    def parse_params(self, params: list) -> dict:
        result = {'x': '', 'y': '', 'hold': 0, 'coord_mode': 'scaled'}
        if len(params) >= 2:
            result['x'] = params[0]
            result['y'] = params[1]
            # 세 번째 파라미터부터 처리
            if len(params) >= 3:
                # 마지막 파라미터가 좌표 모드인지 확인
                if params[-1] in ['scaled', 'offset']:
                    result['coord_mode'] = params[-1]
                    # hold 시간은 그 앞 파라미터
                    if len(params) >= 4:
                        try:
                            result['hold'] = float(params[2])
                        except ValueError:
                            result['hold'] = 0
                else:
                    # 좌표 모드가 없으면 기본값 사용하고 hold 시간만 파싱
                    try:
                        result['hold'] = float(params[2])
                    except ValueError:
                        result['hold'] = 0
        return result
    
    def set_ui_values(self, params: dict):
        self.click_x.setText(params.get('x', ''))
        self.click_y.setText(params.get('y', ''))
        self.click_hold_input.setText(str(params.get('hold', 0)))
        # 좌표 모드 설정
        coord_mode = params.get('coord_mode', 'scaled')
        if coord_mode == 'offset':
            self.coord_mode_combo.setCurrentIndex(1)
        else:
            self.coord_mode_combo.setCurrentIndex(0)
    
    def get_command_string(self) -> str:
        x = self.click_x.text().strip()
        y = self.click_y.text().strip()
        hold = self.click_hold_input.text().strip()
        coord_mode = 'offset' if self.coord_mode_combo.currentIndex() == 1 else 'scaled'
        
        if x and y and hold and hold != '0':
            return f"click {x} {y} {hold} {coord_mode}"
        elif x and y:
            return f"click {x} {y} {coord_mode}"
        else:
            return 'click'
    
    def execute(self, params: dict, window_coords=None, processor_state=None):
        x = params.get('x')
        y = params.get('y')
        hold_time = float(params.get('hold', 0))
        
        if not x or not y:
            print("click requires x and y coordinates")
            return
        if not window_coords:
            print("Window coordinates not available for click")
            return
            
        try:
            x, y = int(x), int(y)
            coord_mode = params.get('coord_mode', 'scaled')
            
            # 좌표 모드에 따라 다른 계산 함수 사용
            if coord_mode == 'offset':
                adjusted_x, adjusted_y = calculate_offset_coordinates(x, y, window_coords)
            else:  # 'scaled'
                adjusted_x, adjusted_y = calculate_adjusted_coordinates(x, y, window_coords)
            
            pyd.moveTo(adjusted_x, adjusted_y)
            
            if hold_time > 0:
                # Hold 모드: 마우스를 누르고 있다가 뗌
                pyd.mouseDown()
                print(f'Mouse down at ({adjusted_x}, {adjusted_y}) (holding for {hold_time}초)')
                
                # 중지 플래그를 체크하면서 대기
                if self._interruptible_sleep(hold_time, params):
                    pyd.mouseUp()
                    print(f'Mouse up at ({adjusted_x}, {adjusted_y}) (중지됨)')
                    return
                    
                pyd.mouseUp()
                print(f'Mouse up at ({adjusted_x}, {adjusted_y})')
            else:
                # 기본 모드: 즉시 클릭                                      
                pyd.mouseDown()
                time.sleep(0.05)
                pyd.mouseUp()
                print(f'Clicked at adjusted coordinates: ({adjusted_x}, {adjusted_y})')
        except (ValueError, TypeError) as e:
            print(f"Invalid coordinates for click: {e}")


class DragCommand(CommandBase):
    """마우스 드래그 명령어 - 새로운 명령어 예시! 🎉"""
    
    @property
    def name(self) -> str:
        return "Drag"
    
    @property
    def description(self) -> str:
        return "Mouse drag from point A to point B"
    
    def create_ui(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout()
        
        # 좌표 입력
        coords_layout = QHBoxLayout()
        self.x1_input = QLineEdit()
        self.y1_input = QLineEdit()
        self.x2_input = QLineEdit()
        self.y2_input = QLineEdit()
        
        self.x1_input.setPlaceholderText('상대 x1')
        self.y1_input.setPlaceholderText('상대 y1')
        self.x2_input.setPlaceholderText('상대 x2')
        self.y2_input.setPlaceholderText('상대 y2')
        
        coords_layout.addWidget(QLabel('시작 (상대):'))
        coords_layout.addWidget(self.x1_input)
        coords_layout.addWidget(self.y1_input)
        coords_layout.addWidget(QLabel('끝 (상대):'))
        coords_layout.addWidget(self.x2_input)
        coords_layout.addWidget(self.y2_input)
        
        # 헬퍼 버튼
        helper_row = QHBoxLayout()
        self.get_start_btn = QPushButton('Get Start Position (F3)')
        self.get_start_btn.setShortcut('F3')
        self.get_start_btn.clicked.connect(self.get_start_position)
        
        self.get_end_btn = QPushButton('Get End Position (F4)')
        self.get_end_btn.setShortcut('F4')
        self.get_end_btn.clicked.connect(self.get_end_position)
        
        helper_row.addWidget(self.get_start_btn)
        helper_row.addWidget(self.get_end_btn)
        
        layout.addLayout(coords_layout)
        layout.addLayout(helper_row)
        widget.setLayout(layout)
        return widget
    
    def get_start_position(self):
        """시작 위치를 상대 좌표로 가져오기"""
        try:
            import pyautogui
            import pygetwindow as gw
            
            # 절대 마우스 위치 가져오기
            abs_x, abs_y = pyautogui.position()
            
            # 메인 앱에서 선택된 윈도우 정보 가져오기
            try:
                from PyQt5.QtWidgets import QApplication
                app = QApplication.instance()
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'window_dropdown'):
                        selected_window = widget.window_dropdown.currentText()
                        if selected_window:
                            windows = gw.getWindowsWithTitle(selected_window)
                            if windows:
                                window = windows[0]
                                win_x, win_y = window.left, window.top
                                
                                # 상대 좌표 계산
                                rel_x = abs_x - win_x
                                rel_y = abs_y - win_y
                                
                                self.x1_input.setText(str(rel_x))
                                self.y1_input.setText(str(rel_y))
                                print(f"시작점 상대 좌표: ({rel_x}, {rel_y}) | 절대 좌표: ({abs_x}, {abs_y})")
                                return
                
                # 윈도우를 찾지 못했을 때 절대 좌표 사용
                self.x1_input.setText(str(abs_x))
                self.y1_input.setText(str(abs_y))
                print(f"윈도우가 선택되지 않음. 절대 좌표 사용: ({abs_x}, {abs_y})")
                
            except Exception as e:
                # 오류 발생시 절대 좌표 사용
                self.x1_input.setText(str(abs_x))
                self.y1_input.setText(str(abs_y))
                print(f"상대 좌표 계산 실패, 절대 좌표 사용: ({abs_x}, {abs_y}) - {e}")
                
        except Exception as e:
            print(f"마우스 좌표를 가져올 수 없음: {e}")
    
    def get_end_position(self):
        """끝 위치를 상대 좌표로 가져오기"""
        try:
            import pyautogui
            import pygetwindow as gw
            
            # 절대 마우스 위치 가져오기
            abs_x, abs_y = pyautogui.position()
            
            # 메인 앱에서 선택된 윈도우 정보 가져오기
            try:
                from PyQt5.QtWidgets import QApplication
                app = QApplication.instance()
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'window_dropdown'):
                        selected_window = widget.window_dropdown.currentText()
                        if selected_window:
                            windows = gw.getWindowsWithTitle(selected_window)
                            if windows:
                                window = windows[0]
                                win_x, win_y = window.left, window.top
                                
                                # 상대 좌표 계산
                                rel_x = abs_x - win_x
                                rel_y = abs_y - win_y
                                
                                self.x2_input.setText(str(rel_x))
                                self.y2_input.setText(str(rel_y))
                                print(f"끝점 상대 좌표: ({rel_x}, {rel_y}) | 절대 좌표: ({abs_x}, {abs_y})")
                                return
                
                # 윈도우를 찾지 못했을 때 절대 좌표 사용
                self.x2_input.setText(str(abs_x))
                self.y2_input.setText(str(abs_y))
                print(f"윈도우가 선택되지 않음. 절대 좌표 사용: ({abs_x}, {abs_y})")
                
            except Exception as e:
                # 오류 발생시 절대 좌표 사용
                self.x2_input.setText(str(abs_x))
                self.y2_input.setText(str(abs_y))
                print(f"상대 좌표 계산 실패, 절대 좌표 사용: ({abs_x}, {abs_y}) - {e}")
                
        except Exception as e:
            print(f"마우스 좌표를 가져올 수 없음: {e}")
    
    def parse_params(self, params: list) -> dict:
        if len(params) >= 4:
            return {
                'x1': params[0],
                'y1': params[1], 
                'x2': params[2],
                'y2': params[3]
            }
        return {}
    
    def set_ui_values(self, params: dict):
        self.x1_input.setText(params.get('x1', ''))
        self.y1_input.setText(params.get('y1', ''))
        self.x2_input.setText(params.get('x2', ''))
        self.y2_input.setText(params.get('y2', ''))
    
    def get_command_string(self) -> str:
        x1 = self.x1_input.text().strip()
        y1 = self.y1_input.text().strip()
        x2 = self.x2_input.text().strip()
        y2 = self.y2_input.text().strip()
        
        if x1 and y1 and x2 and y2:
            return f"drag {x1} {y1} {x2} {y2}"
        return 'drag'
    
    def execute(self, params: dict, window_coords=None, processor_state=None):
        """드래그 실행"""
        try:
            x1 = int(params.get('x1', 0))
            y1 = int(params.get('y1', 0))
            x2 = int(params.get('x2', 0))
            y2 = int(params.get('y2', 0))
            
            # 윈도우 좌표 조정
            if window_coords:
                adj_x1, adj_y1 = calculate_adjusted_coordinates(x1, y1, window_coords)
                adj_x2, adj_y2 = calculate_adjusted_coordinates(x2, y2, window_coords)
            else:
                adj_x1, adj_y1 = x1, y1
                adj_x2, adj_y2 = x2, y2
            
            # 드래그 실행
            pyd.moveTo(adj_x1, adj_y1)
            pyd.mouseDown()
            time.sleep(0.1)
            pyd.moveTo(adj_x2, adj_y2)
            time.sleep(0.1)
            pyd.mouseUp()
            print(f'Dragged from ({adj_x1}, {adj_y1}) to ({adj_x2}, {adj_y2})')
        except ValueError as e:
            print(f"Invalid coordinates for drag: {e}")


# 간단한 명령어들 (파라미터가 없는 것들)
class I2sCommand(CommandBase):
    @property
    def name(self): return "I2S"
    @property 
    def description(self): return "OCR in English"
    def create_ui(self): 
        w = QWidget()
        l = QHBoxLayout()
        l.addWidget(QLabel('No parameters'))
        w.setLayout(l)
        return w
    def parse_params(self, params): return {}
    def set_ui_values(self, params): pass
    def get_command_string(self): return "i2s"
    def execute(self, params, window_coords=None, processor_state=None):
        if processor_state and processor_state.get('screenshot_path'):
            processor_state['extracted_text'] = image_to_text(processor_state['screenshot_path'], lang='eng')
            print(f'OCR (English): {processor_state["extracted_text"]}')


class I2skrCommand(CommandBase):
    @property
    def name(self): return "I2SKR"
    @property
    def description(self): return "OCR in Korean"
    def create_ui(self): 
        w = QWidget()
        l = QHBoxLayout()
        l.addWidget(QLabel('No parameters'))
        w.setLayout(l)
        return w
    def parse_params(self, params): return {}
    def set_ui_values(self, params): pass
    def get_command_string(self): return "i2skr"
    def execute(self, params, window_coords=None, processor_state=None):
        if processor_state and processor_state.get('screenshot_path'):
            processor_state['extracted_text'] = image_to_text(processor_state['screenshot_path'], lang='kor')
            print(f'OCR (Korean): {processor_state["extracted_text"]}')


class OCRCommand(CommandBase):
    """개선된 OCR 명령어 - 자동 언어 감지 및 다중 줄 지원"""
    
    @property
    def name(self): return "OCR"
    
    @property
    def description(self): return "OCR with Auto Language Detection (영어+한글 자동 감지, 다중 줄 지원)"
    
    def create_ui(self): 
        widget = QWidget()
        layout = QVBoxLayout()
        
        info_label = QLabel("📝 개선된 OCR - 자동 언어 감지 및 다중 줄 지원")
        info_label.setStyleSheet("font-weight: bold; color: #2E8B57;")
        layout.addWidget(info_label)
        
        desc_label = QLabel("최신 스크린샷에서 텍스트를 추출합니다.\n"
                           "• 영어와 한글을 자동으로 감지\n"
                           "• 여러 줄의 텍스트 지원\n"
                           "• 다양한 전처리로 인식률 향상")
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        widget.setLayout(layout)
        return widget
    
    def parse_params(self, params): 
        return {}
    
    def set_ui_values(self, params): 
        pass
    
    def get_command_string(self): 
        return "ocr"
    
    def execute(self, params, window_coords=None, processor_state=None):
        if processor_state and processor_state.get('screenshot_path'):
            # 자동 언어 감지 사용
            extracted = image_to_text(processor_state['screenshot_path'], lang='auto')
            processor_state['extracted_text'] = extracted if extracted else ""
            print(f'🔍 OCR (자동 감지): {processor_state["extracted_text"]}')

class WaitUntilCommand(CommandBase):
    """텍스트가 나타날 때까지 대기하는 명령어"""
    
    @property
    def name(self): 
        return "WaitUntil"
    
    @property
    def description(self): 
        return "스크린샷을 1초마다 찍어서 입력한 텍스트가 출력될 때까지 반복"
    
    def _interruptible_sleep(self, duration, params, context=""):
        """중지 플래그를 체크하면서 대기하는 함수
        
        Args:
            duration: 대기 시간 (초)
            params: 파라미터 딕셔너리 (processor_stop_flag 체크용)
            context: 대기 컨텍스트 (디버깅용)
        
        Returns:
            bool: True if interrupted (중지됨), False if completed (완료됨)
        """
        if duration <= 0:
            return False
            
        # 0.1초 간격으로 중지 플래그 체크
        check_interval = 0.1
        total_slept = 0
        
        while total_slept < duration:
            # params에서 CommandProcessor의 실시간 stop_flag 체크
            processor = params.get('processor') if params else None
            if processor and hasattr(processor, 'stop_flag') and processor.stop_flag:
                print(f"⚠️ WaitUntil {context} 중지됨 (경과시간: {total_slept:.1f}초/{duration}초)")
                return True  # 중지됨
            
            sleep_time = min(check_interval, duration - total_slept)
            time.sleep(sleep_time)
            total_slept += sleep_time
            
        return False  # 완료됨
    
    def create_ui(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # 좌표 입력
        coord_layout = QHBoxLayout()
        coord_layout.addWidget(QLabel('상대 X:'))
        self.x_input = QSpinBox()
        self.x_input.setRange(-9999, 9999)  # 상대 좌표는 음수일 수 있음
        coord_layout.addWidget(self.x_input)
        
        coord_layout.addWidget(QLabel('상대 Y:'))
        self.y_input = QSpinBox()
        self.y_input.setRange(-9999, 9999)  # 상대 좌표는 음수일 수 있음
        coord_layout.addWidget(self.y_input)
        
        coord_layout.addWidget(QLabel('Width:'))
        self.width_input = QSpinBox()
        self.width_input.setRange(1, 9999)
        self.width_input.setValue(100)
        coord_layout.addWidget(self.width_input)
        
        coord_layout.addWidget(QLabel('Height:'))
        self.height_input = QSpinBox()
        self.height_input.setRange(1, 9999)
        self.height_input.setValue(50)
        coord_layout.addWidget(self.height_input)
        
        layout.addLayout(coord_layout)
        
        # 좌표 선택 및 테스트 버튼
        coord_btn_layout = QHBoxLayout()
        self.get_coord_btn = QPushButton('좌표 선택 (드래그)')
        self.get_coord_btn.clicked.connect(self.on_get_coordinates)
        self.test_ocr_btn = QPushButton('OCR 테스트')
        self.test_ocr_btn.clicked.connect(self.on_test_ocr)
        coord_btn_layout.addWidget(self.get_coord_btn)
        coord_btn_layout.addWidget(self.test_ocr_btn)
        layout.addLayout(coord_btn_layout)
        
        # OCR 언어 선택
        ocr_layout = QHBoxLayout()
        ocr_layout.addWidget(QLabel('OCR:'))
        self.ocr_combo = QComboBox()
        self.ocr_combo.addItems(['i2s (English)', 'i2skr (Korean)'])
        ocr_layout.addWidget(self.ocr_combo)
        layout.addLayout(ocr_layout)
        
        # 타겟 텍스트 입력
        text_layout = QHBoxLayout()
        text_layout.addWidget(QLabel('찾을 텍스트:'))
        self.text_input = QLineEdit()
        self.text_input.setPlaceholderText("찾을 텍스트를 입력하세요")
        text_layout.addWidget(self.text_input)
        layout.addLayout(text_layout)
        
        # 매칭 모드 선택
        match_layout = QHBoxLayout()
        match_layout.addWidget(QLabel('매칭 모드:'))
        self.match_mode_combo = QComboBox()
        self.match_mode_combo.addItems(['일부 포함', '완전 일치'])
        match_layout.addWidget(self.match_mode_combo)
        layout.addLayout(match_layout)
        
        # 최대 시도 횟수
        tries_layout = QHBoxLayout()
        tries_layout.addWidget(QLabel('최대 시도:'))
        self.max_tries_input = QSpinBox()
        self.max_tries_input.setRange(1, 10000)
        self.max_tries_input.setValue(10)
        self.max_tries_input.setSuffix('회')
        tries_layout.addWidget(self.max_tries_input)
        layout.addLayout(tries_layout)
        
        # 좌표 모드 선택
        coord_mode_layout = QHBoxLayout()
        coord_mode_layout.addWidget(QLabel('좌표 모드:'))
        self.coord_mode_combo = QComboBox()
        self.coord_mode_combo.addItems(['스케일링 (기준해상도 기반)', '오프셋 (단순 위치이동)'])
        self.coord_mode_combo.setCurrentIndex(1)  # 기본값: 스케일링
        coord_mode_layout.addWidget(self.coord_mode_combo)
        layout.addLayout(coord_mode_layout)
        
        widget.setLayout(layout)
        return widget
    
    def parse_params(self, params):
        if len(params) < 6:
            return {}
        
        try:
            # 파라미터 파싱: x y width height ocr_type "target_text" [exact_match] [max_tries]
            parsed = {
                'x': int(params[0]),
                'y': int(params[1]), 
                'width': int(params[2]),
                'height': int(params[3]),
                'ocr_type': params[4],
            }
            
            # target_text 파싱 (따옴표 처리)
            if params[5].startswith('"') and params[5].endswith('"'):
                parsed['target_text'] = params[5][1:-1]  # 따옴표 제거
            else:
                parsed['target_text'] = params[5]
            
            # 추가 옵션 파라미터 처리
            param_idx = 6
            parsed['exact_match'] = False  # 기본값: 일부 포함
            parsed['max_tries'] = 10       # 기본값: 10회
            parsed['coord_mode'] = 'scaled'  # 기본값: 스케일링
            
            # exact_match 옵션 확인 (exact 또는 contains)
            if len(params) > param_idx:
                if params[param_idx].lower() in ['exact', 'true', '1']:
                    parsed['exact_match'] = True
                    param_idx += 1
                elif params[param_idx].lower() in ['contains', 'false', '0']:
                    parsed['exact_match'] = False
                    param_idx += 1
            
            # max_tries 확인
            if len(params) > param_idx:
                try:
                    parsed['max_tries'] = int(params[param_idx])
                    param_idx += 1
                except ValueError:
                    # 숫자가 아니면 좌표 모드일 수 있음
                    pass
                    
            # coord_mode 확인 (마지막 파라미터)
            if len(params) > param_idx:
                if params[param_idx].lower() in ['offset', 'scaled']:
                    parsed['coord_mode'] = params[param_idx].lower()
                    
            return parsed
        except (ValueError, IndexError):
            return {}
    
    def set_ui_values(self, params):
        if not params:
            return
            
        self.x_input.setValue(params.get('x', 0))
        self.y_input.setValue(params.get('y', 0))
        self.width_input.setValue(params.get('width', 100))
        self.height_input.setValue(params.get('height', 50))
        
        ocr_type = params.get('ocr_type', 'i2s')
        if ocr_type == 'i2skr':
            self.ocr_combo.setCurrentIndex(1)
        else:
            self.ocr_combo.setCurrentIndex(0)
            
        self.text_input.setText(params.get('target_text', ''))
        
        # 매칭 모드 설정 (완전일치면 1, 일부포함이면 0)
        exact_match = params.get('exact_match', False)
        self.match_mode_combo.setCurrentIndex(1 if exact_match else 0)
        
        self.max_tries_input.setValue(params.get('max_tries', 10))
        
        # 좌표 모드 설정
        coord_mode = params.get('coord_mode', 'scaled')
        if coord_mode == 'offset':
            self.coord_mode_combo.setCurrentIndex(1)
        else:
            self.coord_mode_combo.setCurrentIndex(0)
    
    def get_command_string(self):
        ocr_type = 'i2skr' if self.ocr_combo.currentIndex() == 1 else 'i2s'
        target_text = f'"{self.text_input.text()}"'  # 텍스트를 따옴표로 묶음
        match_mode = 'exact' if self.match_mode_combo.currentIndex() == 1 else 'contains'
        coord_mode = 'offset' if self.coord_mode_combo.currentIndex() == 1 else 'scaled'
        return f"waituntil {self.x_input.value()} {self.y_input.value()} {self.width_input.value()} {self.height_input.value()} {ocr_type} {target_text} {match_mode} {self.max_tries_input.value()} {coord_mode}"
    
    def execute(self, params, window_coords=None, processor_state=None):
        if not params or 'target_text' not in params:
            print("오류: waituntil 명령어에 필요한 파라미터가 없습니다.")
            return
            
        x = params.get('x', 0)
        y = params.get('y', 0) 
        width = params.get('width', 100)
        height = params.get('height', 50)
        ocr_type = params.get('ocr_type', 'i2s')
        target_text = params.get('target_text', '')
        max_tries = params.get('max_tries', 10)
        exact_match = params.get('exact_match', False)
        
        # 좌표 보정
        if window_coords:
            coord_mode = params.get('coord_mode', 'scaled')
            
            # 좌표 모드에 따라 다른 계산 함수 사용
            if coord_mode == 'offset':
                adjusted_coords = calculate_offset_coordinates(x, y, window_coords)
            else:  # 'scaled'
                adjusted_coords = calculate_adjusted_coordinates(x, y, window_coords)
            x, y = adjusted_coords
        
        match_mode_text = "완전일치" if exact_match else "일부포함"
        print(f"'{target_text}' 텍스트가 나타날 때까지 대기 중... (매칭모드: {match_mode_text}, 최대 {max_tries}회 시도)")
        
        for i in range(max_tries):
            # 중지 플래그 체크 (각 반복 시작 시)
            processor = params.get('processor') if params else None
            if processor and hasattr(processor, 'stop_flag') and processor.stop_flag:
                print(f"⚠️ WaitUntil 중지됨 ({i}/{max_tries}번째 시도)")
                return
            
            # processor_state에서 중지 요청 확인
            if processor_state and processor_state.get('stop_requested', False):
                print(f"⚠️ WaitUntil 중지됨 (state 플래그, {i}/{max_tries}번째 시도)")
                return
            
            try:
                # 스크린샷 촬영
                screenshot_path = take_screenshot_with_coords(x, y, width, height)
                if not screenshot_path:
                    print(f"[{i+1}/{max_tries}] 스크린샷 촬영 실패")
                    # 중단 가능한 1초 대기
                    if self._interruptible_sleep(1, params, f"screenshot retry wait ({i+1}/{max_tries})"):
                        return
                    continue
                
                # OCR 실행
                if ocr_type == 'i2s':
                    extracted_text = image_to_text(screenshot_path, lang='eng')
                elif ocr_type == 'i2skr':
                    extracted_text = image_to_text(screenshot_path, lang='kor')
                else:
                    print(f"지원하지 않는 OCR 타입: {ocr_type}")
                    return
                
                print(f"[{i+1}/{max_tries}] OCR 결과: {extracted_text}")
                
                # processor_state에 결과 저장
                if processor_state is not None:
                    processor_state['screenshot_path'] = screenshot_path
                    processor_state['extracted_text'] = extracted_text
                
                # 타겟 텍스트 확인 (완전일치 vs 일부포함)
                match_found = False
                if exact_match:
                    # 완전일치: OCR 결과가 타겟 텍스트와 정확히 일치하는지 확인
                    match_found = extracted_text.strip() == target_text.strip()
                    match_type = "완전일치"
                else:
                    # 일부포함: OCR 결과에 타겟 텍스트가 포함되어 있는지 확인
                    match_found = target_text in extracted_text
                    match_type = "일부포함"
                
                if match_found:
                    print(f"✓ '{target_text}' 텍스트를 찾았습니다! ({match_type}, {i+1}번째 시도)")
                    return
                
                print(f"[{i+1}/{max_tries}] '{target_text}' 텍스트를 찾지 못했습니다. 1초 후 재시도...")
                
            except Exception as e:
                print(f"[{i+1}/{max_tries}] 오류 발생: {e}")
            
            # 중단 가능한 1초 대기
            if self._interruptible_sleep(1, params, f"waituntil retry ({i+1}/{max_tries})"):
                return
        
        print(f"✗ 타임아웃: {max_tries}회 시도 후에도 '{target_text}' 텍스트를 찾지 못했습니다. (매칭모드: {match_mode_text})")
    
    def on_get_coordinates(self):
        """드래그로 영역 선택하여 좌표 설정"""
        from PyQt5.QtWidgets import QApplication, QRubberBand
        from PyQt5.QtCore import QRect, QPoint, QSize
        
        class DragSelector(QWidget):
            def __init__(self):
                super().__init__()
                self.setWindowFlag(Qt.FramelessWindowHint)
                self.setAttribute(Qt.WA_TransparentForMouseEvents, False)
                self.setWindowOpacity(0.15)
                self.screen_geo = QApplication.primaryScreen().geometry()
                self.setGeometry(self.screen_geo)
                self.rubber_band = QRubberBand(QRubberBand.Rectangle, self)
                self.origin = QPoint()
                self.setCursor(Qt.CrossCursor)
                self.rect_result = None

            def mousePressEvent(self, event):
                self.origin = event.pos()
                self.rubber_band.setGeometry(QRect(self.origin, QSize()))
                self.rubber_band.show()

            def mouseMoveEvent(self, event):
                self.rubber_band.setGeometry(QRect(self.origin, event.pos()).normalized())

            def mouseReleaseEvent(self, event):
                r = self.rubber_band.geometry().getRect()
                self.rect_result = r
                self.rubber_band.hide()
                self.close()

            def get_rect(self):
                return self.rect_result

        app = QApplication.instance()
        drag_selector = DragSelector()
        drag_selector.setWindowOpacity(0.20)
        drag_selector.setWindowModality(Qt.ApplicationModal)
        drag_selector.show()

        while drag_selector.isVisible():
            app.processEvents()

        rect = drag_selector.get_rect()
        if rect:
            x, y, w, h = rect
            
            # 선택된 윈도우 기준으로 상대 좌표 계산
            try:
                import pygetwindow as gw
                from PyQt5.QtWidgets import QApplication
                
                app = QApplication.instance()
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'window_dropdown'):
                        selected_window = widget.window_dropdown.currentText()
                        if selected_window:
                            windows = gw.getWindowsWithTitle(selected_window)
                            if windows:
                                window = windows[0]
                                win_x, win_y = window.left, window.top
                                
                                # 상대 좌표 계산
                                rel_x = x - win_x
                                rel_y = y - win_y
                                
                                self.x_input.setValue(rel_x)
                                self.y_input.setValue(rel_y)
                                self.width_input.setValue(w)
                                self.height_input.setValue(h)
                                print(f"상대 좌표가 설정되었습니다: ({rel_x}, {rel_y}, {w}, {h}) | 절대 좌표: ({x}, {y}, {w}, {h})")
                                return
                
                # 윈도우를 찾지 못했을 때 절대 좌표 사용
                self.x_input.setValue(x)
                self.y_input.setValue(y)
                self.width_input.setValue(w)
                self.height_input.setValue(h)
                print(f"윈도우 미선택. 절대 좌표가 설정되었습니다: ({x}, {y}, {w}, {h})")
                
            except Exception as e:
                # 오류 발생시 절대 좌표 사용
                self.x_input.setValue(x)
                self.y_input.setValue(y)
                self.width_input.setValue(w)
                self.height_input.setValue(h)
                print(f"상대 좌표 계산 실패. 절대 좌표가 설정되었습니다: ({x}, {y}, {w}, {h}) - {e}")
    
    def on_test_ocr(self):
        """현재 설정으로 OCR 테스트 - 통합 함수 사용"""
        try:
            import time
            total_start = time.time()
            
            x = self.x_input.value()
            y = self.y_input.value()
            width = self.width_input.value()
            height = self.height_input.value()
            ocr_type = 'i2skr' if self.ocr_combo.currentIndex() == 1 else 'i2s'
            target_text = self.text_input.text().strip()
            exact_match = self.match_mode_combo.currentIndex() == 1
            
            # 상대 좌표를 절대 좌표로 변환 후 스크린샷 촬영
            try:
                import pygetwindow as gw
                from PyQt5.QtWidgets import QApplication
                
                # 선택된 윈도우 정보 가져오기
                app = QApplication.instance()
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'window_dropdown'):
                        selected_window = widget.window_dropdown.currentText()
                        if selected_window:
                            windows = gw.getWindowsWithTitle(selected_window)
                            if windows:
                                window = windows[0]
                                win_x, win_y = window.left, window.top
                                
                                # 상대 좌표를 절대 좌표로 변환
                                abs_x = x + win_x
                                abs_y = y + win_y
                                
                                screenshot_path = take_screenshot_with_coords(abs_x, abs_y, width, height)
                                break
                        else:
                            # 윈도우가 선택되지 않았을 때는 절대 좌표로 처리
                            screenshot_path = take_screenshot_with_coords(x, y, width, height)
                            break
                else:
                    # 메인 앱을 찾지 못했을 때
                    screenshot_path = take_screenshot_with_coords(x, y, width, height)
                    
            except Exception:
                # 오류 발생시 절대 좌표로 처리
                screenshot_path = take_screenshot_with_coords(x, y, width, height)
            
            if not screenshot_path:
                QMessageBox.warning(None, "테스트 오류", "스크린샷 촬영에 실패했습니다.")
                return
            
            # OCR 실행 (expected_text를 전달하여 시도 정보를 _last_ocr_attempts에 저장)
            if ocr_type == 'i2s':
                extracted_text = image_to_text(screenshot_path, lang='eng', expected_text=target_text, exact_match=exact_match)
                ocr_lang = "영어"
            else:
                extracted_text = image_to_text(screenshot_path, lang='kor', expected_text=target_text, exact_match=exact_match)
                ocr_lang = "한국어"
            
            if not extracted_text:
                extracted_text = ""
            
            total_time = time.time() - total_start
            
            # 통합 다이얼로그 호출
            show_unified_ocr_test_dialog(
                x, y, width, height,
                screenshot_path,
                ocr_lang,
                extracted_text=extracted_text,  # 실제 추출된 텍스트 전달
                expected_text=target_text,
                exact_match=exact_match,
                ocr_attempts=None,  # _last_ocr_attempts 사용
                total_time=total_time
            )
            
        except Exception as e:
            QMessageBox.critical(None, "OCR 테스트 오류", f"테스트 중 오류가 발생했습니다:\n{e}")


class MouseWheelCommand(CommandBase):
    """마우스 휠 조작 명령어"""
    
    @property
    def name(self):
        return "MouseWheel"
    
    @property 
    def description(self):
        return "마우스 휠을 조작하여 스크롤 (방향, 강도 설정 가능)"
    
    def create_ui(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # 휠 방향 선택
        direction_layout = QHBoxLayout()
        direction_layout.addWidget(QLabel('휠 방향:'))
        self.direction_combo = QComboBox()
        self.direction_combo.addItems(['Up (위로)', 'Down (아래로)'])
        direction_layout.addWidget(self.direction_combo)
        layout.addLayout(direction_layout)
        
        # 휠 강도 (스크롤 횟수)
        strength_layout = QHBoxLayout()
        strength_layout.addWidget(QLabel('스크롤 횟수:'))
        self.strength_input = QSpinBox()
        self.strength_input.setRange(1, 50)
        self.strength_input.setValue(3)
        self.strength_input.setSuffix('회')
        strength_layout.addWidget(self.strength_input)
        layout.addLayout(strength_layout)
        
        # 마우스 위치 설정
        position_layout = QHBoxLayout()
        position_layout.addWidget(QLabel('마우스 위치:'))
        self.position_combo = QComboBox()
        self.position_combo.addItems(['현재 위치', '지정 좌표'])
        self.position_combo.currentTextChanged.connect(self.on_position_changed)
        position_layout.addWidget(self.position_combo)
        layout.addLayout(position_layout)
        
        # 좌표 입력 (지정 좌표 선택 시에만 활성화)
        coord_layout = QHBoxLayout()
        coord_layout.addWidget(QLabel('X:'))
        self.x_input = QSpinBox()
        self.x_input.setRange(-9999, 9999)
        self.x_input.setValue(0)
        self.x_input.setEnabled(False)
        coord_layout.addWidget(self.x_input)
        
        coord_layout.addWidget(QLabel('Y:'))
        self.y_input = QSpinBox()
        self.y_input.setRange(-9999, 9999)
        self.y_input.setValue(0)
        self.y_input.setEnabled(False)
        coord_layout.addWidget(self.y_input)
        layout.addLayout(coord_layout)
        
        # 좌표 모드 선택
        coord_mode_layout = QHBoxLayout()
        coord_mode_layout.addWidget(QLabel('좌표 모드:'))
        self.coord_mode_combo = QComboBox()
        self.coord_mode_combo.addItems(['스케일링 (기준해상도 기반)', '오프셋 (단순 위치이동)'])
        self.coord_mode_combo.setCurrentIndex(1)
        self.coord_mode_combo.setEnabled(False)
        coord_mode_layout.addWidget(self.coord_mode_combo)
        layout.addLayout(coord_mode_layout)
        
        # 대기 시간
        delay_layout = QHBoxLayout()
        delay_layout.addWidget(QLabel('실행 후 대기:'))
        self.delay_input = QSpinBox()
        self.delay_input.setRange(0, 5000)
        self.delay_input.setValue(500)
        self.delay_input.setSuffix('ms')
        delay_layout.addWidget(self.delay_input)
        layout.addLayout(delay_layout)
        
        widget.setLayout(layout)
        return widget
    
    def on_position_changed(self, text):
        """위치 모드 변경 시 호출"""
        is_custom = (text == '지정 좌표')
        self.x_input.setEnabled(is_custom)
        self.y_input.setEnabled(is_custom)
        self.coord_mode_combo.setEnabled(is_custom)
    
    def parse_params(self, params):
        if len(params) < 2:
            return {}
        
        try:
            parsed = {
                'direction': params[0],  # 'up' or 'down'
                'strength': int(params[1]),  # 스크롤 횟수
                'position_mode': 'current',  # 기본값
                'x': 0,
                'y': 0,
                'coord_mode': 'scaled',
                'delay': 500
            }
            
            # 추가 파라미터들 (선택적)
            if len(params) > 2:
                parsed['position_mode'] = params[2]  # 'current' or 'custom'
            
            if len(params) > 3:
                parsed['x'] = int(params[3])
            
            if len(params) > 4:
                parsed['y'] = int(params[4])
            
            if len(params) > 5:
                parsed['coord_mode'] = params[5]
            
            if len(params) > 6:
                parsed['delay'] = int(params[6])
            
            return parsed
            
        except (ValueError, IndexError):
            return {}
    
    def set_ui_values(self, params):
        if not params:
            return
        
        # 방향 설정
        direction = params.get('direction', 'up')
        if direction.lower() == 'down':
            self.direction_combo.setCurrentIndex(1)
        else:
            self.direction_combo.setCurrentIndex(0)
        
        # 강도 설정
        self.strength_input.setValue(params.get('strength', 3))
        
        # 위치 모드 설정
        position_mode = params.get('position_mode', 'current')
        if position_mode == 'custom':
            self.position_combo.setCurrentIndex(1)
        else:
            self.position_combo.setCurrentIndex(0)
        
        # 좌표 설정
        self.x_input.setValue(params.get('x', 0))
        self.y_input.setValue(params.get('y', 0))
        
        # 좌표 모드 설정
        coord_mode = params.get('coord_mode', 'scaled')
        if coord_mode == 'offset':
            self.coord_mode_combo.setCurrentIndex(1)
        else:
            self.coord_mode_combo.setCurrentIndex(0)
        
        # 대기 시간 설정
        self.delay_input.setValue(params.get('delay', 500))
    
    def get_command_string(self):
        direction = 'down' if self.direction_combo.currentIndex() == 1 else 'up'
        strength = self.strength_input.value()
        
        if self.position_combo.currentIndex() == 1:  # 지정 좌표
            position_mode = 'custom'
            x = self.x_input.value()
            y = self.y_input.value()
            coord_mode = 'offset' if self.coord_mode_combo.currentIndex() == 1 else 'scaled'
            delay = self.delay_input.value()
            return f"mousewheel {direction} {strength} {position_mode} {x} {y} {coord_mode} {delay}"
        else:  # 현재 위치
            delay = self.delay_input.value()
            return f"mousewheel {direction} {strength} current 0 0 scaled {delay}"
    
    def execute(self, params, window_coords=None, processor_state=None):
        if not params or 'direction' not in params or 'strength' not in params:
            print("오류: mousewheel 명령어에 필요한 파라미터가 없습니다.")
            return
        
        direction = params.get('direction', 'up').lower()
        strength = params.get('strength', 3)
        position_mode = params.get('position_mode', 'current')
        x = params.get('x', 0)
        y = params.get('y', 0)
        coord_mode = params.get('coord_mode', 'scaled')
        delay = params.get('delay', 500)
        
        try:
            import pyautogui as pag
            
            # 마우스 위치 설정
            if position_mode == 'custom':
                # 좌표 보정
                if window_coords:
                    if coord_mode == 'offset':
                        adjusted_coords = calculate_offset_coordinates(x, y, window_coords)
                    else:  # 'scaled'
                        adjusted_coords = calculate_adjusted_coordinates(x, y, window_coords)
                    final_x, final_y = adjusted_coords
                else:
                    final_x, final_y = x, y
                
                # 마우스를 지정 위치로 이동
                pag.moveTo(final_x, final_y)
                print(f"마우스를 ({final_x}, {final_y})로 이동")
            else:
                final_x, final_y = pag.position()
                print(f"현재 마우스 위치: ({final_x}, {final_y})")
            
            # 휠 스크롤 실행
            scroll_direction = strength if direction == 'up' else -strength
            pag.scroll(scroll_direction)
            
            direction_text = "위로" if direction == 'up' else "아래로"
            print(f"마우스 휠 {direction_text} {strength}회 스크롤 완료")
            
            # 대기 시간
            if delay > 0:
                time.sleep(delay / 1000.0)  # ms to seconds
                print(f"{delay}ms 대기 완료")
            
        except Exception as e:
            print(f"마우스 휠 조작 중 오류 발생: {e}")


class TestTextCommand(CommandBase):
    """텍스트 추출 기반 Pass/Fail 판별 명령어"""
    
    @property
    def name(self): 
        return "TestText"
    
    @property
    def description(self): 
        return "스크린샷에서 텍스트를 추출하여 기대값과 비교해 Pass/Fail 판별"
    
    def _interruptible_sleep(self, duration, params, context=""):
        """중지 플래그를 체크하면서 대기하는 함수
        
        Args:
            duration: 대기 시간 (초)
            params: 파라미터 딕셔너리 (processor_stop_flag 체크용)
            context: 대기 컨텍스트 (디버깅용)
        
        Returns:
            bool: True if interrupted (중지됨), False if completed (완료됨)
        """
        if duration <= 0:
            return False
            
        # 0.1초 간격으로 중지 플래그 체크
        check_interval = 0.1
        total_slept = 0
        
        while total_slept < duration:
            # params에서 CommandProcessor의 실시간 stop_flag 체크
            processor = params.get('processor') if params else None
            if processor and hasattr(processor, 'stop_flag') and processor.stop_flag:
                print(f"⚠️ TestText {context} 중지됨 (경과시간: {total_slept:.1f}초/{duration}초)")
                return True  # 중지됨
            
            sleep_time = min(check_interval, duration - total_slept)
            time.sleep(sleep_time)
            total_slept += sleep_time
            
        return False  # 완료됨
    
    def create_ui(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Title 입력 (체크리스트용)
        title_layout = QHBoxLayout()
        title_layout.addWidget(QLabel('Title:'))
        self.title_input = QLineEdit()
        self.title_input.setPlaceholderText("체크리스트에 표시될 제목 (예: 라이브 버튼 출력 확인)")
        title_layout.addWidget(self.title_input)
        layout.addLayout(title_layout)
        
        # 좌표 입력
        coord_layout = QHBoxLayout()
        coord_layout.addWidget(QLabel('상대 X:'))
        self.x_input = QSpinBox()
        self.x_input.setRange(-9999, 9999)  # 상대 좌표는 음수일 수 있음
        coord_layout.addWidget(self.x_input)
        
        coord_layout.addWidget(QLabel('상대 Y:'))
        self.y_input = QSpinBox()
        self.y_input.setRange(-9999, 9999)  # 상대 좌표는 음수일 수 있음
        coord_layout.addWidget(self.y_input)
        
        coord_layout.addWidget(QLabel('Width:'))
        self.width_input = QSpinBox()
        self.width_input.setRange(1, 9999)
        self.width_input.setValue(100)
        coord_layout.addWidget(self.width_input)
        
        coord_layout.addWidget(QLabel('Height:'))
        self.height_input = QSpinBox()
        self.height_input.setRange(1, 9999)
        self.height_input.setValue(50)
        coord_layout.addWidget(self.height_input)
        
        layout.addLayout(coord_layout)
        
        # 좌표 선택 및 테스트 버튼
        coord_btn_layout = QHBoxLayout()
        self.get_coord_btn = QPushButton('좌표 선택 (드래그)')
        self.get_coord_btn.clicked.connect(self.on_get_coordinates)
        self.test_ocr_btn = QPushButton('OCR 테스트')
        self.test_ocr_btn.clicked.connect(self.on_test_ocr)
        coord_btn_layout.addWidget(self.get_coord_btn)
        coord_btn_layout.addWidget(self.test_ocr_btn)
        layout.addLayout(coord_btn_layout)
        
        # OCR 언어 선택
        ocr_layout = QHBoxLayout()
        ocr_layout.addWidget(QLabel('OCR:'))
        self.ocr_combo = QComboBox()
        self.ocr_combo.addItems(['i2s (English)', 'i2skr (Korean)'])
        ocr_layout.addWidget(self.ocr_combo)
        layout.addLayout(ocr_layout)
        
        # 기대 텍스트 입력
        text_layout = QHBoxLayout()
        text_layout.addWidget(QLabel('기대 텍스트:'))
        self.text_input = QLineEdit()
        self.text_input.setPlaceholderText("기대되는 텍스트를 입력하세요 (예: 라이브)")
        text_layout.addWidget(self.text_input)
        layout.addLayout(text_layout)
        
        # 매칭 모드 선택
        match_layout = QHBoxLayout()
        match_layout.addWidget(QLabel('매칭 모드:'))
        self.match_mode_combo = QComboBox()
        self.match_mode_combo.addItems(['일부 포함', '완전 일치'])
        match_layout.addWidget(self.match_mode_combo)
        layout.addLayout(match_layout)
        
        # 좌표 모드 선택
        coord_mode_layout = QHBoxLayout()
        coord_mode_layout.addWidget(QLabel('좌표 모드:'))
        self.coord_mode_combo = QComboBox()
        self.coord_mode_combo.addItems(['스케일링 (기준해상도 기반)', '오프셋 (단순 위치이동)'])
        self.coord_mode_combo.setCurrentIndex(1)  # 기본값: 스케일링
        coord_mode_layout.addWidget(self.coord_mode_combo)
        layout.addLayout(coord_mode_layout)
        
        # 반복 확인 옵션
        repeat_layout = QHBoxLayout()
        repeat_layout.addWidget(QLabel('반복 확인:'))
        self.repeat_checkbox = QCheckBox('활성화')
        self.repeat_checkbox.toggled.connect(self.on_repeat_toggled)
        repeat_layout.addWidget(self.repeat_checkbox)
        
        repeat_layout.addWidget(QLabel('최대 시도:'))
        self.max_tries_input = QSpinBox()
        self.max_tries_input.setRange(1, 10000)
        self.max_tries_input.setValue(10)
        self.max_tries_input.setSuffix('회')
        self.max_tries_input.setEnabled(False)  # 기본적으로 비활성화
        repeat_layout.addWidget(self.max_tries_input)
        
        repeat_layout.addWidget(QLabel('대기 시간:'))
        self.wait_interval_input = QSpinBox()
        self.wait_interval_input.setRange(1, 60)
        self.wait_interval_input.setValue(1)
        self.wait_interval_input.setSuffix('초')
        self.wait_interval_input.setEnabled(False)  # 기본적으로 비활성화
        repeat_layout.addWidget(self.wait_interval_input)
        
        layout.addLayout(repeat_layout)
        
        widget.setLayout(layout)
        return widget
    
    def on_repeat_toggled(self, checked):
        """반복 확인 체크박스 토글 시 호출"""
        self.max_tries_input.setEnabled(checked)
        self.wait_interval_input.setEnabled(checked)
    
    def parse_params(self, params):
        # 전체 명령어 문자열 재구성
        full_command = 'testtext ' + ' '.join(params)
        print(f"testtext 전체 명령어: {full_command}")
        
        try:
            import re
            # 정규식으로 파라미터 추출: testtext "title" x y width height ocr_type "expected_text" [match_mode] [coord_mode] [repeat_mode] [max_tries] [wait_interval]
            pattern = r'testtext\s+"([^"]+)"\s+(-?\d+)\s+(-?\d+)\s+(\d+)\s+(\d+)\s+(\w+)\s+"([^"]*)"\s*(\w*)\s*(\w*)\s*(\w*)\s*(\d*)\s*(\d*)'
            match = re.match(pattern, full_command)
            
            if not match:
                print(f"testtext 정규식 매칭 실패")
                print(f"기대 형식: testtext \"title\" x y width height ocr_type \"expected_text\" [match_mode] [coord_mode]")
                print(f"입력된 명령어: {full_command}")
                return {}
            
            groups = match.groups()
            parsed = {
                'title': groups[0],
                'x': int(groups[1]),
                'y': int(groups[2]),
                'width': int(groups[3]),
                'height': int(groups[4]),
                'ocr_type': groups[5],
                'expected_text': groups[6],
                'exact_match': False,  # 기본값
                'coord_mode': 'scaled',  # 기본값
                'repeat_mode': False,  # 기본값
                'max_tries': 10,  # 기본값
                'wait_interval': 1  # 기본값
            }
            
            # match_mode 처리 (선택적)
            if len(groups) > 7 and groups[7]:
                match_mode = groups[7].lower()
                if match_mode in ['exact', 'true', '1']:
                    parsed['exact_match'] = True
                    
            # coord_mode 처리 (선택적)
            if len(groups) > 8 and groups[8]:
                coord_mode = groups[8].lower()
                if coord_mode in ['offset', 'scaled']:
                    parsed['coord_mode'] = coord_mode
            
            # repeat_mode 처리 (선택적)
            if len(groups) > 9 and groups[9]:
                repeat_mode = groups[9].lower()
                if repeat_mode in ['repeat', 'true', '1']:
                    parsed['repeat_mode'] = True
            
            # max_tries 처리 (선택적)
            if len(groups) > 10 and groups[10]:
                try:
                    parsed['max_tries'] = int(groups[10])
                except ValueError:
                    pass
            
            # wait_interval 처리 (선택적)
            if len(groups) > 11 and groups[11]:
                try:
                    parsed['wait_interval'] = int(groups[11])
                except ValueError:
                    pass
            
            print(f"testtext 파싱 성공: {parsed}")
            return parsed
            
        except (ValueError, AttributeError) as e:
            print(f"testtext 파싱 오류: {e}")
            print(f"입력 파라미터: {params}")
            return {}
    
    def set_ui_values(self, params):
        if not params:
            return
            
        self.title_input.setText(params.get('title', ''))
        self.x_input.setValue(params.get('x', 0))
        self.y_input.setValue(params.get('y', 0))
        self.width_input.setValue(params.get('width', 100))
        self.height_input.setValue(params.get('height', 50))
        
        ocr_type = params.get('ocr_type', 'i2s')
        if ocr_type == 'i2skr':
            self.ocr_combo.setCurrentIndex(1)
        else:
            self.ocr_combo.setCurrentIndex(0)
            
        self.text_input.setText(params.get('expected_text', ''))
        
        # 매칭 모드 설정 (완전일치면 1, 일부포함이면 0)
        exact_match = params.get('exact_match', False)
        self.match_mode_combo.setCurrentIndex(1 if exact_match else 0)
        
        # 좌표 모드 설정
        coord_mode = params.get('coord_mode', 'scaled')
        if coord_mode == 'offset':
            self.coord_mode_combo.setCurrentIndex(1)
        else:
            self.coord_mode_combo.setCurrentIndex(0)
        
        # 반복 모드 설정
        repeat_mode = params.get('repeat_mode', False)
        self.repeat_checkbox.setChecked(repeat_mode)
        
        # 최대 시도 횟수
        max_tries = params.get('max_tries', 10)
        self.max_tries_input.setValue(max_tries)
        
        # 대기 시간
        wait_interval = params.get('wait_interval', 1)
        self.wait_interval_input.setValue(wait_interval)
    
    def get_command_string(self):
        ocr_type = 'i2skr' if self.ocr_combo.currentIndex() == 1 else 'i2s'
        title = f'"{self.title_input.text()}"'  # 제목을 따옴표로 묶음
        expected_text = f'"{self.text_input.text()}"'  # 텍스트를 따옴표로 묶음
        match_mode = 'exact' if self.match_mode_combo.currentIndex() == 1 else 'contains'
        coord_mode = 'offset' if self.coord_mode_combo.currentIndex() == 1 else 'scaled'
        
        # 반복 모드 처리
        command_str = f"testtext {title} {self.x_input.value()} {self.y_input.value()} {self.width_input.value()} {self.height_input.value()} {ocr_type} {expected_text} {match_mode} {coord_mode}"
        
        if self.repeat_checkbox.isChecked():
            repeat_mode = 'repeat'
            max_tries = self.max_tries_input.value()
            wait_interval = self.wait_interval_input.value()
            command_str += f" {repeat_mode} {max_tries} {wait_interval}"
        
        return command_str
    
    def execute(self, params, window_coords=None, processor_state=None):
        if not params or 'expected_text' not in params or 'title' not in params:
            print("오류: testtext 명령어에 필요한 파라미터가 없습니다.")
            return
            
        title = params.get('title', '테스트 항목')
        x = params.get('x', 0)
        y = params.get('y', 0) 
        width = params.get('width', 100)
        height = params.get('height', 50)
        ocr_type = params.get('ocr_type', 'i2s')
        expected_text = params.get('expected_text', '')
        exact_match = params.get('exact_match', False)
        
        # 반복 확인 설정
        repeat_mode = params.get('repeat_mode', False)
        max_tries = params.get('max_tries', 10)
        wait_interval = params.get('wait_interval', 1)
        
        # 좌표 보정
        if window_coords:
            coord_mode = params.get('coord_mode', 'scaled')
            
            # 좌표 모드에 따라 다른 계산 함수 사용
            if coord_mode == 'offset':
                adjusted_coords = calculate_offset_coordinates(x, y, window_coords)
            else:  # 'scaled'
                adjusted_coords = calculate_adjusted_coordinates(x, y, window_coords)
            x, y = adjusted_coords
        
        match_mode_text = "완전일치" if exact_match else "일부포함"
        print(f"테스트 실행: {title} - 기대텍스트: '{expected_text}' (매칭모드: {match_mode_text})")
        
        # 반복 실행 로직
        final_result = None
        current_try = 0
        max_attempts = max_tries if repeat_mode else 1
        
        while current_try < max_attempts:
            # 중지 플래그 체크 (각 반복 시작 시)
            processor = params.get('processor') if params else None
            if processor and hasattr(processor, 'stop_flag') and processor.stop_flag:
                print(f"⚠️ testtext 중지됨 ({current_try}/{max_attempts}번째 시도)")
                return
            
            current_try += 1
            
            if repeat_mode:
                print(f"[{current_try}/{max_attempts}] 텍스트 검사 시도 중...")
            
            try:
                # 스크린샷 촬영
                screenshot_path = take_screenshot_with_coords(x, y, width, height)
                if not screenshot_path:
                    print("스크린샷 촬영 실패")
                    if not repeat_mode:
                        return
                    # 중지 플래그를 체크하면서 대기
                    if self._interruptible_sleep(wait_interval, params, "screenshot retry"):
                        return
                    continue
                
                # OCR 실행 (조기 종료 최적화: expected_text와 exact_match 전달)
                if ocr_type == 'i2s':
                    extracted_text = image_to_text(screenshot_path, lang='eng', expected_text=expected_text, exact_match=exact_match)
                elif ocr_type == 'i2skr':
                    extracted_text = image_to_text(screenshot_path, lang='kor', expected_text=expected_text, exact_match=exact_match)
                else:
                    print(f"지원하지 않는 OCR 타입: {ocr_type}")
                    return
                
                if not extracted_text:
                    extracted_text = ""
                
                print(f"OCR 결과: '{extracted_text}'")
                
                # 텍스트 매칭 확인
                match_found = False
                if exact_match:
                    # 완전일치: OCR 결과가 기대 텍스트와 정확히 일치하는지 확인
                    match_found = extracted_text.strip() == expected_text.strip()
                    match_type = "완전일치"
                else:
                    # 일부포함: OCR 결과에 기대 텍스트가 포함되어 있는지 확인
                    match_found = expected_text in extracted_text
                    match_type = "일부포함"
                
                # 결과 판별
                result = "Pass" if match_found else "Fail"
                
                # 성공하면 루프 종료
                if match_found:
                    print(f"✓ {result}: '{expected_text}' 텍스트를 찾았습니다! ({match_type})")
                    final_result = {
                        'title': title,
                        'expected_text': expected_text,
                        'extracted_text': extracted_text,
                        'result': result,
                        'screenshot_path': screenshot_path,
                        'match_mode': match_type,
                        'attempt': current_try
                    }
                    break
                else:
                    if repeat_mode and current_try < max_attempts:
                        print(f"✗ '{expected_text}' 텍스트를 찾지 못했습니다. {wait_interval}초 후 재시도...")
                        # 중지 플래그를 체크하면서 대기
                        if self._interruptible_sleep(wait_interval, params, f"retry wait ({current_try}/{max_attempts})"):
                            return
                    else:
                        print(f"✗ {result}: '{expected_text}' 텍스트를 찾지 못했습니다. ({match_type})")
                        final_result = {
                            'title': title,
                            'expected_text': expected_text,
                            'extracted_text': extracted_text,
                            'result': result,
                            'screenshot_path': screenshot_path,
                            'match_mode': match_type,
                            'attempt': current_try
                        }
                
            except Exception as e:
                print(f"테스트 실행 중 오류 발생: {e}")
                if not repeat_mode or current_try >= max_attempts:
                    final_result = {
                        'title': title,
                        'expected_text': expected_text,
                        'extracted_text': f"오류: {str(e)}",
                        'result': "Fail",
                        'screenshot_path': screenshot_path if 'screenshot_path' in locals() else "N/A",
                        'match_mode': match_type if 'match_type' in locals() else "N/A",
                        'attempt': current_try
                    }
                    break
                else:
                    print(f"오류 발생, {wait_interval}초 후 재시도...")
                    # 중지 플래그를 체크하면서 대기
                    if self._interruptible_sleep(wait_interval, params, f"error retry wait ({current_try}/{max_attempts})"):
                        return
        
        # 최종 결과 처리
        if final_result:
            # processor_state에 결과 저장
            if processor_state is not None:
                if 'test_results' not in processor_state:
                    processor_state['test_results'] = []
                processor_state['test_results'].append(final_result)
            
            print(f"테스트 결과가 저장되었습니다: {title}")
            
            if repeat_mode:
                attempts_text = f" (총 {final_result.get('attempt', 1)}회 시도)"
                print(f"반복 테스트 완료{attempts_text}")
            
    
    def on_get_coordinates(self):
        """드래그로 영역 선택하여 좌표 설정 (WaitUntilCommand와 동일)"""
        from PyQt5.QtWidgets import QApplication, QRubberBand
        from PyQt5.QtCore import QRect, QPoint, QSize
        
        class DragSelector(QWidget):
            def __init__(self):
                super().__init__()
                self.setWindowFlag(Qt.FramelessWindowHint)
                self.setAttribute(Qt.WA_TransparentForMouseEvents, False)
                self.setWindowOpacity(0.15)
                self.screen_geo = QApplication.primaryScreen().geometry()
                self.setGeometry(self.screen_geo)
                self.rubber_band = QRubberBand(QRubberBand.Rectangle, self)
                self.origin = QPoint()
                self.setCursor(Qt.CrossCursor)
                self.rect_result = None

            def mousePressEvent(self, event):
                self.origin = event.pos()
                self.rubber_band.setGeometry(QRect(self.origin, QSize()))
                self.rubber_band.show()

            def mouseMoveEvent(self, event):
                self.rubber_band.setGeometry(QRect(self.origin, event.pos()).normalized())

            def mouseReleaseEvent(self, event):
                r = self.rubber_band.geometry().getRect()
                self.rect_result = r
                self.rubber_band.hide()
                self.close()

            def get_rect(self):
                return self.rect_result

        app = QApplication.instance()
        drag_selector = DragSelector()
        drag_selector.setWindowOpacity(0.20)
        drag_selector.setWindowModality(Qt.ApplicationModal)
        drag_selector.show()

        while drag_selector.isVisible():
            app.processEvents()

        rect = drag_selector.get_rect()
        if rect:
            x, y, w, h = rect
            
            # 선택된 윈도우 기준으로 상대 좌표 계산
            try:
                import pygetwindow as gw
                from PyQt5.QtWidgets import QApplication
                
                app = QApplication.instance()
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'window_dropdown'):
                        selected_window = widget.window_dropdown.currentText()
                        if selected_window:
                            windows = gw.getWindowsWithTitle(selected_window)
                            if windows:
                                window = windows[0]
                                win_x, win_y = window.left, window.top
                                
                                # 절대 좌표를 상대 좌표로 변환
                                rel_x = x - win_x
                                rel_y = y - win_y
                                
                                # UI에 상대 좌표 설정
                                self.x_input.setValue(rel_x)
                                self.y_input.setValue(rel_y)
                                self.width_input.setValue(w)
                                self.height_input.setValue(h)
                                print(f"상대 좌표가 설정되었습니다: ({rel_x}, {rel_y}, {w}, {h}) | 절대 좌표: ({x}, {y}, {w}, {h})")
                                return
                
                # 윈도우를 찾지 못했을 때 절대 좌표 사용
                self.x_input.setValue(x)
                self.y_input.setValue(y)
                self.width_input.setValue(w)
                self.height_input.setValue(h)
                print(f"윈도우 미선택. 절대 좌표가 설정되었습니다: ({x}, {y}, {w}, {h})")
                
            except Exception as e:
                # 오류 발생시 절대 좌표 사용
                self.x_input.setValue(x)
                self.y_input.setValue(y)
                self.width_input.setValue(w)
                self.height_input.setValue(h)
                print(f"상대 좌표 계산 실패. 절대 좌표가 설정되었습니다: ({x}, {y}, {w}, {h}) - {e}")
    
    def on_test_ocr(self):
        """현재 설정으로 OCR 테스트 - 통합 함수 사용 (WaitUntil과 동일)"""
        try:
            import time
            total_start = time.time()
            
            x = self.x_input.value()
            y = self.y_input.value()
            width = self.width_input.value()
            height = self.height_input.value()
            ocr_type = 'i2skr' if self.ocr_combo.currentIndex() == 1 else 'i2s'
            expected_text = self.text_input.text().strip()
            exact_match = self.match_mode_combo.currentIndex() == 1
            
            # 상대 좌표를 절대 좌표로 변환 후 스크린샷 촬영
            try:
                import pygetwindow as gw
                from PyQt5.QtWidgets import QApplication
                
                # 선택된 윈도우 정보 가져오기
                app = QApplication.instance()
                for widget in app.topLevelWidgets():
                    if hasattr(widget, 'window_dropdown'):
                        selected_window = widget.window_dropdown.currentText()
                        if selected_window:
                            windows = gw.getWindowsWithTitle(selected_window)
                            if windows:
                                window = windows[0]
                                win_x, win_y = window.left, window.top
                                
                                # 상대 좌표를 절대 좌표로 변환
                                abs_x = x + win_x
                                abs_y = y + win_y
                                
                                screenshot_path = take_screenshot_with_coords(abs_x, abs_y, width, height)
                                break
                        else:
                            # 윈도우가 선택되지 않았을 때는 절대 좌표로 처리
                            screenshot_path = take_screenshot_with_coords(x, y, width, height)
                            break
                else:
                    # 메인 앱을 찾지 못했을 때
                    screenshot_path = take_screenshot_with_coords(x, y, width, height)
                    
            except Exception:
                # 오류 발생시 절대 좌표로 처리
                screenshot_path = take_screenshot_with_coords(x, y, width, height)
            
            if not screenshot_path:
                QMessageBox.warning(None, "테스트 오류", "스크린샷 촬영에 실패했습니다.")
                return
            
            # OCR 실행 (expected_text를 전달하여 시도 정보를 _last_ocr_attempts에 저장)
            if ocr_type == 'i2s':
                extracted_text = image_to_text(screenshot_path, lang='eng', expected_text=expected_text, exact_match=exact_match)
                ocr_lang = "영어"
            else:
                extracted_text = image_to_text(screenshot_path, lang='kor', expected_text=expected_text, exact_match=exact_match)
                ocr_lang = "한국어"
            
            if not extracted_text:
                extracted_text = ""
            
            total_time = time.time() - total_start
            
            # 통합 다이얼로그 호출
            show_unified_ocr_test_dialog(
                x, y, width, height,
                screenshot_path,
                ocr_lang,
                extracted_text=extracted_text,  # 실제 추출된 텍스트 전달
                expected_text=expected_text,
                exact_match=exact_match,
                ocr_attempts=None,  # _last_ocr_attempts 사용
                total_time=total_time
            )
            
        except Exception as e:
            QMessageBox.critical(None, "OCR 테스트 오류", f"테스트 중 오류가 발생했습니다:\n{e}")


class ShowTestResultsCommand(CommandBase):
    """저장된 테스트 결과를 표시하는 명령어"""
    
    @property
    def name(self): 
        return "ShowResults"
    
    @property
    def description(self): 
        return "현재까지 저장된 testtext 결과들을 콘솔에 출력"
    
    def create_ui(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        info_label = QLabel("현재까지 저장된 테스트 결과를 콘솔에 출력합니다.")
        layout.addWidget(info_label)
        
        widget.setLayout(layout)
        return widget
    
    def parse_params(self, params):
        # 파라미터 없음
        return {}
    
    def set_ui_values(self, params):
        # UI 설정할 것 없음
        pass
    
    def get_command_string(self):
        return "showresults"
    
    def execute(self, params, window_coords=None, processor_state=None):
        from datetime import datetime
        
        print("\n" + "="*50)
        print("테스트 결과 요약")
        print("="*50)
        
        if processor_state is None or 'test_results' not in processor_state:
            print("저장된 테스트 결과가 없습니다.")
            return
        
        test_results = processor_state['test_results']
        if not test_results:
            print("저장된 테스트 결과가 없습니다.")
            return
        
        # 테스트 세션 정보
        test_title = processor_state.get('test_session_title', '알 수 없는 테스트')
        start_time = processor_state.get('test_session_start')
        
        print(f"📋 테스트 제목: {test_title}")
        
        if start_time:
            current_time = datetime.now()
            duration = current_time - start_time
            print(f"⏰ 시작 시간: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"⏰ 현재 시간: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"⏱️  소요 시간: {duration}")
        
        # 테스트 결과 요약
        total_tests = len(test_results)
        passed_tests = len([r for r in test_results if r['result'] == 'Pass'])
        failed_tests = total_tests - passed_tests
        
        print(f"\n📊 결과 요약:")
        print(f"   총 테스트: {total_tests}개")
        print(f"   Pass: {passed_tests}개")
        print(f"   Fail: {failed_tests}개")
        print(f"   성공률: {(passed_tests/total_tests*100):.1f}%")
        
        # Pass한 테스트 제목 표시
        if passed_tests > 0:
            passed_titles = [r['title'] for r in test_results if r['result'] == 'Pass']
            print(f"\n✅ 성공한 테스트:")
            for i, title in enumerate(passed_titles, 1):
                print(f"   {i}. {title}")
        
        # Fail한 테스트 제목 표시
        if failed_tests > 0:
            failed_titles = [r['title'] for r in test_results if r['result'] == 'Fail']
            print(f"\n❌ 실패한 테스트:")
            for i, title in enumerate(failed_titles, 1):
                print(f"   {i}. {title}")
        
        if passed_tests == 0 and failed_tests == 0:
            print(f"\n📋 테스트 결과가 없습니다.")
        
        print("="*50)


class ExportResultCommand(CommandBase):
    """테스트 결과를 다양한 형태로 내보내는 명령어 (엑셀, 텍스트, 슬랙알림)"""
    
    @property
    def name(self): 
        return "ExportResult"
    
    @property
    def description(self): 
        return "현재까지 저장된 testtext 결과들을 다양한 형태로 내보내기 (엑셀, 텍스트, 슬랙알림)"
    
    def create_ui(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        info_label = QLabel("현재까지 저장된 테스트 결과를 선택된 형태로 내보냅니다.")
        layout.addWidget(info_label)
        
        # 테스트 제목 입력
        title_layout = QHBoxLayout()
        title_layout.addWidget(QLabel('테스트 제목:'))
        self.title_input = QLineEdit()
        self.title_input.setPlaceholderText("예: 게임 로그인 테스트, UI 동작 검증 (띄어쓰기 가능)")
        title_layout.addWidget(self.title_input)
        layout.addLayout(title_layout)
        
        # 구분선
        separator1 = QLabel("─" * 60)
        separator1.setStyleSheet("color: gray;")
        layout.addWidget(separator1)
        
        # 엑셀 파일 관련 옵션
        excel_title = QLabel("📊 엑셀 파일 옵션")
        excel_title.setStyleSheet("font-weight: bold; color: #2E8B57;")
        layout.addWidget(excel_title)
        
        self.export_excel_checkbox = QCheckBox('엑셀 파일 생성')
        self.export_excel_checkbox.setChecked(True)  # 기본적으로 활성화
        layout.addWidget(self.export_excel_checkbox)
        
        # 이미지 포함 옵션 (엑셀 하위 옵션)
        image_layout = QHBoxLayout()
        image_layout.addSpacing(20)  # 들여쓰기
        self.include_images_checkbox = QCheckBox('엑셀에 스크린샷 이미지 포함 (체크 해제 시 안전 모드)')
        self.include_images_checkbox.setChecked(False)  # 기본값: 안전 모드
        self.include_images_checkbox.setToolTip('체크 해제하면 이미지 없이 안전하게 엑셀 파일을 생성합니다.')
        image_layout.addWidget(self.include_images_checkbox)
        layout.addLayout(image_layout)
        
        # 특정 파일명 지정 옵션 (엑셀 하위 옵션)
        filename_layout = QHBoxLayout()
        filename_layout.addSpacing(20)  # 들여쓰기
        filename_layout.addWidget(QLabel('특정 파일명:'))
        self.excel_filename_input = QLineEdit()
        self.excel_filename_input.setPlaceholderText("비워두면 타임스탬프 파일명으로 생성, 입력하면 해당 파일에 추가")
        self.excel_filename_input.setToolTip('파일명을 지정하면 해당 엑셀 파일이 있을 경우 이어서 추가하고, 없으면 새로 생성합니다.')
        filename_layout.addWidget(self.excel_filename_input)
        layout.addLayout(filename_layout)
        
        # 구분선
        separator2 = QLabel("─" * 60)
        separator2.setStyleSheet("color: gray;")
        layout.addWidget(separator2)
        
        # 텍스트 파일 관련 옵션
        text_title = QLabel("📄 텍스트 요약 파일 옵션")
        text_title.setStyleSheet("font-weight: bold; color: #4169E1;")
        layout.addWidget(text_title)
        
        self.export_text_checkbox = QCheckBox('텍스트 요약 파일 생성 (슬랙 전송용)')
        self.export_text_checkbox.setChecked(True)  # 기본적으로 활성화
        layout.addWidget(self.export_text_checkbox)
        
        # 구분선
        separator3 = QLabel("─" * 60)
        separator3.setStyleSheet("color: gray;")
        layout.addWidget(separator3)
        
        # 슬랙 알림 관련 옵션
        slack_title = QLabel("💬 슬랙 알림 옵션")
        slack_title.setStyleSheet("font-weight: bold; color: #FF6347;")
        layout.addWidget(slack_title)
        
        self.send_slack_checkbox = QCheckBox('슬랙 알림 발송')
        self.send_slack_checkbox.setChecked(False)  # 기본적으로 비활성화
        layout.addWidget(self.send_slack_checkbox)
        
        # 웹훅 URL 입력 (슬랙 하위 옵션)
        webhook_layout = QHBoxLayout()
        webhook_layout.addSpacing(20)  # 들여쓰기
        webhook_layout.addWidget(QLabel('Webhook URL:'))
        self.webhook_url_input = QLineEdit()
        self.webhook_url_input.setPlaceholderText("https://hooks.slack.com/services/...")
        self.webhook_url_input.setEnabled(False)  # 초기에는 비활성화
        webhook_layout.addWidget(self.webhook_url_input)
        layout.addLayout(webhook_layout)
        
        # 구분선
        separator4 = QLabel("─" * 60)
        separator4.setStyleSheet("color: gray;")
        layout.addWidget(separator4)
        
        # Jira 이슈 생성 관련 옵션
        jira_title = QLabel("🎯 Jira 이슈 생성 옵션")
        jira_title.setStyleSheet("font-weight: bold; color: #8A2BE2;")
        layout.addWidget(jira_title)
        
        self.create_jira_checkbox = QCheckBox('실패한 테스트별로 Jira 이슈 생성')
        self.create_jira_checkbox.setChecked(False)  # 기본적으로 비활성화
        layout.addWidget(self.create_jira_checkbox)
        
        # Jira 서버 URL 입력 (Jira 하위 옵션)
        jira_url_layout = QHBoxLayout()
        jira_url_layout.addSpacing(20)  # 들여쓰기
        jira_url_layout.addWidget(QLabel('Jira 서버 URL:'))
        self.jira_url_input = QLineEdit()
        self.jira_url_input.setPlaceholderText("https://company.atlassian.net")
        self.jira_url_input.setEnabled(False)  # 초기에는 비활성화
        jira_url_layout.addWidget(self.jira_url_input)
        layout.addLayout(jira_url_layout)
        
        # Jira 프로젝트 키 입력
        jira_project_layout = QHBoxLayout()
        jira_project_layout.addSpacing(20)  # 들여쓰기
        jira_project_layout.addWidget(QLabel('프로젝트 키:'))
        self.jira_project_input = QLineEdit()
        self.jira_project_input.setPlaceholderText("예: TEST, QA")
        self.jira_project_input.setEnabled(False)  # 초기에는 비활성화
        jira_project_layout.addWidget(self.jira_project_input)
        layout.addLayout(jira_project_layout)
        
        # Jira 사용자 이메일
        jira_email_layout = QHBoxLayout()
        jira_email_layout.addSpacing(20)  # 들여쓰기
        jira_email_layout.addWidget(QLabel('사용자 이메일:'))
        self.jira_email_input = QLineEdit()
        self.jira_email_input.setPlaceholderText("user@company.com")
        self.jira_email_input.setEnabled(False)  # 초기에는 비활성화
        jira_email_layout.addWidget(self.jira_email_input)
        layout.addLayout(jira_email_layout)
        
        # Jira 인증 정보
        jira_auth_layout = QHBoxLayout()
        jira_auth_layout.addSpacing(20)  # 들여쓰기
        jira_auth_layout.addWidget(QLabel('API 토큰:'))
        self.jira_token_input = QLineEdit()
        self.jira_token_input.setPlaceholderText("Jira API 토큰 또는 비밀번호")
        self.jira_token_input.setEchoMode(QLineEdit.Password)  # 비밀번호 형태로 표시
        self.jira_token_input.setEnabled(False)  # 초기에는 비활성화
        jira_auth_layout.addWidget(self.jira_token_input)
        layout.addLayout(jira_auth_layout)
        
        # 슬랙 체크박스와 웹훅 URL 입력란 연동
        def on_slack_checkbox_changed():
            self.webhook_url_input.setEnabled(self.send_slack_checkbox.isChecked())
            if not self.send_slack_checkbox.isChecked():
                self.webhook_url_input.clear()
        
        self.send_slack_checkbox.toggled.connect(on_slack_checkbox_changed)
        
        # Jira 체크박스와 관련 입력란들 연동
        def on_jira_checkbox_changed():
            enabled = self.create_jira_checkbox.isChecked()
            self.jira_url_input.setEnabled(enabled)
            self.jira_project_input.setEnabled(enabled)
            self.jira_email_input.setEnabled(enabled)
            self.jira_token_input.setEnabled(enabled)
            if not enabled:
                self.jira_url_input.clear()
                self.jira_project_input.clear()
                self.jira_email_input.clear()
                self.jira_token_input.clear()
        
        self.create_jira_checkbox.toggled.connect(on_jira_checkbox_changed)
        
        # 엑셀 체크박스와 하위 옵션들 연동
        def on_excel_checkbox_changed():
            enabled = self.export_excel_checkbox.isChecked()
            self.include_images_checkbox.setEnabled(enabled)
            self.excel_filename_input.setEnabled(enabled)
            if not enabled:
                self.include_images_checkbox.setChecked(False)
        
        self.export_excel_checkbox.toggled.connect(on_excel_checkbox_changed)
        
        widget.setLayout(layout)
        return widget
    
    def parse_params(self, params):
        # 전체 명령어 문자열 재구성
        full_command = 'exportresult ' + ' '.join(params)
        print(f"exportresult 전체 명령어: {full_command}")
        
        try:
            # 1. 토큰 분할 (큰따옴표 고려)
            def tokenize_command(command):
                tokens = []
                current_token = ""
                in_quotes = False
                
                i = 0
                while i < len(command):
                    char = command[i]
                    
                    if char == '"':
                        in_quotes = not in_quotes
                    elif char == ' ' and not in_quotes:
                        if current_token:
                            tokens.append(current_token)
                            current_token = ""
                    else:
                        current_token += char
                    i += 1
                
                if current_token:
                    tokens.append(current_token)
                
                return tokens
            
            tokens = tokenize_command(full_command)
            print(f"토큰 분할 결과: {tokens}")
            
            # 'exportresult' 명령어 제거
            if tokens and tokens[0] == 'exportresult':
                tokens = tokens[1:]
            
            # 기본값 설정
            parsed = {
                'title': '',
                'export_excel': True,
                'include_images': False,
                'export_text': True,
                'send_slack': False,
                'webhook_url': '',
                'create_jira': False,
                'jira_url': '',
                'jira_project': '',
                'jira_email': '',
                'jira_token': '',
                'excel_filename': ''  # 맨 뒤에 추가 (하위 호환성 유지)
            }
            
            # 파라미터 파싱: [title] [export_excel] [include_images] [export_text] [send_slack] [webhook_url] [create_jira] [jira_url] [jira_project] [jira_email] [jira_token] [excel_filename]
            if len(tokens) > 0:
                value = tokens[0].strip('"')  # 큰따옴표 제거
                parsed['title'] = '' if value in ["''", '""', ''] else value
            if len(tokens) > 1:
                parsed['export_excel'] = tokens[1].lower() == 'true'
            if len(tokens) > 2:
                parsed['include_images'] = tokens[2].lower() == 'true'
            if len(tokens) > 3:
                parsed['export_text'] = tokens[3].lower() == 'true'
            if len(tokens) > 4:
                parsed['send_slack'] = tokens[4].lower() == 'true'
            if len(tokens) > 5:
                value = tokens[5].strip('"')  # 큰따옴표 제거
                parsed['webhook_url'] = '' if value in ["''", '""', ''] else value
            if len(tokens) > 6:
                parsed['create_jira'] = tokens[6].lower() == 'true'
            if len(tokens) > 7:
                value = tokens[7].strip('"')  # 큰따옴표 제거
                parsed['jira_url'] = '' if value in ["''", '""', ''] else value
            if len(tokens) > 8:
                value = tokens[8].strip('"')  # 큰따옴표 제거
                parsed['jira_project'] = '' if value in ["''", '""', ''] else value
            if len(tokens) > 9:
                value = tokens[9].strip('"')  # 큰따옴표 제거
                parsed['jira_email'] = '' if value in ["''", '""', ''] else value
            if len(tokens) > 10:
                value = tokens[10].strip('"')  # 큰따옴표 제거
                parsed['jira_token'] = '' if value in ["''", '""', ''] else value
            if len(tokens) > 11:
                value = tokens[11].strip('"')  # 큰따옴표 제거
                parsed['excel_filename'] = '' if value in ["''", '""', ''] else value
            
            print(f"exportresult 파싱 성공: {parsed}")
            return parsed
            
        except Exception as e:
            print(f"exportresult 파싱 오류: {e}")
            print(f"입력 파라미터: {params}")
            # 오류 시 기본값 반환
            return {
                'title': '',
                'export_excel': True,
                'include_images': False,
                'export_text': True,
                'send_slack': False,
                'webhook_url': '',
                'create_jira': False,
                'jira_url': '',
                'jira_project': '',
                'jira_email': '',
                'jira_token': '',
                'excel_filename': ''  # 맨 뒤에 추가 (하위 호환성 유지)
            }
    
    def set_ui_values(self, params):
        if not params:
            return
        self.title_input.setText(params.get('title', ''))
        self.export_excel_checkbox.setChecked(params.get('export_excel', True))
        self.include_images_checkbox.setChecked(params.get('include_images', False))
        self.excel_filename_input.setText(params.get('excel_filename', ''))
        self.export_text_checkbox.setChecked(params.get('export_text', True))
        self.send_slack_checkbox.setChecked(params.get('send_slack', False))
        self.webhook_url_input.setText(params.get('webhook_url', ''))
        self.create_jira_checkbox.setChecked(params.get('create_jira', False))
        self.jira_url_input.setText(params.get('jira_url', ''))
        self.jira_project_input.setText(params.get('jira_project', ''))
        self.jira_email_input.setText(params.get('jira_email', ''))
        self.jira_token_input.setText(params.get('jira_token', ''))
    
    def get_command_string(self):
        title = self.title_input.text().strip()
        export_excel = self.export_excel_checkbox.isChecked()
        include_images = self.include_images_checkbox.isChecked()
        export_text = self.export_text_checkbox.isChecked()
        send_slack = self.send_slack_checkbox.isChecked()
        webhook_url = self.webhook_url_input.text().strip()
        create_jira = self.create_jira_checkbox.isChecked()
        jira_url = self.jira_url_input.text().strip()
        jira_project = self.jira_project_input.text().strip()
        jira_email = self.jira_email_input.text().strip()
        jira_token = self.jira_token_input.text().strip()
        excel_filename = self.excel_filename_input.text().strip()  # 맨 뒤로 이동
        
        # 띄어쓰기가 있는 제목은 큰따옴표로 감싸기
        if title and ' ' in title:
            title = f'"{title}"'
        elif not title:
            title = "''"
        
        # URL들도 큰따옴표 처리
        if webhook_url and (' ' in webhook_url or not webhook_url):
            webhook_url = f'"{webhook_url}"'
        elif not webhook_url:
            webhook_url = "''"
            
        if jira_url and (' ' in jira_url or not jira_url):
            jira_url = f'"{jira_url}"'
        elif not jira_url:
            jira_url = "''"
            
        if not jira_project:
            jira_project = "''"
            
        if not jira_email:
            jira_email = "''"
            
        if not jira_token:
            jira_token = "''"
        
        # 엑셀 파일명 처리 (맨 뒤)
        if excel_filename and ' ' in excel_filename:
            excel_filename = f'"{excel_filename}"'
        elif not excel_filename:
            excel_filename = "''"
            
        return f"exportresult {title} {export_excel} {include_images} {export_text} {send_slack} {webhook_url} {create_jira} {jira_url} {jira_project} {jira_email} {jira_token} {excel_filename}"
    
    def execute(self, params, window_coords=None, processor_state=None):
        print("-"*50)
        print("📋 테스트 결과 내보내기 (exportresult)")
        print("-"*50)
        
        if processor_state is None or 'test_results' not in processor_state:
            print("저장된 테스트 결과가 없습니다.")
            return
        
        test_results = processor_state['test_results']
        if not test_results:
            print("저장된 테스트 결과가 없습니다.")
            return
        
        # 윈도우 실행 정보 출력 (간소화)
        window_info = processor_state.get('window_info', {})
        executed_apps = processor_state.get('executed_apps', [])
        
        # 현재 실제 선택된 윈도우로 업데이트
        self._update_current_window_info(window_info)
        
        if window_info or executed_apps:
            print(" 실행 환경 정보:")
            print("-" * 30)
            
            # 기본 윈도우 정보
            if window_info:
                target_app = window_info.get('target_app', '알 수 없음')
                print(f"• 대상 윈도우: {target_app}")
                
                execution_file = window_info.get('execution_file')
                if execution_file:
                    print(f"• 명령어 파일: {execution_file}")
                else:
                    print("• 명령어 파일: 없음 (직접 설정)")
            
            # 대상 앱 실행 경로 (runapp으로 실행된 앱이 있을 때만)
            if executed_apps:
                for app_info in executed_apps:
                    if app_info.get('file_path'):
                        print(f"• 대상 앱 실행 경로: {app_info['file_path']}")
                        break  # 첫 번째 실행 파일만 표시
            
            print("")
        
        # 파라미터 파싱
        export_excel = params.get('export_excel', True) if params else True
        include_images = params.get('include_images', False) if params else False
        excel_filename = params.get('excel_filename', '') if params else ''
        export_text = params.get('export_text', True) if params else True
        send_slack = params.get('send_slack', False) if params else False
        webhook_url = params.get('webhook_url', '') if params else ''
        create_jira = params.get('create_jira', False) if params else False
        jira_url = params.get('jira_url', '') if params else ''
        jira_project = params.get('jira_project', '') if params else ''
        jira_email = params.get('jira_email', '') if params else ''
        jira_token = params.get('jira_token', '') if params else ''
        
        # 파일명 생성
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        title = params.get('title', '') if params else ''
        
        # 엑셀 파일명 처리
        if excel_filename:
            # 사용자가 파일명을 지정한 경우
            # .xlsx 확장자가 없으면 추가
            if not excel_filename.endswith('.xlsx'):
                excel_filename = f"{excel_filename}.xlsx"
            excel_path = os.path.join(test_results_dir, excel_filename)
            
            # 기본 파일명은 엑셀 파일명 기반으로 생성
            base_filename = excel_filename.replace('.xlsx', '')
        else:
            # 타임스탬프 기반 파일명 생성
            if title:
                # 파일명에 사용할 수 없는 문자들을 안전한 문자로 치환
                safe_title = title.replace(' ', '_').replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
                base_filename = f"{safe_title}_{timestamp}"
            else:
                base_filename = f"testresult_{timestamp}"
            
            excel_filename = f"{base_filename}.xlsx"
            excel_path = os.path.join(test_results_dir, excel_filename)
        
        text_filename = f"{base_filename}_summary.txt"
        text_path = os.path.join(test_results_dir, text_filename)
        
        # test_results_dir 존재 확인
        if not os.path.exists(test_results_dir):
            try:
                os.makedirs(test_results_dir)
                print(f"테스트 결과 디렉토리 생성: {test_results_dir}")
            except Exception as e:
                print(f"테스트 결과 디렉토리 생성 실패: {e}")
                return
        
        # 실행할 작업들 확인
        tasks = []
        if export_excel:
            tasks.append("엑셀 파일")
        if export_text:
            tasks.append("텍스트 요약")
        if send_slack and webhook_url:
            tasks.append("슬랙 알림")
        if create_jira and jira_url and jira_project and jira_email and jira_token:
            tasks.append("Jira 이슈 생성")
        
        if not tasks:
            print("❌ 실행할 작업이 선택되지 않았습니다.")
            return
        
        print(f"📋 실행할 작업: {', '.join(tasks)}")
        print("-" * 50)
        
        # 1. 엑셀 파일 생성
        excel_success = False
        if export_excel:
            # 파일 존재 여부 확인
            file_exists = os.path.exists(excel_path)
            if file_exists:
                print(f"📝 기존 엑셀 파일 발견: {excel_filename}")
                print(f"   → 기존 파일에 결과를 이어서 추가합니다.")
            
            if include_images:
                # 이미지 포함 모드
                mode_text = "이어쓰기" if file_exists else "새로 생성"
                print(f"엑셀 파일 {mode_text} 중... (이미지 포함 모드)")
                try:
                    self._create_excel_report(test_results, excel_path, processor_state, append=file_exists)
                    print(f"✓ 엑셀 파일 저장됨 (이미지 포함): {excel_path}")
                    excel_success = True
                except Exception as e:
                    print(f"❌ 이미지 포함 모드 실패: {e}")
                    # 이미지 포함 모드 실패 시 안전 모드로 fallback
                    try:
                        print("안전 모드로 fallback 시도...")
                        self._create_excel_report_safe(test_results, excel_path, processor_state, append=file_exists)
                        print(f"✓ 엑셀 파일 저장됨 (안전 모드 fallback): {excel_path}")
                        excel_success = True
                    except Exception as e2:
                        print(f"❌ 안전 모드 fallback도 실패: {e2}")
            else:
                # 안전 모드 (기본값)
                mode_text = "이어쓰기" if file_exists else "새로 생성"
                print(f"엑셀 파일 {mode_text} 중... (안전 모드 - 이미지 제외)")
                try:
                    self._create_excel_report_safe(test_results, excel_path, processor_state, append=file_exists)
                    print(f"✓ 엑셀 파일 저장됨 (안전 모드): {excel_path}")
                    excel_success = True
                except Exception as e:
                    print(f"❌ 안전 모드 엑셀 생성 실패: {e}")
        
        # 2. 텍스트 요약 파일 생성
        txt_success = False
        if export_text:
            print("텍스트 요약 파일 생성 중...")
            try:
                self._create_text_summary(test_results, text_path, processor_state, title)
                print(f"✓ 텍스트 요약 저장됨: {text_path}")
                txt_success = True
            except Exception as e:
                print(f"❌ 텍스트 요약 생성 실패: {e}")
        
        # 3. 슬랙 알림 발송
        slack_success = False
        if send_slack and webhook_url:
            print("슬랙 알림 발송 중...")
            try:
                slack_success = self._send_slack_notification(test_results, webhook_url, processor_state, base_filename, title)
                if slack_success:
                    print("✓ 슬랙 알림 발송 완료")
                else:
                    print("❌ 슬랙 알림 발송 실패")
            except Exception as e:
                print(f"❌ 슬랙 알림 발송 중 오류: {e}")
        
        # 4. Jira 이슈 생성
        jira_success = False
        if create_jira and jira_url and jira_project and jira_email and jira_token:
            failed_tests = [r for r in test_results if r['result'] == 'Fail']
            if failed_tests:
                print("Jira 이슈 생성 중...")
                try:
                    jira_success = self._create_jira_issues(failed_tests, jira_url, jira_project, jira_email, jira_token, title)
                    if jira_success:
                        print(f"✓ Jira 이슈 생성 완료 ({len(failed_tests)}개 실패 테스트)")
                    else:
                        print("❌ Jira 이슈 생성 실패")
                except Exception as e:
                    print(f"❌ Jira 이슈 생성 중 오류: {e}")
            else:
                print("실패한 테스트가 없어 Jira 이슈를 생성하지 않습니다.")
                jira_success = True  # 실패한 테스트가 없으면 성공으로 간주
        
        # 결과 요약
        print("-" * 50)
        success_count = sum([excel_success, txt_success, slack_success, jira_success])
        total_tasks = len([t for t in [export_excel, export_text, send_slack and webhook_url, create_jira and jira_url and jira_project and jira_email and jira_token] if t])
        
        if success_count == total_tasks and total_tasks > 0:
            print(f"✅ 모든 작업이 성공적으로 완료되었습니다. (총 {len(test_results)}개 테스트 결과)")
        elif success_count > 0:
            print(f"⚠️ 일부 작업만 완료됨 ({success_count}/{total_tasks})")
            print(f"   엑셀: {'✓' if excel_success else '❌'}, "
                  f"텍스트: {'✓' if txt_success else '❌'}, "
                  f"슬랙: {'✓' if slack_success else '❌'}, "
                  f"Jira: {'✓' if jira_success else '❌'}")
        else:
            print(f"❌ 모든 작업이 실패했습니다.")
        
        # 생성된 리포트 파일 경로를 processor_state에 저장 (메인 앱에서 열기 위해)
        if processor_state is not None:
            processor_state['last_report_txt_path'] = text_path if txt_success else None
            processor_state['last_report_excel_path'] = excel_path if excel_success else None
    
    def _create_excel_report(self, test_results, excel_path, processor_state=None, append=False):
        """엑셀 리포트 생성 (스크린샷 이미지 포함)
        
        Args:
            test_results: 테스트 결과 리스트
            excel_path: 엑셀 파일 경로
            processor_state: 프로세서 상태
            append: True면 기존 파일에 추가, False면 새로 생성
        """
        from openpyxl.drawing.image import Image as OpenpyxlImage
        import os
        import pyautogui
        from datetime import datetime
        
        # 이미지 삽입 여부 확인 (PIL 불필요 - openpyxl 직접 사용)
        insert_images = True
        
        # 기존 파일이 있고 append 모드면 로드, 없으면 새로 생성
        if append and os.path.exists(excel_path):
            try:
                wb = load_workbook(excel_path)
                ws = wb.active
                start_row = ws.max_row + 1  # 다음 행부터 추가
                print(f"   기존 데이터 {ws.max_row - 1}개, 새 데이터 {len(test_results)}개 추가")
            except Exception as e:
                print(f"   ⚠️ 기존 파일 로드 실패, 새로 생성합니다: {e}")
                wb = Workbook()
                ws = wb.active
                ws.title = "테스트 결과"
                start_row = 2
                
                # 헤더 설정
                headers = ['번호', '제목', '결과', '기대값', '추출값', '매칭모드', '스크린샷 경로', '스크린샷', '전체 스크린샷']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                ws.row_dimensions[1].height = 25
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "테스트 결과"
            start_row = 2
            
            # 헤더 설정 (전체 스크린샷 컬럼 추가)
            headers = ['번호', '제목', '결과', '기대값', '추출값', '매칭모드', '스크린샷 경로', '스크린샷', '전체 스크린샷']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 행 높이 설정 (모든 행을 기본 높이로 고정)
            ws.row_dimensions[1].height = 25  # 헤더 행
        
        default_row_height = 22  # 기본 셀 높이 (포인트 단위)
        
        # 데이터 입력 및 스크린샷 이미지 삽입
        for idx, result in enumerate(test_results):
            row = start_row + idx
            # 기본 데이터 입력
            ws.cell(row=row, column=1, value=row-1)  # 번호
            ws.cell(row=row, column=2, value=result['title'])
            ws.cell(row=row, column=3, value=result['result'])
            ws.cell(row=row, column=4, value=result['expected_text'])
            ws.cell(row=row, column=5, value=result['extracted_text'])
            ws.cell(row=row, column=6, value=result.get('match_mode', 'N/A'))
            ws.cell(row=row, column=7, value=result['screenshot_path'])
            
            # 행 높이를 기본값으로 고정 (이미지 크기와 무관)
            ws.row_dimensions[row].height = default_row_height
            
            # 스크린샷 이미지 삽입 (원본을 기본 셀 높이에 맞춤)
            screenshot_path = result['screenshot_path']
            
            if screenshot_path and os.path.exists(screenshot_path) and insert_images:
                try:
                    # 원본 스크린샷 파일로 바로 openpyxl Image 객체 생성
                    img = OpenpyxlImage(screenshot_path)
                    
                    # 원본 이미지의 비율 계산
                    original_width = img.width
                    original_height = img.height
                    aspect_ratio = original_width / original_height
                    
                    # 기본 셀 높이에 맞춰 이미지 크기 조정 (포인트 → 픽셀 변환: 1 포인트 ≈ 1.33 픽셀)
                    target_height = default_row_height * 1.33  # 픽셀로 변환
                    target_width = target_height * aspect_ratio
                    
                    # 이미지 크기를 기본 셀 높이에 맞춤 (원본은 유지되며, 표시만 축소)
                    img.width = int(target_width)
                    img.height = int(target_height)
                    
                    # 이미지를 스크린샷 컬럼(H열)에 배치
                    cell_ref = f'H{row}'
                    ws.add_image(img, cell_ref)
                    
                    ws.cell(row=row, column=8, value="이미지 삽입됨")
                    print(f"  ✓ 이미지 삽입 성공: {result['title']} (원본: {original_width}x{original_height}, 표시: {int(target_width)}x{int(target_height)})")
                        
                except Exception as e:
                    print(f"  ❌ 이미지 삽입 실패 ({result['title']}): {e}")
                    ws.cell(row=row, column=8, value=f"삽입 실패: {str(e)[:20]}")
                            
            elif screenshot_path and os.path.exists(screenshot_path) and not insert_images:
                ws.cell(row=row, column=8, value="이미지 제외됨")
            else:
                ws.cell(row=row, column=8, value="스크린샷 없음")
            
            # Fail인 경우 해당 앱의 전체 화면 스크린샷 추가 캡처 및 삽입
            if result['result'] == 'Fail' and insert_images:
                try:
                    import pygetwindow as gw
                    
                    # 타임스탬프 및 경로 설정
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                    full_screenshot_dir = os.path.dirname(excel_path)
                    full_screenshot_filename = f"fail_app_fullscreen_{row-1}_{timestamp}.png"
                    full_screenshot_path = os.path.join(full_screenshot_dir, full_screenshot_filename)
                    
                    # processor_state에서 현재 앱 정보 가져오기
                    app_captured = False
                    if processor_state and 'window_info' in processor_state:
                        target_app = processor_state['window_info'].get('target_app', '')
                        if target_app:
                            # 해당 앱 윈도우 찾기
                            windows = gw.getWindowsWithTitle(target_app)
                            if windows:
                                window = windows[0]
                                # 앱 윈도우 영역만 캡처
                                region = (window.left, window.top, window.width, window.height)
                                app_screenshot = pyautogui.screenshot(region=region)
                                app_screenshot.save(full_screenshot_path)
                                app_captured = True
                                print(f"  📸 Fail 항목 앱 전체 스크린샷 저장: {full_screenshot_filename} (앱: {target_app})")
                    
                    # 앱 정보가 없거나 윈도우를 찾지 못한 경우 전체 화면 캡처
                    if not app_captured:
                        full_screenshot = pyautogui.screenshot()
                        full_screenshot.save(full_screenshot_path)
                        print(f"  📸 Fail 항목 전체 화면 스크린샷 저장: {full_screenshot_filename} (앱 찾지 못함)")
                    
                    # 전체 스크린샷 이미지 삽입 (원본 해상도 유지)
                    full_img = OpenpyxlImage(full_screenshot_path)
                    
                    # 원본 이미지 크기
                    full_original_width = full_img.width
                    full_original_height = full_img.height
                    full_aspect_ratio = full_original_width / full_original_height
                    
                    # 기본 셀 높이에 맞춰 표시 크기만 조정 (원본은 유지)
                    full_target_height = default_row_height * 1.33
                    full_target_width = full_target_height * full_aspect_ratio
                    
                    # anchor를 사용하여 원본 해상도 유지
                    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
                    
                    # 이미지 크기를 기본 셀 높이에 맞춤 (원본 해상도는 유지)
                    full_img.width = int(full_target_width)
                    full_img.height = int(full_target_height)
                    
                    # 이미지를 전체 스크린샷 컬럼(I열)에 배치
                    cell_ref = f'I{row}'
                    ws.add_image(full_img, cell_ref)
                    
                    ws.cell(row=row, column=9, value="앱 전체 캡처됨" if app_captured else "전체 화면 캡처됨")
                    print(f"  ✓ 전체 스크린샷 삽입 성공: {result['title']} (원본: {full_original_width}x{full_original_height}, 표시: {int(full_target_width)}x{int(full_target_height)})")
                    
                except Exception as e:
                    print(f"  ❌ 전체 스크린샷 삽입 실패 ({result['title']}): {e}")
                    import traceback
                    print(f"     상세 오류: {traceback.format_exc()}")
                    ws.cell(row=row, column=9, value=f"캡처 실패: {str(e)[:20]}")
            else:
                # Pass이거나 이미지 삽입 안 하는 경우
                if result['result'] == 'Pass':
                    ws.cell(row=row, column=9, value="Pass (불필요)")
                else:
                    ws.cell(row=row, column=9, value="-")
        
        # 요약 시트 추가
        ws_summary = wb.create_sheet("요약")
        total_tests = len(test_results)
        passed_tests = len([r for r in test_results if r['result'] == 'Pass'])
        failed_tests = total_tests - passed_tests
        
        ws_summary.cell(row=1, column=1, value="테스트 요약")
        ws_summary.cell(row=2, column=1, value="총 테스트")
        ws_summary.cell(row=2, column=2, value=total_tests)
        ws_summary.cell(row=3, column=1, value="Pass")
        ws_summary.cell(row=3, column=2, value=passed_tests)
        ws_summary.cell(row=4, column=1, value="Fail")
        ws_summary.cell(row=4, column=2, value=failed_tests)
        ws_summary.cell(row=5, column=1, value="성공률")
        ws_summary.cell(row=5, column=2, value=f"{(passed_tests/total_tests*100):.1f}%")
        
        # 실패한 테스트 목록
        if failed_tests > 0:
            ws_summary.cell(row=7, column=1, value="실패한 테스트:")
            failed_titles = [r['title'] for r in test_results if r['result'] == 'Fail']
            for i, title in enumerate(failed_titles, 8):
                ws_summary.cell(row=i, column=1, value=f"• {title}")
        
        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            # 스크린샷 컬럼(H열, I열)은 고정 너비 적용
            if column == 'H':
                ws.column_dimensions[column].width = 40  # 스크린샷 컬럼은 넓게
                continue
            if column == 'I':
                ws.column_dimensions[column].width = 40  # 전체 스크린샷 컬럼도 넓게
                continue
                
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 파일 저장 (UTF-8 인코딩으로 처리)
        wb.save(excel_path)
    
    def _create_excel_report_safe(self, test_results, excel_path, processor_state=None, append=False):
        """엑셀 리포트 생성 (안전 모드 - 이미지 없음)
        
        Args:
            test_results: 테스트 결과 리스트
            excel_path: 엑셀 파일 경로
            processor_state: 프로세서 상태
            append: True면 기존 파일에 추가, False면 새로 생성
        """
        # 기존 파일이 있고 append 모드면 로드, 없으면 새로 생성
        if append and os.path.exists(excel_path):
            try:
                wb = load_workbook(excel_path)
                ws = wb.active
                start_row = ws.max_row + 1  # 다음 행부터 추가
                print(f"   기존 데이터 {ws.max_row - 1}개, 새 데이터 {len(test_results)}개 추가")
            except Exception as e:
                print(f"   ⚠️ 기존 파일 로드 실패, 새로 생성합니다: {e}")
                wb = Workbook()
                ws = wb.active
                ws.title = "테스트 결과"
                start_row = 2
                
                # 헤더 설정
                headers = ['번호', '제목', '결과', '기대값', '추출값', '매칭모드', '스크린샷 경로']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "테스트 결과"
            start_row = 2
            
            # 헤더 설정 (스크린샷 경로만 포함, 실제 이미지 제외)
            headers = ['번호', '제목', '결과', '기대값', '추출값', '매칭모드', '스크린샷 경로']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        
        # 데이터 입력 (이미지 삽입 없음)
        for idx, result in enumerate(test_results):
            row = start_row + idx
            ws.cell(row=row, column=1, value=row-1)  # 번호
            ws.cell(row=row, column=2, value=result['title'])
            ws.cell(row=row, column=3, value=result['result'])
            ws.cell(row=row, column=4, value=result['expected_text'])
            ws.cell(row=row, column=5, value=result['extracted_text'])
            ws.cell(row=row, column=6, value=result.get('match_mode', 'N/A'))
            ws.cell(row=row, column=7, value=result['screenshot_path'])
        
        # 요약 시트 추가
        ws_summary = wb.create_sheet("요약")
        total_tests = len(test_results)
        passed_tests = len([r for r in test_results if r['result'] == 'Pass'])
        failed_tests = total_tests - passed_tests
        
        ws_summary.cell(row=1, column=1, value="테스트 요약")
        ws_summary.cell(row=2, column=1, value="총 테스트")
        ws_summary.cell(row=2, column=2, value=total_tests)
        ws_summary.cell(row=3, column=1, value="Pass")
        ws_summary.cell(row=3, column=2, value=passed_tests)
        ws_summary.cell(row=4, column=1, value="Fail")
        ws_summary.cell(row=4, column=2, value=failed_tests)
        ws_summary.cell(row=5, column=1, value="성공률")
        ws_summary.cell(row=5, column=2, value=f"{(passed_tests/total_tests*100):.1f}%")
        
        # 실행 환경 정보 추가
        current_row = 7
        if processor_state:
            window_info = processor_state.get('window_info', {})
            executed_apps = processor_state.get('executed_apps', [])
            # 현재 실제 선택된 윈도우로 업데이트
            self._update_current_window_info(window_info)
            
            if window_info or executed_apps:
                ws_summary.cell(row=current_row, column=1, value="실행 환경 정보:")
                current_row += 1
                
                if window_info:
                    target_app = window_info.get('target_app', '알 수 없음')
                    ws_summary.cell(row=current_row, column=1, value="• 대상 윈도우:")
                    ws_summary.cell(row=current_row, column=2, value=target_app)
                    current_row += 1
                    
                    execution_file = window_info.get('execution_file')
                    if execution_file:
                        ws_summary.cell(row=current_row, column=1, value="• 명령어 파일:")
                        ws_summary.cell(row=current_row, column=2, value=execution_file)
                    else:
                        ws_summary.cell(row=current_row, column=1, value="• 명령어 파일:")
                        ws_summary.cell(row=current_row, column=2, value="없음 (직접 설정)")
                    current_row += 1
                
                if executed_apps:
                    for app_info in executed_apps:
                        if app_info.get('file_path'):
                            ws_summary.cell(row=current_row, column=1, value="• 대상 앱 실행 경로:")
                            ws_summary.cell(row=current_row, column=2, value=app_info['file_path'])
                            current_row += 1
                            break  # 첫 번째 실행 파일만 표시
                
                current_row += 1  # 빈 행 추가
        
        # 실패한 테스트 목록
        if failed_tests > 0:
            ws_summary.cell(row=current_row, column=1, value="실패한 테스트:")
            current_row += 1
            failed_titles = [r['title'] for r in test_results if r['result'] == 'Fail']
            for i, title in enumerate(failed_titles):
                ws_summary.cell(row=current_row + i, column=1, value=f"• {title}")
        
        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 파일 저장
        wb.save(excel_path)
    
    def _create_text_summary(self, test_results, text_path, processor_state=None, title=""):
        """텍스트 요약 파일 생성 (슬랙 전송용)"""
        total_tests = len(test_results)
        passed_tests = len([r for r in test_results if r['result'] == 'Pass'])
        failed_tests = total_tests - passed_tests
        
        # 테스트 제목 - title 파라미터를 우선적으로 사용
        test_title = title if title else (processor_state.get('test_session_title', '알 수 없는 테스트') if processor_state else '테스트')
        start_time = processor_state.get('test_session_start') if processor_state else None
        current_time = datetime.now()
        
        # UTF-8 인코딩으로 텍스트 파일 생성
        with open(text_path, 'w', encoding='utf-8') as f:
            f.write("="*50 + "\n")
            f.write("테스트 결과 요약\n")
            f.write("="*50 + "\n")
            f.write(f"📋 테스트 제목: {test_title}\n")
            f.write(f"📅 생성 시간: {current_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            
            if start_time:
                duration = current_time - start_time
                f.write(f"⏰ 시작 시간: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"⏱️ 소요 시간: {duration}\n")
            
            # 윈도우 실행 정보 추가 (간소화)
            window_info = processor_state.get('window_info', {}) if processor_state else {}
            executed_apps = processor_state.get('executed_apps', []) if processor_state else []
            # 현재 실제 선택된 윈도우로 업데이트
            if processor_state:
                self._update_current_window_info(window_info)
            
            if window_info or executed_apps:
                f.write("\n📱 실행 환경 정보:\n")
                
                # 기본 윈도우 정보
                if window_info:
                    target_app = window_info.get('target_app', '알 수 없음')
                    f.write(f"   • 대상 윈도우: {target_app}\n")
                    
                    execution_file = window_info.get('execution_file')
                    if execution_file:
                        f.write(f"   • 명령어 파일: {execution_file}\n")
                    else:
                        f.write(f"   • 명령어 파일: 없음 (직접 설정)\n")
                
                # 대상 앱 실행 경로 (runapp으로 실행된 앱이 있을 때만)
                if executed_apps:
                    for app_info in executed_apps:
                        if app_info.get('file_path'):
                            f.write(f"   • 대상 앱 실행 경로: {app_info['file_path']}\n")
                            break  # 첫 번째 실행 파일만 표시
            
            f.write("\n📊 테스트 결과:\n")
            f.write(f"   • 총 테스트: {total_tests}개\n")
            f.write(f"   • Pass: {passed_tests}개\n")
            f.write(f"   • Fail: {failed_tests}개\n")
            f.write(f"   • 성공률: {(passed_tests/total_tests*100):.1f}%\n")
            
            # Pass 항목
            if passed_tests > 0:
                passed_titles = [r['title'] for r in test_results if r['result'] == 'Pass']
                f.write(f"\n✅ Pass 항목:\n")
                for i, title in enumerate(passed_titles, 1):
                    f.write(f"   {i}. {title}\n")
            
            # Fail 항목
            if failed_tests > 0:
                failed_titles = [r['title'] for r in test_results if r['result'] == 'Fail']
                f.write(f"\n❌ Fail 항목:\n")
                for i, title in enumerate(failed_titles, 1):
                    f.write(f"   {i}. {title}\n")
            
            if passed_tests == 0 and failed_tests == 0:
                f.write(f"\n📋 테스트 결과가 없습니다.\n")
            
            f.write("\n" + "="*50 + "\n")
    
    def _send_slack_notification(self, test_results, webhook_url, processor_state=None, base_filename="", title=""):
        """슬랙 알림 발송 - 텍스트 파일과 동일한 형태"""
        try:
            import requests
            
            total_tests = len(test_results)
            passed_tests = len([r for r in test_results if r['result'] == 'Pass'])
            failed_tests = total_tests - passed_tests
            success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
            
            # 테스트 제목 - title 파라미터를 우선적으로 사용
            test_title = title if title else (processor_state.get('test_session_title', '테스트') if processor_state else '테스트')
            start_time = processor_state.get('test_session_start') if processor_state else None
            current_time = datetime.now()
            
            # 텍스트 파일과 동일한 형태로 메시지 구성
            message_lines = []
            message_lines.append("=" * 50)
            message_lines.append("테스트 결과 요약")
            message_lines.append("=" * 50)
            message_lines.append(f"📋 테스트 제목: {test_title}")
            message_lines.append(f"📅 생성 시간: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
            
            if start_time:
                duration = current_time - start_time
                message_lines.append(f"⏰ 시작 시간: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
                message_lines.append(f"⏱️ 소요 시간: {duration}")

            # 윈도우 실행 정보 추가 (간소화)
            window_info = processor_state.get('window_info', {}) if processor_state else {}
            executed_apps = processor_state.get('executed_apps', []) if processor_state else []
            # 현재 실제 선택된 윈도우로 업데이트
            if processor_state:
                self._update_current_window_info(window_info)

            if window_info or executed_apps:
                message_lines.append("")
                message_lines.append("📱 실행 환경 정보:")

                # 기본 윈도우 정보
                if window_info:
                    target_app = window_info.get('target_app', '알 수 없음')
                    message_lines.append(f"   • 대상 윈도우: {target_app}")

                    execution_file = window_info.get('execution_file')
                    if execution_file:
                        message_lines.append(f"   • 명령어 파일: {execution_file}")
                    else:
                        message_lines.append(f"   • 명령어 파일: 없음 (직접 설정)")

                # 대상 앱 실행 경로 (runapp으로 실행된 앱이 있을 때만)
                if executed_apps:
                    for app_info in executed_apps:
                        if app_info.get('file_path'):
                            message_lines.append(f"   • 대상 앱 실행 경로: {app_info['file_path']}")
                            break  # 첫 번째 실행 파일만 표시
            
            message_lines.append("")
            message_lines.append("📊 테스트 결과:")
            message_lines.append(f"   • 총 테스트 항목: {total_tests}개")
            message_lines.append(f"   • Pass: {passed_tests}개")
            message_lines.append(f"   • Fail: {failed_tests}개")
            message_lines.append(f"   • 성공률: {success_rate:.1f}%")
            
            # Pass 항목
            if passed_tests > 0:
                passed_titles = [r['title'] for r in test_results if r['result'] == 'Pass']
                message_lines.append("")
                message_lines.append("✅ Pass 항목:")
                for i, test_title in enumerate(passed_titles, 1):
                    message_lines.append(f"   {i}. {test_title}")
            
            # Fail 항목
            if failed_tests > 0:
                failed_titles = [r['title'] for r in test_results if r['result'] == 'Fail']
                message_lines.append("")
                message_lines.append("❌ Fail 항목:")
                for i, test_title in enumerate(failed_titles, 1):
                    message_lines.append(f"   {i}. {test_title}")
            
            if passed_tests == 0 and failed_tests == 0:
                message_lines.append("")
                message_lines.append("📋 테스트 결과가 없습니다.")
            
            message_lines.append("")
            message_lines.append("=" * 50)
            
            # 메시지를 하나의 문자열로 결합
            message_text = "\n".join(message_lines)
            
            # 슬랙 메시지 색상 결정
            color = "good" if failed_tests == 0 else ("warning" if success_rate >= 70 else "danger")
            
            # 슬랙 페이로드 구성 (간단한 텍스트 형태)
            payload = {
                "text": f"🧪 테스트 결과 알림",
                "attachments": [
                    {
                        "color": color,
                        "text": f"```{message_text}```",
                        "fallback": f"테스트 결과: {passed_tests}/{total_tests} 성공 ({success_rate:.1f}%)"
                    }
                ]
            }
            
            # 슬랙으로 전송
            response = requests.post(webhook_url, json=payload, timeout=10)
            
            if response.status_code == 200:
                return True
            else:
                print(f"슬랙 전송 실패: HTTP {response.status_code} - {response.text}")
                return False
                
        except ImportError:
            print("❌ requests 모듈이 필요합니다. 'pip install requests' 로 설치해주세요.")
            return False
        except Exception as e:
            print(f"❌ 슬랙 알림 발송 중 오류: {e}")
            return False
    
    def _create_jira_issues(self, failed_tests, jira_url, jira_project, jira_email, jira_token, test_title=""):
        """실패한 테스트별로 Jira 이슈 생성"""
        try:
            import requests
            import base64
            
            # Jira API 설정
            api_url = f"{jira_url.rstrip('/')}/rest/api/2/issue"
            
            # 기본 인증 헤더 (username:token 형태로 base64 인코딩)
            # API 토큰을 사용 (username은 사용자 이메일 주소)
            auth_string = f"{jira_email}:{jira_token}"
            auth_bytes = auth_string.encode('ascii')
            auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
            
            headers = {
                'Authorization': f'Basic {auth_b64}',
                'Content-Type': 'application/json'
            }
            
            created_issues = []
            failed_issues = []
            
            for i, test in enumerate(failed_tests, 1):
                try:
                    # 이슈 제목 생성
                    issue_summary = f"[TEST FAIL] {test['title']}"
                    if test_title:
                        issue_summary = f"[{test_title}] {test['title']} - 테스트 실패"
                    
                    # 이슈 설명 생성
                    description = f"""테스트 자동화에서 실패한 테스트입니다.

*테스트 정보:*
• 테스트 제목: {test['title']}
• 테스트 결과: {test['result']}
• 기대값: {test.get('expected_text', 'N/A')}
• 추출값: {test.get('extracted_text', 'N/A')}
• 매칭모드: {test.get('match_mode', 'N/A')}

*스크린샷:*
• 경로: {test.get('screenshot_path', 'N/A')}

*테스트 실행 정보:*
• 실행 시간: {test.get('timestamp', 'N/A')}
• 전체 테스트: {test_title}

이 이슈는 테스트 자동화 시스템에서 자동으로 생성되었습니다."""
                    
                    # Jira 이슈 페이로드
                    issue_data = {
                        "fields": {
                            "project": {
                                "key": jira_project
                            },
                            "summary": issue_summary,
                            "description": description,
                            "issuetype": {
                                "name": "Bug"  # 기본적으로 Bug 타입으로 생성
                            },
                            "priority": {
                                "name": "Medium"  # 기본 우선순위
                            }
                        }
                    }
                    
                    # Jira API 호출
                    response = requests.post(api_url, json=issue_data, headers=headers, timeout=30)
                    
                    if response.status_code == 201:
                        issue_info = response.json()
                        issue_key = issue_info.get('key', 'Unknown')
                        issue_url = f"{jira_url}/browse/{issue_key}"
                        created_issues.append({
                            'test_title': test['title'],
                            'issue_key': issue_key,
                            'issue_url': issue_url
                        })
                        print(f"  ✓ [{i}/{len(failed_tests)}] {test['title']} → {issue_key}")
                    else:
                        error_msg = f"HTTP {response.status_code}"
                        try:
                            error_detail = response.json()
                            if 'errors' in error_detail:
                                error_msg += f": {error_detail['errors']}"
                        except:
                            error_msg += f": {response.text[:100]}"
                        
                        failed_issues.append({
                            'test_title': test['title'],
                            'error': error_msg
                        })
                        print(f"  ❌ [{i}/{len(failed_tests)}] {test['title']} → {error_msg}")
                    
                except Exception as e:
                    failed_issues.append({
                        'test_title': test['title'],
                        'error': str(e)
                    })
                    print(f"  ❌ [{i}/{len(failed_tests)}] {test['title']} → 오류: {e}")
            
            # 결과 요약
            if created_issues:
                print(f"\n✅ 성공적으로 생성된 Jira 이슈 ({len(created_issues)}개):")
                for issue in created_issues:
                    print(f"   • {issue['issue_key']}: {issue['test_title']}")
                    print(f"     URL: {issue['issue_url']}")
            
            if failed_issues:
                print(f"\n❌ 생성에 실패한 Jira 이슈 ({len(failed_issues)}개):")
                for issue in failed_issues:
                    print(f"   • {issue['test_title']}: {issue['error']}")
            
            # 전체 성공 여부 반환
            return len(created_issues) > 0 and len(failed_issues) == 0
            
        except ImportError:
            print("❌ requests 모듈이 필요합니다. 'pip install requests' 로 설치해주세요.")
            return False
        except Exception as e:
            print(f"❌ Jira 이슈 생성 중 오류: {e}")
            return False


class RunAppCommand(CommandBase):
    """특정 폴더에서 최신 파일을 찾아 실행하고 윈도우 자동 설정하는 명령어"""
    
    @property
    def name(self): 
        return "RunApp"
    
    @property
    def description(self): 
        return "앱 실행 및 윈도우 자동 설정 (폴더: 최신 파일 검색, 직접: 절대 경로 지정, 윈도우: 실행 없이 윈도우만 인식)"
    
    def _get_window_titles(self):
        """현재 열려있는 윈도우의 제목들을 반환 (탐색기 제외)"""
        window_titles = []
        
        try:
            import win32gui
            
            def enum_windows_proc(hwnd, param):
                if win32gui.IsWindowVisible(hwnd) and win32gui.GetWindowText(hwnd):
                    title = win32gui.GetWindowText(hwnd)
                    class_name = win32gui.GetClassName(hwnd)
                    
                    # 윈도우 탐색기와 기타 시스템 윈도우 제외
                    excluded_classes = ['ExploreWClass', 'CabinetWClass', 'Shell_TrayWnd', 'DV2ControlHost']
                    excluded_titles = ['바탕 화면', 'Desktop', '작업 표시줄', 'Taskbar', 'Program Manager']
                    
                    if (class_name not in excluded_classes and 
                        title not in excluded_titles and 
                        not title.startswith('Windows ') and
                        len(title.strip()) > 0):
                        window_titles.append(title)
                
                return True
            
            win32gui.EnumWindows(enum_windows_proc, None)
            
        except ImportError:
            # win32gui가 없는 경우 pygetwindow로 fallback
            print("⚠️ win32gui 모듈이 없습니다. pygetwindow로 대체 사용합니다.")
            try:
                import pygetwindow as gw
                all_windows = gw.getAllWindows()
                for window in all_windows:
                    try:
                        title = window.title
                        if title and len(title.strip()) > 0:
                            # 시스템 윈도우 제외
                            excluded_titles = ['바탕 화면', 'Desktop', '작업 표시줄', 'Taskbar', 'Program Manager']
                            if (title not in excluded_titles and 
                                not title.startswith('Windows ') and
                                not title.startswith('Microsoft Text Input Application')):
                                window_titles.append(title)
                    except Exception:
                        continue
            except Exception as e:
                print(f"pygetwindow로 윈도우 목록 가져오기 실패: {e}")
                # 완전 실패 시 기본 윈도우 목록 제공
                window_titles = ["메모장", "Chrome", "Firefox", "Edge", "Visual Studio Code"]
                
        except Exception as e:
            print(f"윈도우 목록 가져오기 실패: {e}")
            # 오류 시 기본 윈도우 목록 제공
            window_titles = ["메모장", "Chrome", "Firefox", "Edge", "Visual Studio Code"]

        # 중복 제거 및 정렬
        return sorted(list(set(window_titles)))
    
    def _check_existing_window(self, window_pattern):
        """이미 열려있는 윈도우를 즉시 확인"""
        if not window_pattern:
            return None
            
        import pygetwindow as gw
        
        try:
            all_windows = gw.getAllTitles()
            
            candidates = []  # 후보 윈도우들
            
            for window_title in all_windows:
                if not window_title or window_title.strip() == "":
                    continue
                
                # 패턴 매칭 시도
                match_found = False
                
                # 1. 정확한 매칭
                if window_title.lower() == window_pattern.lower():
                    candidates.append((window_title, 100))  # 우선순위 100
                    match_found = True
                
                # 2. 부분 매칭 (패턴이 윈도우 제목에 포함)
                elif window_pattern.lower() in window_title.lower():
                    candidates.append((window_title, 80))  # 우선순위 80
                    match_found = True
                
                # 3. 윈도우 제목이 패턴에 포함 (패턴이 더 긴 경우)
                elif window_title.lower() in window_pattern.lower():
                    candidates.append((window_title, 60))  # 우선순위 60
                    match_found = True
                
                # 4. 패턴의 일부 단어들이 포함되어 있는지
                elif len(window_pattern) > 3:  # 패턴이 너무 짧으면 스킵
                    pattern_words = window_pattern.lower().split()
                    title_lower = window_title.lower()
                    
                    matched_words = sum(1 for word in pattern_words if word in title_lower)
                    if matched_words > 0:
                        score = (matched_words / len(pattern_words)) * 40  # 최대 40점
                        if score >= 20:  # 최소 50% 일치
                            candidates.append((window_title, score))
                            match_found = True
            
            # 후보가 있으면 가장 높은 점수의 윈도우 반환
            if candidates:
                candidates.sort(key=lambda x: x[1], reverse=True)  # 점수 순으로 정렬
                best_window = candidates[0][0]
                print(f"윈도우 매칭 결과: '{best_window}' (점수: {candidates[0][1]})")
                return best_window
            
            return None
        
        except Exception as e:
            print(f"윈도우 확인 중 오류 발생: {e}")
            return None
    
    def _refresh_window_list(self):
        """윈도우 목록을 새로고침하여 콤보박스에 추가"""
        if not hasattr(self, 'window_pattern_input'):
            return
            
        current_text = self.window_pattern_input.currentText()
        self.window_pattern_input.clear()
        
        # 빈 항목 추가 (직접 입력용)
        self.window_pattern_input.addItem("")
        
        # 현재 열려있는 윈도우 목록 가져와서 추가
        window_titles = self._get_window_titles()
        for title in window_titles:
            self.window_pattern_input.addItem(title)
        
        # 이전에 입력했던 텍스트 복원
        if current_text:
            index = self.window_pattern_input.findText(current_text)
            if index >= 0:
                self.window_pattern_input.setCurrentIndex(index)
            else:
                self.window_pattern_input.setEditText(current_text)
    
    def create_ui(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # 실행 모드 선택
        from PyQt5.QtWidgets import QRadioButton, QButtonGroup
        mode_layout = QVBoxLayout()
        mode_layout.addWidget(QLabel('실행 모드:'))
        
        mode_buttons_layout = QHBoxLayout()
        self.folder_mode_radio = QRadioButton('폴더에서 최신 파일 찾기')
        self.direct_mode_radio = QRadioButton('직접 파일 경로 지정')
        self.window_mode_radio = QRadioButton('윈도우 인식만 (실행 없음)')
        self.folder_mode_radio.setChecked(True)  # 기본값
        
        # 버튼 그룹으로 묶어서 하나만 선택되도록
        self.mode_group = QButtonGroup()
        self.mode_group.addButton(self.folder_mode_radio, 0)
        self.mode_group.addButton(self.direct_mode_radio, 1)
        self.mode_group.addButton(self.window_mode_radio, 2)
        
        # 모드 변경 시 UI 업데이트
        self.folder_mode_radio.toggled.connect(self._update_ui_mode)
        self.direct_mode_radio.toggled.connect(self._update_ui_mode)
        self.window_mode_radio.toggled.connect(self._update_ui_mode)
        
        mode_buttons_layout.addWidget(self.folder_mode_radio)
        mode_buttons_layout.addWidget(self.direct_mode_radio)
        mode_buttons_layout.addWidget(self.window_mode_radio)
        mode_buttons_layout.addStretch()
        mode_layout.addLayout(mode_buttons_layout)
        layout.addLayout(mode_layout)
        
        # === 폴더 모드 UI ===
        self.folder_mode_widget = QWidget()
        folder_mode_layout = QVBoxLayout()
        
        # 폴더 경로 입력
        folder_layout = QHBoxLayout()
        folder_layout.addWidget(QLabel('폴더 경로:'))
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText("예: C:\\Games\\MyGame")
        self.folder_browse_btn = QPushButton('Browse')
        self.folder_browse_btn.clicked.connect(self.browse_folder)
        folder_layout.addWidget(self.folder_input)
        folder_layout.addWidget(self.folder_browse_btn)
        folder_mode_layout.addLayout(folder_layout)
        
        # 파일명 패턴 입력
        pattern_layout = QHBoxLayout()
        pattern_layout.addWidget(QLabel('파일명 패턴:'))
        self.pattern_input = QLineEdit()
        self.pattern_input.setPlaceholderText("예: *.exe, MyGame*.exe, *launcher*")
        pattern_layout.addWidget(self.pattern_input)
        folder_mode_layout.addLayout(pattern_layout)
        
        self.folder_mode_widget.setLayout(folder_mode_layout)
        layout.addWidget(self.folder_mode_widget)
        
        # === 직접 실행 모드 UI ===
        self.direct_mode_widget = QWidget()
        direct_mode_layout = QVBoxLayout()
        
        # 실행 파일 경로 입력
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel('실행 파일 경로:'))
        self.file_input = QLineEdit()
        self.file_input.setPlaceholderText("예: C:\\Games\\MyGame\\game.exe")
        self.file_browse_btn = QPushButton('Browse')
        self.file_browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(self.file_input)
        file_layout.addWidget(self.file_browse_btn)
        direct_mode_layout.addLayout(file_layout)
        
        self.direct_mode_widget.setLayout(direct_mode_layout)
        self.direct_mode_widget.setVisible(False)  # 기본적으로 숨김
        layout.addWidget(self.direct_mode_widget)
        
        # === 윈도우 인식 모드 UI ===
        self.window_mode_widget = QWidget()
        window_mode_layout = QVBoxLayout()
        
        # 설명 라벨
        info_label = QLabel('이미 실행 중인 앱의 윈도우를 찾아서 좌표만 갱신합니다.')
        info_label.setStyleSheet("color: #666; font-style: italic;")
        window_mode_layout.addWidget(info_label)
        
        self.window_mode_widget.setLayout(window_mode_layout)
        self.window_mode_widget.setVisible(False)  # 기본적으로 숨김
        layout.addWidget(self.window_mode_widget)
        
        # 윈도우 제목 패턴 (자동 인식용)
        window_layout = QHBoxLayout()
        window_layout.addWidget(QLabel('윈도우 제목 패턴:'))
        self.window_pattern_input = QComboBox()
        self.window_pattern_input.setEditable(True)  # 사용자가 직접 입력도 가능하도록
        self.window_pattern_input.setPlaceholderText("예: MyGame, Launcher (윈도우 자동 선택용)")
        window_layout.addWidget(self.window_pattern_input)
        
        # 윈도우 목록 새로고침 버튼
        self.refresh_windows_btn = QPushButton('🔄')
        self.refresh_windows_btn.setToolTip('현재 열려있는 윈도우 목록 새로고침')
        self.refresh_windows_btn.setMaximumWidth(30)
        self.refresh_windows_btn.clicked.connect(self._refresh_window_list)
        window_layout.addWidget(self.refresh_windows_btn)
        
        layout.addLayout(window_layout)
        
        # 초기 윈도우 목록 로드
        self._refresh_window_list()
        
        # 옵션들
        options_layout = QVBoxLayout()
        self.wait_checkbox = QCheckBox('앱이 완전히 로드될 때까지 대기')
        self.wait_checkbox.setChecked(True)
        self.auto_window_checkbox = QCheckBox('윈도우 자동 선택 및 새로고침')
        self.auto_window_checkbox.setChecked(True)
        options_layout.addWidget(self.wait_checkbox)
        options_layout.addWidget(self.auto_window_checkbox)
        layout.addLayout(options_layout)
        
        # 대기 시간 설정
        timeout_layout = QHBoxLayout()
        timeout_layout.addWidget(QLabel('최대 대기 시간:'))
        self.timeout_input = QSpinBox()
        self.timeout_input.setRange(5, 120)
        self.timeout_input.setValue(30)
        self.timeout_input.setSuffix('초')
        timeout_layout.addWidget(self.timeout_input)
        timeout_layout.addStretch()
        layout.addLayout(timeout_layout)
        
        widget.setLayout(layout)
        return widget
    
    def _update_ui_mode(self):
        """모드 변경에 따른 UI 업데이트"""
        is_folder_mode = self.folder_mode_radio.isChecked()
        is_direct_mode = self.direct_mode_radio.isChecked()
        is_window_mode = self.window_mode_radio.isChecked()
        
        self.folder_mode_widget.setVisible(is_folder_mode)
        self.direct_mode_widget.setVisible(is_direct_mode)
        self.window_mode_widget.setVisible(is_window_mode)
    
    def browse_folder(self):
        """폴더 선택 다이얼로그"""
        from PyQt5.QtWidgets import QFileDialog
        folder = QFileDialog.getExistingDirectory(None, "폴더 선택")
        if folder:
            self.folder_input.setText(folder)
    
    def browse_file(self):
        """실행 파일 선택 다이얼로그"""
        from PyQt5.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getOpenFileName(
            None, 
            "실행 파일 선택", 
            "", 
            "실행 파일 (*.exe);;모든 파일 (*.*)"
        )
        if file_path:
            self.file_input.setText(file_path)
    
    def parse_params(self, params):
        # 전체 명령어 문자열 재구성
        full_command = 'runapp ' + ' '.join(params)
        print(f"runapp 전체 명령어: {full_command}")
        
        try:
            import re
            
            # 1. 토큰 분할 (따옴표 고려)
            tokens = []
            current_token = ""
            in_quotes = False
            
            i = 0
            while i < len(full_command):
                char = full_command[i]
                
                if char == '"':
                    in_quotes = not in_quotes
                elif char == ' ' and not in_quotes:
                    if current_token:
                        tokens.append(current_token)
                        current_token = ""
                else:
                    current_token += char
                i += 1
            
            if current_token:
                tokens.append(current_token)
            
            print(f"토큰 분할 결과: {tokens}")
            
            # 'runapp' 제거
            if tokens and tokens[0] == 'runapp':
                tokens = tokens[1:]
            
            if len(tokens) < 1:
                print(f"runapp 파라미터 부족: 모드 지정 필요")
                return {}
            
            # 모드 확인 (첫 번째 파라미터)
            mode = tokens[0].lower()
            
            if mode == 'folder':
                # 폴더 모드: runapp folder "folder_path" "file_pattern" "window_pattern" wait auto_window timeout
                if len(tokens) < 3:
                    print(f"runapp folder 모드 파라미터 부족: folder_path, file_pattern 필요")
                    return {}
                
                parsed = {
                    'mode': 'folder',
                    'folder_path': tokens[1],
                    'file_pattern': tokens[2],
                    'window_pattern': tokens[3] if len(tokens) > 3 else '',
                    'wait_for_load': True,  # 기본값
                    'auto_window': True,    # 기본값
                    'timeout': 30           # 기본값
                }
                
                # 불린/숫자 파라미터 처리
                if len(tokens) > 4:
                    parsed['wait_for_load'] = tokens[4].lower() in ['true', '1', 'yes']
                
                if len(tokens) > 5:
                    parsed['auto_window'] = tokens[5].lower() in ['true', '1', 'yes']
                
                if len(tokens) > 6:
                    parsed['timeout'] = int(tokens[6])
                    
            elif mode == 'direct':
                # 직접 모드: runapp direct "file_path" "window_pattern" wait auto_window timeout
                if len(tokens) < 2:
                    print(f"runapp direct 모드 파라미터 부족: file_path 필요")
                    return {}
                
                parsed = {
                    'mode': 'direct',
                    'file_path': tokens[1],
                    'window_pattern': tokens[2] if len(tokens) > 2 else '',
                    'wait_for_load': True,  # 기본값
                    'auto_window': True,    # 기본값
                    'timeout': 30           # 기본값
                }
                
                # 불린/숫자 파라미터 처리
                if len(tokens) > 3:
                    parsed['wait_for_load'] = tokens[3].lower() in ['true', '1', 'yes']
                
                if len(tokens) > 4:
                    parsed['auto_window'] = tokens[4].lower() in ['true', '1', 'yes']
                
                if len(tokens) > 5:
                    parsed['timeout'] = int(tokens[5])
                    
            elif mode == 'window':
                # 윈도우 인식 모드: runapp window "window_pattern" timeout
                if len(tokens) < 2:
                    print(f"runapp window 모드 파라미터 부족: window_pattern 필요")
                    return {}
                
                parsed = {
                    'mode': 'window',
                    'window_pattern': tokens[1],
                    'timeout': 30           # 기본값
                }
                
                # 숫자 파라미터 처리
                if len(tokens) > 2:
                    parsed['timeout'] = int(tokens[2])
            
            else:
                # 구 버전 호환성을 위해 기본적으로 폴더 모드로 처리
                print(f"구 버전 호환성: 폴더 모드로 처리")
                if len(tokens) < 2:
                    print(f"runapp 파라미터 부족: 최소 2개 필요 (folder_path, file_pattern)")
                    return {}
                
                parsed = {
                    'mode': 'folder',
                    'folder_path': tokens[0],
                    'file_pattern': tokens[1],
                    'window_pattern': tokens[2] if len(tokens) > 2 else '',
                    'wait_for_load': True,  # 기본값
                    'auto_window': True,    # 기본값
                    'timeout': 30           # 기본값
                }
                
                # 불린/숫자 파라미터 처리
                if len(tokens) > 3:
                    parsed['wait_for_load'] = tokens[3].lower() in ['true', '1', 'yes']
                
                if len(tokens) > 4:
                    parsed['auto_window'] = tokens[4].lower() in ['true', '1', 'yes']
                
                if len(tokens) > 5:
                    parsed['timeout'] = int(tokens[5])
            
            print(f"runapp 파싱 성공: {parsed}")
            return parsed
            
        except (ValueError, IndexError) as e:
            print(f"runapp 파싱 오류: {e}")
            print(f"입력 파라미터: {params}")
            return {}
    
    def set_ui_values(self, params):
        if not params:
            return
        
        # 모드에 따라 UI 설정
        mode = params.get('mode', 'folder')
        if mode == 'direct':
            self.direct_mode_radio.setChecked(True)
            self.file_input.setText(params.get('file_path', ''))
        elif mode == 'window':
            self.window_mode_radio.setChecked(True)
        else:
            self.folder_mode_radio.setChecked(True)
            self.folder_input.setText(params.get('folder_path', ''))
            self.pattern_input.setText(params.get('file_pattern', ''))
        
        # 공통 설정
        self.window_pattern_input.setEditText(params.get('window_pattern', ''))
        
        # 윈도우 모드가 아닐 때만 설정 (윈도우 모드는 이 옵션들이 없음)
        if mode != 'window':
            self.wait_checkbox.setChecked(params.get('wait_for_load', True))
            self.auto_window_checkbox.setChecked(params.get('auto_window', True))
            
        self.timeout_input.setValue(params.get('timeout', 30))
        
        # UI 모드 업데이트
        self._update_ui_mode()
    
    def get_command_string(self):
        window_pattern = f'"{self.window_pattern_input.currentText()}"' if self.window_pattern_input.currentText() else '""'
        timeout = str(self.timeout_input.value())
        
        if self.window_mode_radio.isChecked():
            # 윈도우 모드: runapp window "window_pattern" timeout
            return f"runapp window {window_pattern} {timeout}"
        elif self.direct_mode_radio.isChecked():
            # 직접 모드: runapp direct "file_path" "window_pattern" wait auto_window timeout
            wait = 'true' if self.wait_checkbox.isChecked() else 'false'
            auto_window = 'true' if self.auto_window_checkbox.isChecked() else 'false'
            file_path = self.file_input.text()
            file_path_quoted = f'"{file_path}"' if ' ' in file_path else file_path
            return f"runapp direct {file_path_quoted} {window_pattern} {wait} {auto_window} {timeout}"
        else:
            # 폴더 모드: runapp folder "folder_path" "file_pattern" "window_pattern" wait auto_window timeout
            wait = 'true' if self.wait_checkbox.isChecked() else 'false'
            auto_window = 'true' if self.auto_window_checkbox.isChecked() else 'false'
            folder = self.folder_input.text()
            folder_quoted = f'"{folder}"' if ' ' in folder else folder
            pattern = f'"{self.pattern_input.text()}"'
            return f"runapp folder {folder_quoted} {pattern} {window_pattern} {wait} {auto_window} {timeout}"
    
    def execute(self, params, window_coords=None, processor_state=None):
        if not params:
            print("오류: runapp 명령어에 필요한 파라미터가 없습니다.")
            return
        
        # CommandProcessor 참조 저장 (중지 신호 체크용)
        self.processor = params.get('processor')
        
        # 모드 확인
        mode = params.get('mode', 'folder')
        window_pattern = params.get('window_pattern', '')
        wait_for_load = params.get('wait_for_load', True)
        auto_window = params.get('auto_window', True)
        timeout = params.get('timeout', 30)
        
        print(f"앱 실행 모드: {mode}")
        
        # 윈도우 모드: 실행 없이 윈도우만 찾기
        if mode == 'window':
            if not window_pattern:
                print("오류: 윈도우 모드에는 window_pattern이 필요합니다.")
                return
                
            print(f"윈도우 인식 모드: '{window_pattern}' 패턴으로 윈도우 검색 중...")
            
            # 기존 윈도우 즉시 확인
            existing_window = self._check_existing_window(window_pattern)
            if existing_window:
                print(f"✓ 윈도우를 발견했습니다: {existing_window}")
                self._auto_select_window(existing_window, processor_state)
                # 윈도우만 인식한 경우에도 정보 저장
                if processor_state is not None:
                    if 'executed_apps' not in processor_state:
                        processor_state['executed_apps'] = []
                    
                    app_info = {
                        'execution_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'file_path': None,  # 윈도우 인식만이므로 파일 없음
                        'file_name': None,
                        'mode': 'window',
                        'window_pattern': window_pattern,
                        'detected_window': existing_window
                    }
                    processor_state['executed_apps'].append(app_info)
                    print(f"윈도우 인식 정보 저장됨: {existing_window}")
                return
            
            # 윈도우를 찾지 못했을 때 대기 시간동안 재시도
            print(f"윈도우를 찾지 못함. {timeout}초 동안 재시도...")
            detected_window = self._wait_for_window(window_pattern, timeout, False)
            if detected_window:
                print(f"✓ 윈도우 감지됨: {detected_window}")
                self._auto_select_window(detected_window, processor_state)
            else:
                print(f"❌ '{window_pattern}' 패턴의 윈도우를 찾을 수 없습니다. (타임아웃: {timeout}초)")
            return
        
        # 1. 윈도우가 이미 열려있는지 확인 (폴더/직접 모드)
        if window_pattern and auto_window:
            existing_window = self._check_existing_window(window_pattern)
            if existing_window:
                print(f"✓ 이미 열려있는 윈도우를 발견했습니다: {existing_window}")
                print("🔄 발견된 윈도우를 자동 선택하고 활성화합니다...")
                self._auto_select_window(existing_window, processor_state)
                return  # 이미 열려있으니까 실행 종료
        
        # 2. 모드별 파일 경로 결정
        if mode == 'direct':
            # 직접 모드: 절대 경로로 파일 지정
            if 'file_path' not in params:
                print("오류: direct 모드에는 file_path 파라미터가 필요합니다.")
                return
            
            file_to_run = params['file_path']
            print(f"직접 실행 파일: {file_to_run}")
            
            # 파일 존재 여부 확인
            if not os.path.exists(file_to_run):
                print(f"❌ 파일이 존재하지 않습니다: {file_to_run}")
                return
                
        else:
            # 폴더 모드: 최신 파일 찾기
            if 'folder_path' not in params or 'file_pattern' not in params:
                print("오류: folder 모드에는 folder_path, file_pattern 파라미터가 필요합니다.")
                return
            
            folder_path = params['folder_path']
            file_pattern = params['file_pattern']
            print(f"폴더 모드 실행: 폴더='{folder_path}', 패턴='{file_pattern}'")
            
            file_to_run = self._find_latest_file(folder_path, file_pattern)
            if not file_to_run:
                print(f"❌ 패턴 '{file_pattern}'에 맞는 파일을 찾을 수 없습니다.")
                return
            
            print(f"✓ 발견된 최신 파일: {file_to_run}")
        
        # 3. 앱 실행 (공통 로직)
        self._execute_file(file_to_run)
        
        # 실행된 앱 정보 저장 (processor_state에 저장)
        if processor_state is not None:
            if 'executed_apps' not in processor_state:
                processor_state['executed_apps'] = []
            
            app_info = {
                'execution_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'file_path': file_to_run,
                'file_name': os.path.basename(file_to_run),
                'mode': mode,
                'window_pattern': window_pattern
            }
            processor_state['executed_apps'].append(app_info)
            print(f"앱 실행 정보 저장됨: {os.path.basename(file_to_run)}")
        
        # 4. 윈도우 대기 및 자동 선택
        if auto_window:
            detected_window = self._wait_for_window(window_pattern, timeout, wait_for_load)
            if detected_window:
                print(f"✓ 윈도우 감지됨: {detected_window}")
                self._auto_select_window(detected_window, processor_state)
            else:
                print(f"⚠️ 윈도우를 감지하지 못했습니다 (타임아웃: {timeout}초)")
    
    def _execute_file(self, file_path):
        """파일 실행 공통 로직"""
        try:
            print(f"앱 실행 중: {file_path}")
            
            # 파일이 있는 디렉토리를 작업 디렉토리로 설정
            file_dir = os.path.dirname(file_path)
            print(f"작업 디렉토리를 {file_dir}로 설정")
            
            # 바로가기 파일(.lnk) 처리
            if file_path.lower().endswith('.lnk'):
                print(f"바로가기 파일 감지: {file_path}")
                # Windows에서 바로가기 파일은 os.startfile로 실행
                os.startfile(file_path)
                print(f"✓ 바로가기 파일이 실행되었습니다")
            else:
                # 일반 실행 파일 - 해당 파일의 디렉토리에서 실행
                process = subprocess.Popen([file_path], cwd=file_dir)
                print(f"✓ 앱이 실행되었습니다 (PID: {process.pid}, 작업디렉토리: {file_dir})")
        except Exception as e:
            print(f"❌ 앱 실행 실패: {e}")
            # 대안으로 os.startfile 시도
            try:
                print(f"대안 실행 시도 중...")
                os.startfile(file_path)
                print(f"✓ 대안 방법으로 파일이 실행되었습니다")
            except Exception as e2:
                print(f"❌ 대안 실행도 실패: {e2}")
                return
    
    def _find_latest_file(self, folder_path, pattern):
        """최적화된 최신 파일 검색 알고리즘"""
        if not os.path.exists(folder_path):
            print(f"❌ 폴더가 존재하지 않습니다: {folder_path}")
            return None
        
        print(f"파일 검색 중... (패턴: {pattern})")
        print(f"검색 폴더: {folder_path}")
        
        latest_file = None
        latest_time = 0
        found_files = []  # 디버깅용
        
        try:
            # os.walk를 사용한 재귀적 검색 (최적화)
            for root, dirs, files in os.walk(folder_path):
                # .git, node_modules 등 숨겨진 폴더만 스킵 (OP.GG 같은 폴더는 유지)
                dirs[:] = [d for d in dirs if not (d.startswith('.') and len(d) > 1) and d not in ['node_modules', '__pycache__']]
                
                #print(f"검색 중인 폴더: {root}")
                #print(f"발견된 파일들: {files}")
                
                # 패턴 매칭
                for file in files:
                    #print(f"파일 체크: {file} vs 패턴: {pattern}")
                    
                    # 다양한 방식으로 매칭 시도
                    match_found = False
                    
                    # 1. 정확한 매칭 (확장자 무시)
                    file_base = os.path.splitext(file)[0]  # 확장자 제거
                    if file_base.lower() == pattern.lower():
                        match_found = True
                        print(f"  ✓ 정확한 매칭 (확장자 무시): {file}")
                    
                    # 2. fnmatch 패턴 매칭
                    elif self._match_pattern(file, pattern):
                        match_found = True
                        print(f"  ✓ fnmatch 매칭: {file}")
                    
                    # 3. 포함 매칭 (pattern이 파일명에 포함)
                    elif pattern.lower() in file.lower():
                        match_found = True
                        print(f"  ✓ 포함 매칭: {file}")
                    
                    if match_found:
                        file_path = os.path.join(root, file)
                        found_files.append(file_path)
                        try:
                            mtime = os.path.getmtime(file_path)
                            if mtime > latest_time:
                                latest_time = mtime
                                latest_file = file_path
                                print(f"  ✓ 새로운 최신 파일: {file_path}")
                        except OSError:
                            print(f"  ❌ 파일 액세스 오류: {file_path}")
                            continue
            
            print(f"검색 완료. 발견된 파일 총 {len(found_files)}개:")
            for f in found_files:
                print(f"  - {f}")
            
            if latest_file:
                print(f"선택된 최신 파일: {latest_file}")
            
            return latest_file
            
        except Exception as e:
            print(f"❌ 파일 검색 중 오류: {e}")
            return None
    
    def _match_pattern(self, filename, pattern):
        """파일명 패턴 매칭 (glob 스타일)"""
        import fnmatch
        return fnmatch.fnmatch(filename.lower(), pattern.lower())
    
    def _wait_for_window(self, window_pattern, timeout, wait_for_load):
        """윈도우가 나타날 때까지 대기"""
        import pygetwindow as gw
        
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            # stop_flag 체크 (중지 버튼 대응)
            if hasattr(self, 'processor') and self.processor:
                if (hasattr(self.processor, 'stop_flag') and self.processor.stop_flag) or \
                   (hasattr(self.processor, 'main_app') and self.processor.main_app and 
                    hasattr(self.processor.main_app, 'stop_flag') and self.processor.main_app.stop_flag):
                    print("⚠️ 윈도우 대기 중 중지 신호 감지됨")
                    return None
            try:
                all_windows = gw.getAllTitles()
                
                candidates = []  # 후보 윈도우들
                
                for window_title in all_windows:
                    # 시스템 윈도우나 일반적인 윈도우 필터링
                    skip_keywords = ['desktop', 'taskbar', 'system', '파일 탐색기', 'file explorer', 
                                   'windows explorer', 'explorer', '제어판', 'control panel',
                                   'settings', '설정', 'task manager', '작업 관리자']
                    
                    if any(skip in window_title.lower() for skip in skip_keywords):
                        # print(f"시스템 윈도우 스킵: {window_title}")  # 로그 스팸 방지를 위해 주석 처리
                        continue
                    
                    # 윈도우 패턴이 있으면 매칭 검사
                    if window_pattern:
                        if window_pattern.lower() in window_title.lower():
                            try:
                                window = gw.getWindowsWithTitle(window_title)[0]
                                
                                # 윈도우 유효성 검사
                                if window.width > 200 and window.height > 150:
                                    # 점수 기반으로 우선순위 결정
                                    score = 0
                                    
                                    # 정확한 매칭일수록 높은 점수
                                    if window_title.lower().startswith(window_pattern.lower()):
                                        score += 100  # 시작 매칭
                                    elif window_pattern.lower() == window_title.lower():
                                        score += 150  # 완전 매칭
                                    else:
                                        score += 50   # 부분 매칭
                                    
                                    # 크기가 클수록 높은 점수 (실제 앱일 가능성)
                                    score += min(window.width * window.height / 10000, 50)
                                    
                                    # 특정 키워드가 있으면 추가 점수
                                    app_keywords = ['game', 'app', 'application', '.exe']
                                    if any(keyword in window_title.lower() for keyword in app_keywords):
                                        score += 25
                                    
                                    candidates.append((window_title, score, window))
                                    # print(f"후보 윈도우: {window_title} (점수: {score})")  # 로그 스팸 방지를 위해 주석 처리
                            except:
                                continue
                    else:
                        # 패턴이 없으면 새로 생성된 윈도우 중 가장 큰 것 선택
                        try:
                            window = gw.getWindowsWithTitle(window_title)[0]
                            if (window.width > 300 and window.height > 200):
                                return window_title
                        except:
                            continue
                
                # 가장 점수가 높은 후보 선택
                if candidates:
                    candidates.sort(key=lambda x: x[1], reverse=True)  # 점수 순 정렬
                    best_candidate = candidates[0]
                    print(f"최적 윈도우 선택: {best_candidate[0]} (점수: {best_candidate[1]})")
                    
                    if wait_for_load:
                        # 윈도우가 완전히 로드될 때까지 추가 대기 (중지 신호 체크 포함)
                        print("윈도우 로딩 완료 대기 중...")
                        for _ in range(20):  # 0.1초씩 20번 = 2초
                            if hasattr(self, 'processor') and self.processor:
                                if (hasattr(self.processor, 'stop_flag') and self.processor.stop_flag) or \
                                   (hasattr(self.processor, 'main_app') and self.processor.main_app and 
                                    hasattr(self.processor.main_app, 'stop_flag') and self.processor.main_app.stop_flag):
                                    print("⚠️ 윈도우 로딩 대기 중 중지 신호 감지됨")
                                    return None
                            time.sleep(0.1)
                    
                    return best_candidate[0]
                
                # 0.5초 대기 중에도 중지 신호 체크
                for _ in range(5):  # 0.1초씩 5번 = 0.5초
                    if hasattr(self, 'processor') and self.processor:
                        if (hasattr(self.processor, 'stop_flag') and self.processor.stop_flag) or \
                           (hasattr(self.processor, 'main_app') and self.processor.main_app and 
                            hasattr(self.processor.main_app, 'stop_flag') and self.processor.main_app.stop_flag):
                            print("⚠️ 윈도우 대기 중 중지 신호 감지됨 (sleep 중)")
                            return None
                    time.sleep(0.1)
                
            except Exception as e:
                print(f"윈도우 검색 중 오류: {e}")
                time.sleep(1)
        
        return None
    
    def _auto_select_window(self, window_title, processor_state=None):
        """윈도우 자동 선택 및 메인 앱 새로고침"""
        try:
            import pygetwindow as gw
            from PyQt5.QtWidgets import QApplication
            import time
            
            # 메인 애플리케이션 인스턴스 찾기
            app = QApplication.instance()
            main_widget = None
            
            for widget in app.topLevelWidgets():
                if hasattr(widget, 'window_dropdown') and hasattr(widget, 'refresh_window_list'):
                    main_widget = widget
                    break
            
            if not main_widget:
                print("⚠️ 메인 애플리케이션을 찾을 수 없습니다.")
                return False
            
            print(f"메인 애플리케이션 찾음: {main_widget.windowTitle()}")
            
            # 원래 prefix 백업
            original_prefix = main_widget.prefix_input.text()
            print(f"기존 prefix: '{original_prefix}'")
            
            # 여러 번 시도하여 윈도우 목록 갱신 및 선택
            max_attempts = 5
            for attempt in range(max_attempts):
                print(f"🔍 드롭다운에서 '{window_title}' 선택 시도... ({attempt + 1}/{max_attempts})")
                
                # prefix를 임시로 비워서 모든 윈도우 표시
                main_widget.prefix_input.setText("")
                
                # 윈도우 목록 새로고침
                main_widget.refresh_window_list()
                
                # 잠깐 대기 (UI 업데이트 시간)
                time.sleep(0.5)
                
                # 현재 윈도우 목록 출력 (디버깅)
                print(f"현재 윈도우 목록 ({main_widget.window_dropdown.count()}개):")
                for i in range(main_widget.window_dropdown.count()):
                    item_text = main_widget.window_dropdown.itemText(i)
                    #print(f"  {i}: {item_text}")
                
                # 감지된 윈도우를 자동 선택 (정확한 매칭)
                found_index = -1
                for i in range(main_widget.window_dropdown.count()):
                    item_text = main_widget.window_dropdown.itemText(i)
                    if window_title == item_text:
                        found_index = i
                        break
                
                # 정확한 매칭이 없으면 부분 매칭 시도
                if found_index == -1:
                    print(f"정확한 매칭 실패, 부분 매칭 시도...")
                    for i in range(main_widget.window_dropdown.count()):
                        item_text = main_widget.window_dropdown.itemText(i)
                        if window_title.lower() in item_text.lower():
                            found_index = i
                            print(f"부분 매칭 발견: '{item_text}'")
                            break
                
                if found_index != -1:
                    # 윈도우 선택
                    main_widget.window_dropdown.setCurrentIndex(found_index)
                    selected_window = main_widget.window_dropdown.currentText()
                    print(f"✅ 드롭다운에서 윈도우 선택 완료: {selected_window}")
                    
                    # processor_state의 target_app도 업데이트
                    if processor_state and 'window_info' in processor_state:
                        processor_state['window_info']['target_app'] = selected_window
                        print(f"📱 대상 윈도우 정보 업데이트: {selected_window}")
                    
                    # 실제 윈도우 활성화 (맨 앞으로 가져오기)
                    try:
                        print(f"윈도우 활성화 시도: {selected_window}")
                        windows = gw.getWindowsWithTitle(selected_window)
                        if windows:
                            target_window = windows[0]
                            print(f"윈도우 객체 찾음: {target_window}")
                            
                            # 여러 방법으로 윈도우 활성화 시도
                            success = False
                            
                            # 방법 1: pygetwindow activate()
                            try:
                                target_window.activate()
                                print("✓ activate() 성공")
                                success = True
                            except Exception as e1:
                                print(f"activate() 실패: {e1}")
                            
                            # 방법 2: restore + maximize
                            if not success:
                                try:
                                    target_window.restore()
                                    time.sleep(0.1)
                                    target_window.maximize()
                                    print("✓ restore() + maximize() 성공")
                                    success = True
                                except Exception as e2:
                                    print(f"restore()/maximize() 실패: {e2}")
                            
                            # 방법 3: Windows API 사용
                            if not success:
                                try:
                                    import ctypes
                                    from ctypes import wintypes
                                    hwnd = target_window._hWnd
                                    
                                    # ShowWindow를 사용해 윈도우 표시
                                    user32 = ctypes.windll.user32
                                    user32.ShowWindow(hwnd, 9)  # SW_RESTORE
                                    time.sleep(0.1)
                                    user32.SetForegroundWindow(hwnd)
                                    print("✓ Windows API로 윈도우 활성화 성공")
                                    success = True
                                except Exception as e3:
                                    print(f"Windows API 활성화 실패: {e3}")
                            
                            # 방법 4: 클릭으로 활성화
                            if not success:
                                try:
                                    center_x = target_window.left + target_window.width // 2
                                    center_y = target_window.top + target_window.height // 2
                                    import pyautogui as pag
                                    pag.click(center_x, center_y)
                                    print("✓ 클릭으로 윈도우 활성화 성공")
                                    success = True
                                except Exception as e4:
                                    print(f"클릭 활성화도 실패: {e4}")
                            
                            if not success:
                                print("⚠️ 모든 윈도우 활성화 방법이 실패했습니다.")
                        else:
                            print(f"⚠️ 윈도우 객체를 찾을 수 없음: {selected_window}")
                    except Exception as e:
                        print(f"윈도우 활성화 실패: {e}")
                    
                    # 윈도우 활성화 완료 대기
                    print("윈도우 활성화 완료 대기 중...")
                    time.sleep(1.0)  # 윈도우가 완전히 활성화될 시간 제공
                    
                    # 좌표 업데이트 강제 실행
                    print("좌표 정보를 업데이트합니다...")
                    if hasattr(main_widget, 'update_coordinates'):
                        coords = main_widget.update_coordinates()
                        print(f"업데이트된 좌표: {coords}")
                    
                    # prefix를 선택된 윈도우에 맞게 업데이트
                    if hasattr(main_widget, 'prefix_input'):
                        # 윈도우 제목에서 의미있는 부분 추출하여 prefix 설정
                        new_prefix = self._extract_prefix_from_window(selected_window)
                        main_widget.prefix_input.setText(new_prefix)
                        print(f"프리픽스 업데이트: '{new_prefix}'")
                        
                        # 새 prefix로 윈도우 목록 다시 새로고침
                        main_widget.refresh_window_list()
                        
                        # 다시 같은 윈도우 선택
                        for j in range(main_widget.window_dropdown.count()):
                            if main_widget.window_dropdown.itemText(j) == selected_window:
                                main_widget.window_dropdown.setCurrentIndex(j)
                                break
                    
                    return True
                
                print(f"윈도우 '{window_title}' 찾기 실패, {1}초 후 재시도...")
                time.sleep(1)
            
            print(f"❌ {max_attempts}번 시도 후에도 윈도우 '{window_title}'를 찾지 못했습니다.")
            print("사용 가능한 윈도우 목록:")
            for i in range(main_widget.window_dropdown.count()):
                print(f"  - {main_widget.window_dropdown.itemText(i)}")
            
            # 실패 시 원래 prefix 복원
            print(f"원래 prefix로 복원: '{original_prefix}'")
            main_widget.prefix_input.setText(original_prefix)
            main_widget.refresh_window_list()
                
        except Exception as e:
            print(f"❌ 윈도우 자동 선택 실패: {e}")
            # 예외 발생 시에도 원래 prefix 복원
            try:
                if 'original_prefix' in locals():
                    main_widget.prefix_input.setText(original_prefix)
                    main_widget.refresh_window_list()
            except:
                pass
        
        return False
    
    def _extract_prefix_from_window(self, window_title):
        """윈도우 제목에서 적절한 prefix 추출"""
        try:
            # 일반적인 구분자들로 분할
            separators = [' - ', ' | ', ' : ', '  ']
            
            # 첫 번째 의미있는 부분 추출
            title = window_title
            for separator in separators:
                if separator in title:
                    title = title.split(separator)[0]
                    break
            
            # 공백으로 나눈 첫 번째 단어 사용 (최대 15자)
            words = title.split()
            if words:
                prefix = words[0][:15]
                return prefix
            
            # 그래도 없으면 전체 제목의 첫 15자
            return window_title[:15]
            
        except Exception as e:
            print(f"prefix 추출 실패: {e}")
            return window_title[:10]  # 기본값
    

class KeepAliveCommand(CommandBase):
    """PC 자동 잠금 방지 제어 명령어"""
    
    @property
    def name(self):
        return "KeepAlive"
    
    @property 
    def description(self):
        return "PC 자동 잠금 방지 기능 제어 (start/stop/status)"
    
    def create_ui_elements(self, form_layout):
        """UI 요소 생성"""
        from PyQt5.QtWidgets import QComboBox, QLabel, QSpinBox
        
        # Action 선택
        self.action_combo = QComboBox()
        self.action_combo.addItems(['start', 'stop', 'status'])
        form_layout.addRow(QLabel("동작:"), self.action_combo)
        
        # Interval (start 시에만 사용)
        self.interval_input = QSpinBox()
        self.interval_input.setRange(1, 60)
        self.interval_input.setValue(12)
        self.interval_input.setSuffix(" 분")
        form_layout.addRow(QLabel("간격 (start 시):"), self.interval_input)
    
    def get_command_string(self):
        """명령어 문자열 생성"""
        action = self.action_combo.currentText()
        if action == 'start':
            interval = self.interval_input.value()
            return f"keepalive {action} {interval}"
        else:
            return f"keepalive {action}"
    
    def parse_params(self, command_str):
        """파라미터 파싱"""
        tokens = command_str.strip().split()
        if len(tokens) < 2:
            return None
        
        params = {'action': tokens[1]}
        
        # start 명령어의 경우 간격 파라미터
        if params['action'] == 'start' and len(tokens) >= 3:
            try:
                params['interval'] = int(tokens[2])
            except ValueError:
                params['interval'] = 12  # 기본값
        
        return params
    
    def set_ui_values(self, params):
        """UI에 파라미터 값 설정"""
        if not params:
            return
        
        action = params.get('action', 'start')
        
        # Action 콤보박스 설정
        for i in range(self.action_combo.count()):
            if self.action_combo.itemText(i) == action:
                self.action_combo.setCurrentIndex(i)
                break
        
        # Interval 설정
        interval = params.get('interval', 12)
        self.interval_input.setValue(interval)
    
    def execute(self, params, window_coords=None, processor_state=None):
        """Keep-alive 명령어 실행"""
        if not params:
            print("오류: keepalive 명령어에 필요한 파라미터가 없습니다.")
            return
        
        action = params.get('action', 'status')
        
        try:
            from utils import start_keep_alive, stop_keep_alive, is_keep_alive_running
            
            if action == 'start':
                interval = params.get('interval', 12)
                if is_keep_alive_running():
                    print("⚠️ Keep-alive가 이미 실행 중입니다.")
                else:
                    start_keep_alive(interval_minutes=interval)
                    print(f"✅ Keep-alive 시작됨 (간격: {interval}분)")
                    
            elif action == 'stop':
                if is_keep_alive_running():
                    stop_keep_alive()
                    print("🛑 Keep-alive 중지됨")
                else:
                    print("⚠️ Keep-alive가 실행되고 있지 않습니다.")
                    
            elif action == 'status':
                if is_keep_alive_running():
                    print("✅ Keep-alive 상태: 실행 중")
                else:
                    print("🛑 Keep-alive 상태: 중지됨")
            else:
                print(f"❌ 알 수 없는 Keep-alive 동작: {action}")
                
        except Exception as e:
            print(f"❌ Keep-alive 명령어 실행 실패: {e}")


# 명령어 레지스트리 - 새 명령어는 여기만 추가하면 됩니다! 🎉
# 명령어들 모두 대문자 추가
COMMAND_REGISTRY = {
    'Press': PressCommand(),
    'Write': WriteCommand(),
    'Wait': WaitCommand(),
    'Screenshot': ScreenshotCommand(),
    'Click': ClickCommand(),
    'Drag': DragCommand(),  # ← 새 명령어 추가! 이것만 하면 끝!
    'MouseWheel': MouseWheelCommand(),  # ← 마우스 휠 조작 명령어
    'I2S': I2sCommand(),
    'I2SKR': I2skrCommand(),
    'OCR': OCRCommand(),  # ← 개선된 OCR (자동 언어 감지, 다중 줄 지원)
    'WaitUntil': WaitUntilCommand(),
    'TestText': TestTextCommand(),  # ← 텍스트 추출 기반 Pass/Fail 판별 명령어
    'ShowResults': ShowTestResultsCommand(),  # ← 테스트 결과 표시 명령어
    'ExportResult': ExportResultCommand(),  # ← 테스트 결과 다양한 형태로 내보내기 명령어 (엑셀, 텍스트, 슬랙)
    'RunApp': RunAppCommand(),  # ← 앱 실행 및 윈도우 자동 설정 명령�
   # 'keepalive': KeepAliveCommand(),  # ← PC 자동 잠금 방지 제어 명령어
}


def get_all_commands():
    """모든 등록된 명령어 반환"""
    return COMMAND_REGISTRY


def get_command(name: str) -> CommandBase:
    """특정 명령어 반환 (대소문자 구분 없음)"""
    # 대소문자 구분 없이 검색
    name_lower = name.lower()
    
    # 정확한 매칭 시도
    if name_lower in COMMAND_REGISTRY:
        return COMMAND_REGISTRY[name_lower]
    
    # 대소문자 구분 없는 매칭 시도
    for key, value in COMMAND_REGISTRY.items():
        if key.lower() == name_lower:
            return value
    
    return None


def set_main_app_for_all_commands(main_app):
    """모든 명령어 객체에 메인 앱 참조 설정"""
    for command in COMMAND_REGISTRY.values():
        if hasattr(command, 'set_main_app'):
            command.set_main_app(main_app)


def get_command_names():
    """모든 명령어 이름 리스트 반환"""
    return list(COMMAND_REGISTRY.keys())