import sys
import time
import pygetwindow as gw
import pyautogui as pag
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton, QLineEdit, QTextEdit, QScrollArea, QShortcut, QFileDialog, QMessageBox, QLineEdit, QPushButton, QCheckBox, QFrame, QDialog
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QKeySequence
from PyQt5.QtWidgets import QMenuBar, QAction
import os
from datetime import datetime
import pydirectinput as pyd
import pyperclip
import subprocess
import pytesseract
import threading
import json
#from PIL import Image
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from tes import image_to_text_with_fallback
#from PIL import Image as PILImage

# We'll set tesseract path at runtime via UI or auto-detection.
def set_pytesseract_cmd(path):
    """Set pytesseract.tesseract_cmd if the path looks valid."""
    if path and os.path.exists(path) and path.lower().endswith('tesseract.exe'):
        pytesseract.pytesseract.tesseract_cmd = path
        return True
    return False

current_dir = os.path.dirname(os.path.abspath(__file__))

screenshot_dir = os.path.join(current_dir, 'screenshot')
if not os.path.exists(screenshot_dir):
    os.makedirs(screenshot_dir)

cl_dir = os.path.join(current_dir, 'checklist')
if not os.path.exists(cl_dir):
    os.makedirs(cl_dir)
    
dir_preset = 'preset'
if not os.path.exists(dir_preset):
    os.makedirs(dir_preset)

recent_txt = "0"


class CommandPopup(QDialog):
    def __init__(self, commands, parent=None):
        super().__init__(parent)
        self.setWindowTitle("명령어 실행")
        self.resize(400, 300)
        self.layout = QVBoxLayout()
        self.labels = []
        for cmd in commands:
            hbox = QHBoxLayout()
            label = QLabel(cmd)
            status = QLabel("")
            hbox.addWidget(status)
            hbox.addWidget(label)
            self.layout.addLayout(hbox)
            self.labels.append((status, label))
        self.stop_btn = QPushButton("중지")
        self.stop_btn.clicked.connect(self.stop_execution)
        self.layout.addWidget(self.stop_btn)
        self.setLayout(self.layout)
        self.stopped = False

    def mark_executed(self, idx):
        self.labels[idx][0].setText("✅")

    def stop_execution(self):
        self.stopped = True
        self.close()

class PbbAutoApp(QWidget):
    def __init__(self):
        super().__init__()
        self.stop_flag = False
        self.initUI()

        self.prefix_input.setText('SM5')
        self.refresh_window_list()
    
    def initUI(self):
        # Layouts
        main_layout = QVBoxLayout()
        
        toggle_layout= QHBoxLayout()

        self.toggle_coord_button = QPushButton('Toggle Coordinates', self)
        self.toggle_coord_button.clicked.connect(lambda: self.toggle_layouts(self.coord_container))
        toggle_layout.addWidget(self.toggle_coord_button)

        self.toggle_preset_button = QPushButton('Toggle Preset', self)
        self.toggle_preset_button.clicked.connect(lambda: self.toggle_layouts(self.preset_container))
        toggle_layout.addWidget(self.toggle_preset_button)

        main_layout.addLayout(toggle_layout)

        self.x = 0
        self.y = 0
        self.width = 0
        self.height = 0
        self.screenshot_path = ""
        
        # Input boxes for coordinates
        self.x_input0 = QLineEdit(self)
        self.y_input0 = QLineEdit(self)
        self.apply_input0_pushButton = QPushButton('⬇️', self)
        self.apply_input0_pushButton.setShortcut(QKeySequence("Shift+F2"))
        self.x_input1 = QLineEdit(self)
        self.y_input1 = QLineEdit(self)
        self.apply_input1_pushButton = QPushButton('⬇️', self)
        self.apply_input1_pushButton.setShortcut(QKeySequence("Shift+F3"))

        self.x_pos = QLineEdit(self)
        self.y_pos = QLineEdit(self)
        self.w_pos = QLineEdit(self)
        self.h_pos = QLineEdit(self)
        self.apply_pos_pushButton = QPushButton('⬇️', self)
        self.apply_pos_pushButton.setShortcut(QKeySequence("Shift+F4"))

        self.apply_input0_pushButton.clicked.connect(self.insert_input0_text)
        self.apply_input1_pushButton.clicked.connect(self.insert_input1_text)
        self.apply_pos_pushButton.clicked.connect(self.insert_pos_text)

        # Setting placeholders for x and y
        self.x_input0.setPlaceholderText("x")
        self.y_input0.setPlaceholderText("y")
        self.x_input1.setPlaceholderText("x")
        self.y_input1.setPlaceholderText("y")

        self.x_pos.setPlaceholderText("x")
        self.y_pos.setPlaceholderText("y")
        self.w_pos.setPlaceholderText("w")
        self.h_pos.setPlaceholderText("h")

        # Adding coordinate inputs to the layout
        coord_layout0 = QHBoxLayout()
        coord_layout0.addWidget(self.x_input0)
        coord_layout0.addWidget(self.y_input0)
        coord_layout0.addWidget(self.apply_input0_pushButton)

        # Adding widgets to main layout
        #main_layout.addLayout(coord_layout0)

        # Adding coordinate inputs to the layout
        coord_layout1 = QHBoxLayout()
        coord_layout1.addWidget(self.x_input1)
        coord_layout1.addWidget(self.y_input1)
        coord_layout1.addWidget(self.apply_input1_pushButton)
        
        # Adding widgets to main layout
        #main_layout.addLayout(coord_layout1)

        # Adding coordinate inputs to the layout
        coord_layout2 = QHBoxLayout()
        coord_layout2.addWidget(self.x_pos)
        coord_layout2.addWidget(self.y_pos)
        coord_layout2.addWidget(self.w_pos)
        coord_layout2.addWidget(self.h_pos)
        coord_layout2.addWidget(self.apply_pos_pushButton)

        # Adding widgets to main layout
        #main_layout.addLayout(coord_layout2)


        self.coord_container = QWidget()
        self.coord_container.setStyleSheet("border: 0.5px solid #1866E1;")
        coord_container_layout = QVBoxLayout()
        coord_container_layout.addLayout(coord_layout0)
        coord_container_layout.addLayout(coord_layout1)
        coord_container_layout.addLayout(coord_layout2)
        self.coord_container.setLayout(coord_container_layout)
        main_layout.addWidget(self.coord_container)
        

        # Setting F3 shortcut for capturing mouse coordinates
        shortcut0 = QShortcut(QKeySequence("F2"), self)
        shortcut0.activated.connect(lambda: self.capture_mouse_position(0))
        # Setting F3 shortcut for capturing mouse coordinates
        shortcut1 = QShortcut(QKeySequence("F3"), self)
        shortcut1.activated.connect(lambda: self.capture_mouse_position(1))

        # 실행파일 경로 인풋박스 및 버튼
        self.file_path_input = QLineEdit(self)
        self.file_path_input.setPlaceholderText("Select .bat file...")
        
        self.file_select_button = QPushButton("Browse", self)
        self.file_select_button.clicked.connect(self.select_bat_file)

        self.file_run_button = QPushButton("Run File", self)
        self.file_run_button.clicked.connect(self.run_bat_file)

        # Adding to layout
        file_layout = QHBoxLayout()
        file_layout.addWidget(self.file_path_input)
        file_layout.addWidget(self.file_select_button)
        file_layout.addWidget(self.file_run_button)
        
        main_layout.addLayout(file_layout)

        # Tesseract executable path input and button
        tesseract_layout = QHBoxLayout()
        self.tesseract_input = QLineEdit(self)
        self.tesseract_input.setPlaceholderText('Path to tesseract.exe (optional)')
        self.tesseract_browse_btn = QPushButton('Browse Tesseract', self)
        self.tesseract_browse_btn.clicked.connect(self.select_tesseract_file)
        tesseract_layout.addWidget(self.tesseract_input)
        tesseract_layout.addWidget(self.tesseract_browse_btn)
        main_layout.addLayout(tesseract_layout)

        '''
        region
        '''
        #region 프리셋
        preset_layout = QHBoxLayout()
        self.preset_prefix = QComboBox()
        self.preset_prefix.setFixedWidth(105)
        preset_layout.addWidget(self.preset_prefix)
        self.preset = QComboBox()
        #layout.addWidget(self.preset)
        preset_layout.addWidget(self.preset)

        # Refresh and Apply Preset Buttons
        self.delete_preset_btn = QPushButton('❌')
        self.delete_preset_btn.setFixedWidth(25)
        self.delete_preset_btn.clicked.connect(self.deletePreset)
        preset_layout.addWidget(self.delete_preset_btn)

        self.refresh_preset_btn = QPushButton('🔄')
        self.refresh_preset_btn.setFixedWidth(25)
        self.refresh_preset_btn.clicked.connect(self.refreshPresets)
        preset_layout.addWidget(self.refresh_preset_btn)

        self.apply_preset_btn = QPushButton('✅')
        self.apply_preset_btn.setToolTip('프리셋 적용')
        self.apply_preset_btn.setFixedWidth(25)
        self.apply_preset_btn.clicked.connect(self.applyPreset)
        preset_layout.addWidget(self.apply_preset_btn)

        #main_layout.addLayout(preset_layout)

        add_preset_layout = QHBoxLayout()

        self.add_preset_line = QLineEdit()
        add_preset_layout.addWidget(self.add_preset_line)
        
        self.save_preset_btn = QPushButton('💾')
        self.save_preset_btn.setFixedWidth(25)
        self.save_preset_btn.clicked.connect(self.savePreset)
        add_preset_layout.addWidget(self.save_preset_btn)

        #main_layout.addLayout(add_preset_layout)


        self.preset_container = QWidget()
        self.preset_container.setStyleSheet("border: 0.5px solid #1866E1;")
        preset_container_layout = QVBoxLayout()
        preset_container_layout.addLayout(preset_layout)
        preset_container_layout.addLayout(add_preset_layout)
        self.preset_container.setLayout(preset_container_layout)
        main_layout.addWidget(self.preset_container)
        #endregion


        # Dropdown, prefix input, refresh button, and coordinate output (x, y, w, h)
        self.prefix_input = QLineEdit(self)
        self.prefix_input.setPlaceholderText("Enter window title prefix...")
        self.refresh_button = QPushButton('Refresh', self)
        self.refresh_button.clicked.connect(self.refresh_window_list)
        self.window_dropdown = QComboBox(self)
        self.multi_checkbox = QCheckBox('Multi',self)
        self.multi_align_button = QPushButton('Align', self)
        self.multi_align_button.clicked.connect(self.align_windows)
        self.coord_label = QLabel('Coordinates: (x, y, w, h)', self)

        # Refresh layout
        refresh_layout = QHBoxLayout()
        refresh_layout.addWidget(self.prefix_input)
        refresh_layout.addWidget(self.refresh_button)
        refresh_layout.addWidget(self.window_dropdown)
        refresh_layout.addWidget(self.multi_checkbox)
        refresh_layout.addWidget(self.multi_align_button)



        # Textarea for commands
        self.textarea = QTextEdit(self)
        self.textarea.setPlaceholderText("Enter commands here...")

        # Execute button
        execute_layout = QHBoxLayout()
        self.execute_count_label = QLabel('실행횟수', self)
        self.execute_count_lineEdit = QLineEdit(self)
        self.execute_count_lineEdit.setFixedWidth(50)
        self.open_report_checkbox = QCheckBox('Open Report', self)
        self.open_report_checkbox.setChecked(True)
        self.open_screenshot_image = QCheckBox('Open Screenshot Image', self)
        self.open_screenshot_image.setChecked(True)
        self.execute_button = QPushButton('Execute (F5)', self)
        self.execute_button.setShortcut('F5')
        self.execute_button.clicked.connect(self.execute_commands)
        execute_layout.addWidget(self.execute_count_label)
        execute_layout.addWidget(self.execute_count_lineEdit)
        execute_layout.addWidget(self.open_report_checkbox)
        execute_layout.addWidget(self.open_screenshot_image)
        execute_layout.addStretch()
        execute_layout.addWidget(self.execute_button)

        # Stop button
        self.stop_button = QPushButton('Stop', self)
        self.stop_button.clicked.connect(self.stop_execution)
        execute_layout.addWidget(self.stop_button)

        # Add widgets to main layout
        main_layout.addLayout(refresh_layout)
        main_layout.addWidget(self.coord_label)
        main_layout.addWidget(self.textarea)
        main_layout.addLayout(execute_layout)

        # Set main layout
        self.setLayout(main_layout)

        # Window properties
        self.setWindowTitle('PbbAuto - Test Automation')
        self.setGeometry(300, 300, 500, 400)

        # Load config (tesseract path) if present; otherwise try to auto-detect
        if not self.load_config():
            self.auto_detect_tesseract()

        # 메뉴바 추가: 메뉴 -> Test OCR and Set Tesseract Path
        menubar = QMenuBar(self)
        menu = menubar.addMenu('메뉴')
        set_tess_action = QAction('Set Tesseract Path', self)
        set_tess_action.triggered.connect(self.select_tesseract_file)
        menu.addAction(set_tess_action)

        test_ocr_action = QAction('Test OCR', self)
        test_ocr_action.triggered.connect(self.test_ocr)
        menu.addAction(test_ocr_action)
        # Place the menubar at the top of the widget using a layout trick
        # Create a small layout to contain the menubar above the main layout
        outer_layout = QVBoxLayout()
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.addWidget(menubar)
        outer_layout.addLayout(main_layout)
        self.setLayout(outer_layout)

    # ----- Config persistence -----
    def load_config(self):
        """Load config.json and set tesseract path. Returns True if loaded and set."""
        config_path = os.path.join(current_dir, 'config.json')
        if not os.path.exists(config_path):
            return False
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            tess_path = cfg.get('tesseract_path')
            if tess_path and set_pytesseract_cmd(tess_path):
                try:
                    self.tesseract_input.setText(tess_path)
                except Exception:
                    pass
                print(f"Loaded tesseract path from config: {tess_path}")
                return True
        except Exception as e:
            print(f"Error loading config.json: {e}")
        return False

    def save_config(self):
        """Save current tesseract path to config.json."""
        config_path = os.path.join(current_dir, 'config.json')
        try:
            cfg = {}
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    try:
                        cfg = json.load(f)
                    except Exception:
                        cfg = {}
            tess_path = self.tesseract_input.text().strip()
            cfg['tesseract_path'] = tess_path
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
            print(f"Saved tesseract path to config: {tess_path}")
            return True
        except Exception as e:
            print(f"Error saving config.json: {e}")
            return False

    # Refresh function to update window list
    def refresh_window_list(self):
        prefix = self.prefix_input.text()
        all_windows = gw.getAllTitles()
        filtered_windows = [w for w in all_windows if prefix in w]
        self.window_dropdown.clear()
        self.window_dropdown.addItems(filtered_windows)
        if filtered_windows:
            self.update_coordinates()

    # Function to get the coordinates of the selected window
    def update_coordinates(self):
        selected_window = self.window_dropdown.currentText()
        if selected_window:
            window_obj = gw.getWindowsWithTitle(selected_window)[0]
            x, y, width, height = window_obj.left, window_obj.top, window_obj.width, window_obj.height
            self.coord_label.setText(f"Coordinates: ({x}, {y}, {width}, {height})")
            return x,y,width, height
        

    # Function to execute commands
    # def execute_commands(self):
    #     selected_window = self.window_dropdown.currentText()
    #     if selected_window:
    #         window_obj = gw.getWindowsWithTitle(selected_window)[0]
    #         window_obj.activate()  # Bring the window to the front

    #         commands = self.textarea.toPlainText().strip().split('\n')
    #         for command in commands:
    #             command = command.split('#')[0].strip()
    #             if command:  # Process only non-empty commands
    #                 self.process_command(command)
    #         try: 
    #             os.startfile(self.cl_path)
    #         except:
    #             print('리포트 파일 열기 오류')

    def execute_commands(self):
        self.stop_flag = False
        commands = [c.split('#')[0].strip() for c in self.textarea.toPlainText().strip().split('\n') if c.strip()]
        self.popup = CommandPopup(commands, self)
        self.popup.show()
        print("명령어 실행 시작")
        #thread = threading.Thread(target=self._execute_commands_worker_with_popup, args=(commands,))
        thread = threading.Thread(target=self._execute_commands_worker)
        thread.start()

    def stop_execution(self):
        self.stop_flag = True
        if hasattr(self, 'popup'):
            self.popup.stopped = True
            self.popup.close()
        print("Execution stopped by user.")

    def _execute_commands_worker(self):
        all_windows = gw.getAllWindows()
        selected_windows = []
        execute_count = self.execute_count_lineEdit.text()
        if execute_count == "" or int(execute_count) == 0:
            execute_count = 1
        else:
            execute_count = int(execute_count)

        for i in range(execute_count):
            if self.stop_flag:
                print("Stopped before window selection.")
                return

            if self.multi_checkbox.isChecked():
                for window in all_windows:
                    if window.title in [self.window_dropdown.itemText(i) for i in range(self.window_dropdown.count())]:
                        selected_windows.append(window)
            else:
                selected_window = self.window_dropdown.currentText()
                selected_windows = []
                for window in all_windows:
                    if window.title == selected_window:
                        selected_windows.append(window)
                        break

            for window in selected_windows:
                if self.stop_flag:
                    print("Stopped before window activation.")
                    return
                try:
                    window.activate()
                except Exception as e:
                    print(f"윈도우 '{window.title}' 활성화 중 오류 발생: {e}")

                commands = self.textarea.toPlainText().strip().split('\n')
                time.sleep(0.2)
                for idx, command in enumerate(commands):
                    if self.stop_flag:
                        print("Stopped during command execution.")
                        return
                    command = command.split('#')[0].strip()
                    if command:
                        self.process_command(command)
                        self.popup.mark_executed(idx)
                        

        try:
            if self.open_report_checkbox.isChecked():
                os.startfile(self.cl_path)
        except Exception as e:
            print('리포트 파일 열기 오류 :', e)

    # def _execute_commands_worker_with_popup(self, commands):
    #     for idx, command in enumerate(commands):
    #         if self.stop_flag or getattr(self.popup, 'stopped', False):
    #             print("Execution stopped by user.")
    #             break
    #         self.process_command(command)
    #         # 명령어 오른쪽에 실행 표시
    #         self.popup.mark_executed(idx)
    #         time.sleep(0.2)
    #     self.popup.close()

    def take_screenshot(self):
        """Take a full-screen screenshot."""


        timestamp = datetime.now().strftime('%y%m%d_%H%M%S')
        screenshot_path = os.path.join(screenshot_dir, f"{timestamp}.jpg")

        screenshot = pag.screenshot()
        screenshot.save(screenshot_path)
        self.screenshot_path = screenshot_path
        return screenshot_path

    def take_screenshot_with_coords(self, x, y, w, h):
        """Take a screenshot at specified coordinates."""

        timestamp = datetime.now().strftime('%y%m%d_%H%M%S')
        screenshot_path = os.path.join(screenshot_dir, f"{timestamp}.jpg")

        screenshot = pag.screenshot(region=(x, y, w, h))
        screenshot.save(screenshot_path)
        self.screenshot_path = screenshot_path
        return screenshot_path

    def image_to_text(self, img_path = "", lang='eng'):
        """Convert the most recent screenshot to text."""
        
        if img_path == "":
        
            screenshot_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'screenshot')
            if not os.path.exists(screenshot_dir):
                print("Screenshot directory does not exist.")
                return None

            screenshots = sorted(
                [f for f in os.listdir(screenshot_dir) if f.endswith('.jpg')],
                key=lambda f: os.path.getmtime(os.path.join(screenshot_dir, f)),
                reverse=True
            )

            if not screenshots:
                print("No screenshots found.")
                return None

            img_path = os.path.join(screenshot_dir, screenshots[0])

        return image_to_text_with_fallback(img_path=img_path, lang=lang, preview=False)
        #return pytesseract.image_to_string(Image.open(img_path), lang=lang)

    # Command execution logic
    def process_command(self, command):
        parts = command.split()
        action = parts[0].strip()

        
        try:
            print(f'{action} : {key}')
        except:
            print(f'{action}')

        if action == "press":
            keys = [part.strip() for part in parts[1:]]
            if len(keys) == 1:
                pyd.press(keys[0])
            elif len(keys) >= 2:
                pyd.keyDown(keys[0])
                pyd.press(keys[1])
                pyd.keyUp(keys[0])
        elif action == "write" or action == "typewrite":
            key = parts[1].strip()
            #pyd.typewrite(key)
            pyperclip.copy(key)                 # Copy the key to clipboard
            pyd.keyDown('ctrl')
            pyd.press('v')                      # Paste from clipboard
            pyd.keyUp('ctrl')
            
        elif action == "wait":
            duration = float(parts[1].strip())
            time.sleep(duration)
            
        elif action == "screenshot":
            if len(parts) == 5:  # If coordinates are provided
                try:
                    x = int(parts[1])
                    y = int(parts[2])
                    w = int(parts[3])
                    h = int(parts[4])
                    screenshot_path = self.take_screenshot_with_coords(x, y, w, h)
                    print(f"Screenshot taken and saved at {screenshot_path}")
                except (IndexError, ValueError):
                    print("Invalid coordinates for screenshot. Use: screenshot x y w h")
            else:  # No coordinates, use default take_screenshot()
                screenshot_path = self.take_screenshot()
                print(f"Screenshot taken and saved at {screenshot_path}")

        elif action == "cheat":
            key = " ".join(parts[1:]).strip()  # Treat everything after "cheat" as the key
            pyd.press('`')                     # Press ` key
            pyperclip.copy(key)                 # Copy the key to clipboard
            pyd.keyDown('ctrl')
            pyd.press('v')                      # Paste from clipboard
            pyd.keyUp('ctrl')
            pyd.press('enter')

        elif action == "click":
            # Extract x and y from the command
            try:
                x = int(parts[1].strip())
                y = int(parts[2].strip())
                self.makeclick(x,y)
                #pyd.click(x, y)  # Perform the click at specified coordinates
                #pag.click(x, y)  # Perform the click at specified coordinates
                print(f'Clicked at ({x}, {y})')
            except (IndexError, ValueError):
                print("Invalid coordinates for click command. Use: click x y")


        # Image-to-text (English)
        elif action == "i2s":
            self.extracted_text = self.image_to_text(img_path=self.screenshot_path, lang='eng')
            #global recent_txt
            recent_txt = self.extracted_text
            print(f'{recent_txt=}')
            if self.extracted_text:
                print(f"Extracted text (English): {self.extracted_text}")
                if len(parts) > 1 and parts[1].lower() == "show":
                    QTimer.singleShot(0, lambda: QMessageBox.information(self, "Extracted Text (English)", self.extracted_text))
            else:
                print("No text found in the image.")
                if len(parts) > 1 and parts[1].lower() == "show":
                    QTimer.singleShot(0, lambda: QMessageBox.information(self, "Extracted Text (English)", "No text found in the image."))

        elif action == "i2skr":
            self.extracted_text = self.image_to_text(img_path=self.screenshot_path, lang='kor')
            #global recent_txt
            recent_txt = self.extracted_text
            if self.extracted_text:
                print(f"Extracted text (English): {self.extracted_text}")
                if len(parts) > 1 and parts[1].lower() == "show":
                    QTimer.singleShot(0, lambda: QMessageBox.information(self, "Extracted Text (English)", self.extracted_text))
            else:
                print("No text found in the image.")
                if len(parts) > 1 and parts[1].lower() == "show":
                    QTimer.singleShot(0, lambda: QMessageBox.information(self, "Extracted Text (English)", "No text found in the image."))

        # Validate extracted text with optional language selection
        elif action == "validate":
            try:
                lang = parts[1].strip()  # Language ('i2s' or 'i2skr')
                self.expected_text = " ".join(parts[2:]).strip()  # Expected text

                # Extract text based on language
                if lang == "i2s":
                    self.extracted_text = self.image_to_text(img_path=self.screenshot_path, lang='eng')
                elif lang == "i2skr":
                    self.extracted_text = self.image_to_text(img_path=self.screenshot_path, lang='kor')
                else:
                    print(f"Unsupported language command: {lang}")
                    return

                # Compare extracted text with expected text
                if self.extracted_text.strip() == self.expected_text:
                    self.last_result = "Passed"
                    print(f"Test Passed: Extracted text matches expected text: \nExpected: '{self.expected_text}'\nExtracted: '{self.extracted_text.strip()}'")
                else:
                    self.last_result = "Failed"
                    print(f"Test Failed: Extracted text does not match expected text.\nExpected: '{self.expected_text}'\nExtracted: '{self.extracted_text.strip()}'")
            except Exception as e:
                print(f"Error during validation: {e}")

        # elif action == "export":
        #     try:
        #         custom_title = " ".join(parts[1:]).strip().strip('"')  # Extract custom title
        #         filename = "checklist.xlsx"  # Default Excel file name
        #         cl_path = os.path.join(cl_dir, filename)

        #         # Check if the Excel file exists; if not, create a new one
        #         if os.path.exists(cl_path):
        #             workbook = load_workbook(cl_path)
        #             sheet = workbook.active
        #         else:
        #             workbook = Workbook()
        #             sheet = workbook.active
        #             # Add headers to the new file
        #             sheet.append(["Title", "Result", "Screenshot"])

        #         # Append new row to the Excel file
        #         new_row = [custom_title, self.last_result, self.screenshot_path if hasattr(self, 'screenshot_path') else ""]
        #         sheet.append(new_row)

        #         # If a screenshot path exists, insert the image into the file
        #         if hasattr(self, 'screenshot_path') and self.screenshot_path:
        #             # Open the screenshot to resize it
        #             img_path = self.screenshot_path
        #             pil_img = PILImage.open(img_path)

        #             # Resize the image to a reasonable width and height (e.g., 150x100 pixels)
        #             max_width = 150  # Maximum width for the image
        #             aspect_ratio = pil_img.height / pil_img.width
        #             resized_height = int(max_width * aspect_ratio)
        #             pil_img = pil_img.resize((max_width, resized_height), PILImage.Resampling.LANCZOS)

        #             # Save the resized image to a temporary file
        #             temp_img_path = os.path.join(screenshot_dir, "temp_resized_image.png")
        #             pil_img.save(temp_img_path)

        #             # Add the resized image to the Excel file
        #             excel_img = ExcelImage(temp_img_path)
        #             img_cell = f"C{sheet.max_row}"  # Place the image in the Screenshot column of the new row
        #             sheet.add_image(excel_img, img_cell)

        #             # Adjust the row height to match the resized image height
        #             row_height = resized_height * 0.75   # Adjust scaling factor for Excel row height
        #             sheet.row_dimensions[sheet.max_row].height = row_height

        #         # Save the Excel file
        #         workbook.save(cl_path)

        #         print(f"Validation result exported to {cl_path} with title: {custom_title}")
        #     except Exception as e:
        #         print(f"Error exporting result: {e}")


        elif action == "export":
            '''
            validate i2s LOGIN DIRECTLY
            export "접속 버튼 출력"
            '''
            try:
                custom_title = " ".join(parts[1:]).strip().strip('"')  # Extract custom title
                self.checklist_file = "checklist.xlsx"  # Default Excel file name
                self.cl_path = os.path.join(cl_dir, self.checklist_file)

                # Check if the Excel file exists; if not, create a new one
                if os.path.exists(self.cl_path):
                    workbook = load_workbook(self.cl_path)
                    sheet = workbook.active
                else:
                    workbook = Workbook()
                    sheet = workbook.active
                    # Add headers to the new file
                    sheet.append(["Title", "Expected Value", "Result", "Experimental Value", "Screenshot Path", "Screenshot"])

                # Append new row to the Excel file
                new_row = [custom_title, self.expected_text, self.last_result, self.extracted_text, self.screenshot_path if hasattr(self, 'screenshot_path') else ""]
                sheet.append(new_row)

                # If a screenshot path exists, insert the image into the file
                if hasattr(self, 'screenshot_path') and self.screenshot_path:
                    img = ExcelImage(self.screenshot_path)
                    img_cell = f"F{sheet.max_row}"  # Place the image in the Screenshot column of the new row
                    sheet.add_image(img, img_cell)

                # === 열 너비 자동 조정 ===
                from openpyxl.utils import get_column_letter
                for col in sheet.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            cell_length = len(str(cell.value)) if cell.value else 0
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    sheet.column_dimensions[col_letter].width = max_length + 2  # 여유공간

                # Save the Excel file
                workbook.save(self.cl_path)

                print(f"Validation result exported to {self.cl_path} with title: {custom_title}")
            except Exception as e:
                print(f"Error exporting result: {e}")

        elif action == "waituntil":
        # 예) waituntil screenshot 2162 1076 44 26 i2s ADS
            if len(parts) < 8:
                print("Usage: waituntil screenshot x y w h i2s|i2skr target_text")
                return

            sc_x, sc_y, sc_w, sc_h = map(int, parts[2:6])
            ocr_cmd = parts[6].strip()
            target_text = parts[7].strip()

            print(f"Waiting until '{target_text}' appears (max 10 tries)...")

            for i in range(10):  # 최대 10회 시도
                if self.stop_flag:
                    print("Stopped during waituntil.")
                    return

                try:
                    # 1) screenshot 실행
                    screenshot_path = self.take_screenshot_with_coords(sc_x, sc_y, sc_w, sc_h)

                    # 2) OCR 실행
                    if ocr_cmd == "i2s":
                        self.extracted_text = self.image_to_text(img_path=screenshot_path, lang='eng')
                    elif ocr_cmd == "i2skr":
                        self.extracted_text = self.image_to_text(img_path=screenshot_path, lang='kor')
                    else:
                        print(f"Unsupported OCR command: {ocr_cmd}")
                        return

                    # OCR 결과 전역에 저장
                    #global recent_txt
                    recent_txt = self.extracted_text

                    if target_text in recent_txt:
                        print(f"Found '{target_text}' in recent_text.")
                        break

                    print(f"[{i+1}/10] '{target_text}' not found. Retrying...")

                except Exception as e:
                    print(f"Error during waituntil loop (try {i+1}): {e}")

                time.sleep(1)  # 예외 발생하더라도 1초 대기 후 다시 시도
            else:
                print(f"Timeout: '{target_text}' not found after 10 tries.")



    def capture_mouse_position(self, num):
        # Get current mouse position
        x, y = pag.position()
        
        if num == 0 :
            # Populate input boxes
            self.x_input0.setText(str(x))
            self.y_input0.setText(str(y))
        else:
            self.x_input1.setText(str(x))
            self.y_input1.setText(str(y))
            if self.x_input0.text() != "" and self.y_input0.text() != "":
                self.x_pos.setText(self.x_input0.text())
                self.y_pos.setText(self.y_input0.text())
                self.w_pos.setText(str(int(self.x_input1.text())- int(self.x_input0.text())))
                self.h_pos.setText(str(int(self.y_input1.text())- int(self.y_input0.text())))

    def select_bat_file(self):
        # 파일 탐색기를 통해 .bat 파일 선택
        file_path, _ = QFileDialog.getOpenFileName(self, "Select .bat file", "", "Batch Files (*.bat)")
        if file_path:
            self.file_path_input.setText(file_path)  # 파일 경로를 인풋박스에 표시

    def run_bat_file(self):
        # 실행파일 경로 가져오기
        file_path = self.file_path_input.text().strip()
        if file_path and os.path.exists(file_path) and file_path.endswith(".bat"):
            # 파일의 디렉토리 경로로 이동하여 실행
            try:
                file_dir = os.path.dirname(file_path)
                subprocess.Popen(file_path, cwd=file_dir, shell=True)  # 지정된 디렉토리에서 .bat 파일 실행
            except Exception as e:
                self.show_error_message(f"Error running file: {e}")
        else:
            self.show_error_message("Invalid file path. Please select a valid .bat file.")


    def select_tesseract_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select tesseract.exe", "", "Executable Files (*.exe)")
        if file_path:
            # Validate and set
            if set_pytesseract_cmd(file_path):
                self.tesseract_input.setText(file_path)
                # Save selection to config
                self.save_config()
                QMessageBox.information(self, "Tesseract", f"Tesseract set to: {file_path}")
            else:
                QMessageBox.warning(self, "Tesseract", "Selected file is not a valid tesseract.exe")

    def auto_detect_tesseract(self):
        # Common install locations on Windows
        possible_paths = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        ]
        for p in possible_paths:
            if os.path.exists(p):
                set_pytesseract_cmd(p)
                # If UI exists, set the input text
                try:
                    self.tesseract_input.setText(p)
                except:
                    pass
                print(f"Auto-detected tesseract at: {p}")
                return True
        return False

    def test_ocr(self):
        """Run OCR on the most recent screenshot or take a new screenshot and show the result."""
        # Prefer the most recent screenshot if available
        img_path = None
        try:
            screenshots = sorted(
                [f for f in os.listdir(screenshot_dir) if f.endswith('.jpg')],
                key=lambda f: os.path.getmtime(os.path.join(screenshot_dir, f)),
                reverse=True
            )
            if screenshots:
                img_path = os.path.join(screenshot_dir, screenshots[0])
        except Exception:
            img_path = None

        if not img_path or not os.path.exists(img_path):
            # Take a fresh screenshot
            img_path = self.take_screenshot()

        # Run OCR
        try:
            extracted = self.image_to_text(img_path=img_path, lang='eng')
            if not extracted:
                extracted = "(No text found)"
            QMessageBox.information(self, "OCR Result", extracted)
        except Exception as e:
            QMessageBox.critical(self, "OCR Error", f"Error running OCR: {e}")


    def show_error_message(self, message):
        # 경고 팝업
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(message)
        msg.setWindowTitle("Error")
        msg.exec_()

    # def makeclick(self,x,y):
    #     pyd.moveTo(x, y)
    #     pyd.mouseDown()
    #     time.sleep(0.05)  # Optional: Add a slight delay to ensure it registers
    #     pyd.mouseUp()

    def makeclick(self, x, y):
    # 현재 선택된 윈도우의 크기 가져오기
        x1,y1,w1,h1 = self.update_coordinates()
        screen_width, screen_height = pag.size()  # 현재 모니터의 해상도를 가져옴

        # 2560x1440 기준으로 비율 계산
        width_ratio = w1 / screen_width
        height_ratio = (h1 - 30) / screen_height

        print(f'width_ratio={width_ratio}, height_ratio={height_ratio}')
        # 입력된 좌표를 현재 윈도우 크기에 맞게 조정
        if w1 != 2560:
            adjusted_x = x1 + int(x * width_ratio)
            adjusted_y = y1 + int(y* height_ratio) + 30#int(30 * height_ratio)#창모드시 상단 바 길이 

        else : 
            adjusted_x = x
            adjusted_y = y
        
        # 조정된 좌표로 클릭 실행
        #pyd.click(adjusted_x, adjusted_y)
        
        pyd.moveTo(adjusted_x, adjusted_y)
        pyd.mouseDown()
        time.sleep(0.05)  # Optional: Add a slight delay to ensure it registers
        pyd.mouseUp()
        print(f'Clicked at adjusted coordinates: ({adjusted_x}, {adjusted_y})')
    #else:
       # print("No window selected. Cannot perform click.")

    def align_windows(self):
        if not self.multi_checkbox.isChecked():
            return

        # windows = [gw.getWindowsWithTitle(self.window_dropdown.itemText(i))[0] 
        #         for i in range(min(self.window_dropdown.count(), 4))]

        all_windows = gw.getAllWindows()
        selected_windows = []
        
        for i in range(min(self.window_dropdown.count(), 4)):
            title = self.window_dropdown.itemText(i)
            matching_windows = [w for w in all_windows if w.title == title]
            if matching_windows:
                selected_windows.append(matching_windows[0])
                all_windows.remove(matching_windows[0])  # 이미 선택된 창 제거
        
        screen_width, screen_height = pag.size()  # 현재 모니터의 해상도를 가져옴

        width = screen_width // 2 #- 200
        height = int(width * 9 / 16) + 30

        positions = [
            (0, 0),           # 좌상단
            (width, 0),       # 우상단
            (0, height),      # 좌하단
            (width, height)   # 우하단
        ]

        for i, window in enumerate(selected_windows):
            x, y = positions[i]
            window.resizeTo(width, height)
            window.moveTo(x, y)
            try:
                window.activate()
            except:
                pass  # activate 실패 시 무시하고 계속 진행
                    
    def toggle_layouts(self, target):
        if target.isVisible():
            target.hide()
        else:
            target.show()


    def savePreset(self):
        new_preset = self.add_preset_line.text().strip()
        if not new_preset:
            QMessageBox.warning(self, "Warning", "Preset name cannot be empty.")
            return
        if not new_preset.endswith('.txt'):
            new_preset += '.txt'
        self.save_textarea_content(new_preset)
        self.refreshPresets()
        print(f'saved preset successfully: {new_preset}')

    def deletePreset(self):
        current_preset = self.preset.currentText()
        if not current_preset:
            return

        confirm_dialog = QMessageBox()
        confirm_dialog.setIcon(QMessageBox.Warning)
        confirm_dialog.setText(f"Are you sure you want to delete the preset '{current_preset}'?")
        confirm_dialog.setWindowTitle("Confirm Deletion")
        confirm_dialog.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)

        result = confirm_dialog.exec_()

        if result == QMessageBox.Ok:
            file_path = os.path.join(dir_preset, f"{current_preset}")
            try:
                os.remove(file_path)
                self.refreshPresets()
                QMessageBox.information(self, "Success", f"Preset '{current_preset}' has been deleted.")
            except OSError as e:
                QMessageBox.critical(self, "Error", f"Failed to delete preset: {str(e)}")

    def refreshPresets(self):
        self.preset.clear()
        self.preset_prefix.clear()  # Clear existing items
        #dir_preset = 'preset'  # Directory path
        preset_files = [f for f in os.listdir(dir_preset) if f.endswith('.txt')]
        
        # Create a set to store unique prefixes
        prefixes = set()

        # Dictionary to store files associated with each prefix
        prefix_to_files = {}

        for filename in preset_files:
            # Split the filename into prefix and the rest using '_'
            parts = filename.split('_')
            if len(parts) > 1:
                prefix = parts[0]  # Use the part before the first '_'
            else:
                prefix = parts[0].replace('.txt', '')  # If no '_', use the whole name

            # Add the prefix to the set
            prefixes.add(prefix)

            # Add files to the dictionary
            if prefix not in prefix_to_files:
                prefix_to_files[prefix] = []
            prefix_to_files[prefix].append(filename)


        # Add a slot to handle changes in preset_prefix selection
        def on_preset_prefix_changed():
            current_prefix = self.preset_prefix.currentText()
            self.preset.clear()
            # Filter preset files based on the selected prefix
            if current_prefix in prefix_to_files:
                self.preset.addItems(prefix_to_files[current_prefix])

        # Add sorted prefixes to preset_prefix combobox
        self.preset_prefix.addItems(sorted(prefixes))
        on_preset_prefix_changed()
        # Connect the function to preset_prefix changes
        self.preset_prefix.currentIndexChanged.connect(on_preset_prefix_changed)

        # Trigger the function to update preset for the default selection
        if self.preset_prefix.count() > 0:
            self.preset_prefix.setCurrentIndex(0)

    def applyPreset(self):
        selected_preset = self.preset.currentText()
        self.add_preset_line.setText(selected_preset)

        if selected_preset:
            self.load_textarea_content(selected_preset)

    def save_textarea_content(self, filename='commands.txt'):
        file_path = os.path.join(dir_preset, filename)
        content = self.textarea.toPlainText()
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(content)
        print(f"Content saved to {file_path}")

    def load_textarea_content(self, filename='commands.txt'):
        file_path = os.path.join(dir_preset, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            self.textarea.setPlainText(content)
            print(f"Content loaded from {file_path}")
        except FileNotFoundError:
            print(f"File not found: {file_path}")


    def insert_input0_text(self):
        x = self.x_input0.text().strip()
        y = self.y_input0.text().strip()
        if x and y:
            self.insert_text_to_cursor(f"click {x} {y}")

    def insert_input1_text(self):
        x = self.x_input1.text().strip()
        y = self.y_input1.text().strip()
        if x and y:
            self.insert_text_to_cursor(f"click {x} {y}")

    def insert_pos_text(self):
        x = self.x_pos.text().strip()
        y = self.y_pos.text().strip()
        w = self.w_pos.text().strip()
        h = self.h_pos.text().strip()
        if x and y and w and h:
            self.insert_text_to_cursor(f"screenshot {x} {y} {w} {h}")

    def insert_text_to_cursor(self, new_text):
        cursor = self.textarea.textCursor()
        current_pos = cursor.position()

        # 현재 줄의 시작과 끝 위치 구하기
        cursor.movePosition(cursor.StartOfBlock, cursor.KeepAnchor)
        start_of_line = cursor.selectionStart()

        if current_pos != start_of_line:
            # 현재 줄의 끝으로 이동
            cursor.movePosition(cursor.EndOfBlock)
            cursor.insertText('\n')  # 줄 끝에 개행 추가

        # 텍스트 삽입
        cursor.insertText(new_text + '\n')
        self.textarea.setTextCursor(cursor)
        self.textarea.setFocus()


# Main function
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = PbbAutoApp()
    ex.show()
    sys.exit(app.exec_())
